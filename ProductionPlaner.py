import datetime
import sys
import os
import ast
import shutil
from openpyxl import Workbook, load_workbook
import configparser

from PyQt6.QtCore import Qt, pyqtSignal, QRegularExpression
from PyQt6.QtWidgets import (
    QApplication,
    QLabel,
    QMainWindow,
    QPushButton,
    QWidget,    
    QLineEdit,
    QHBoxLayout,
    QVBoxLayout,       
    QComboBox,
    QAbstractItemView,    
    QFileDialog,    
    QDateEdit,    
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QTabWidget,
    QToolBar, 
    QCheckBox, 
    QTextEdit,
    QMessageBox,
    QInputDialog          
)
from PyQt6.QtGui import QIcon, QAction, QIntValidator, QRegularExpressionValidator, QTextDocument, QTextCursor, QTextTableFormat, QPageLayout, QTextCharFormat, QColor
from PyQt6.QtPrintSupport import QPrintPreviewDialog

try:
    from ctypes import windll  # Only exists on Windows.
    myappid = 'realblack7.productionmanager.v1.3'
    windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
except ImportError:
    pass

class AddBatchWindow(QWidget):
    added = pyqtSignal(list)
    finished = pyqtSignal()

    def __init__(self, attr1, attr2, attrPack, attrLab, timenormal, timedensity, timemechanics, timereach, colorList, dictArticleColors, dispoNoList):
        super().__init__()      

        self.attr1 = attr1
        self.attr2 = attr2        
        self.attrPackaging = attrPack
        self.attrLab = attrLab  
        self.timeNormal = timenormal
        self.timeDensity = timedensity
        self.timeMechanics = timemechanics
        self.timeReach = timereach 
        self.colorList = colorList
        self.dictArticleColors = dictArticleColors
        self.dispoNoList = dispoNoList
        self.imagePath = os.path.dirname(__file__)

        self.closeMenu = True
        self.setWindowTitle('Charge hinzufügen')
        self._createGUI()

    def _createGUI(self):        
        layout2 = QVBoxLayout()
        layout3 = QVBoxLayout()
        layout4 = QHBoxLayout() 
        layout5 = QVBoxLayout()   
        layout6 = QHBoxLayout()   
        
        self.listCostumer = QComboBox()  
        self.listCostumer.addItem('')                 
        self.listCostumer.addItems(self.attr1)        
        self.listCostumer.setEditable(True) 
        self.listCostumer.InsertPolicy.InsertAlphabetically        
        self.listCostumer.setFixedWidth(150)

        rx = QRegularExpression("SP\\d{1,7}[a-zA-Z]")
        self.listDispo = QLineEdit() 
        self.listDispo.setFixedWidth(100) 
        self.listDispo.setMaxLength(10)
        self.listDispo.setText('SP')
        self.listDispo.setValidator(QRegularExpressionValidator(rx, self))

        rx2 = QRegularExpression("3\\d{1}\\.\\d{1,4}\\-\\d{1}") 
        self.listArticle = QComboBox()  
        self.listArticle.addItem('32.')                 
        self.listArticle.addItems(self.attr2)        
        self.listArticle.setEditable(True) 
        self.listArticle.InsertPolicy.InsertAlphabetically        
        self.listArticle.setValidator(QRegularExpressionValidator(rx2, self))   
        for article in range(len(self.attr2)):
                        
            if article+1 <= len(self.attr2):
                whichIcon = self.colorList[self.dictArticleColors[self.attr2[article]]][0]+'.png'               
                self.listArticle.setItemIcon(article+1, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))     
        

        self.listDeliveryDate = QDateEdit()
        self.listDeliveryDate.setFixedWidth(100)
        self.listDeliveryDate.setDate(datetime.datetime.now() + datetime.timedelta(days=self.timeNormal))
        self.listDeliveryDate.setMouseTracking(False)
        
        self.listBatchSize = QLineEdit()
        self.listBatchSize.setFixedWidth(100)        
        self.listBatchSize.setValidator(QIntValidator(00, 99))

        self.labelCharge = QLabel('Kunde')
        self.labelDispo = QLabel('Dispo-Nr.')
        self.labelArticle = QLabel('Artikle-Nr.')
        self.labelDeliveryDate = QLabel('Lieferdatum')
        self.labelBatchSize = QLabel('Tonnage')   
        self.labelPackaging = QLabel('Zusatz')
        self.labelLab = QLabel('Labor')       

        self.listPackaging = QComboBox()
        self.listPackaging.addItems(self.attrPackaging)
                       
        self.listLab = QComboBox()
        self.listLab.addItems(self.attrLab)
        self.listLab.currentIndexChanged.connect(self.labChanged)

        self.closeButton = QPushButton('Schließen')
        self.closeButton.setFixedWidth(80)
        self.closeButton.clicked.connect(self.close) 

        self.addButton = QPushButton('Hinzufügen')
        self.addButton.setFixedWidth(80)
        self.addButton.clicked.connect(self.addBatchToList)  
        self.addButton.setShortcut("Return")   

        layout2.addWidget(self.listCostumer)
        layout2.addWidget(self.listDispo)
        layout2.addWidget(self.listArticle)
        layout2.addWidget(self.listDeliveryDate)
        layout2.addWidget(self.listBatchSize)
        layout2.addWidget(self.listPackaging)
        layout2.addWidget(self.listLab)            

        layout3.addWidget(self.labelCharge) 
        layout3.addWidget(self.labelDispo)
        layout3.addWidget(self.labelArticle)
        layout3.addWidget(self.labelDeliveryDate)
        layout3.addWidget(self.labelBatchSize) 
        layout3.addWidget(self.labelPackaging)
        layout3.addWidget(self.labelLab)     
      
        layout4.addLayout(layout3)
        layout4.addLayout(layout2) 

        layout6.addWidget(self.addButton)
        layout6.addWidget(self.closeButton)
        layout6.addStretch()

        layout5.addLayout(layout4)
        layout5.addLayout(layout6)     

        self.setLayout(layout5)        

    def addBatchToList(self): 
        

        if self.listDispo.text() in self.dispoNoList.keys():
            dispoInUse = QMessageBox()
            dispoInUse.setWindowTitle("Achtung!")
            dispoInUse.setText("Die Charge konnte nicht hinzugefügt werden.\nDie angegebene Dispo-Nr. ist bereits vergeben!")
            dispoInUse.exec()
            return
        if self.listArticle.currentText() == '32.':
            dispoInUse = QMessageBox()
            dispoInUse.setWindowTitle("Achtung!")
            dispoInUse.setText("Die Charge konnte nicht hinzugefügt werden.\nBitte wähle einen Artikel aus!")
            dispoInUse.exec()
            return
        else:        
            batchArray = ['', '', '', '', self.listArticle.currentText(), '', self.listDispo.text(), self.listCostumer.currentText(), self.listPackaging.currentIndex(), self.listLab.currentIndex(), self.listDeliveryDate.date().toString('dd.MM.yyyy'), self.listBatchSize.text(), '' ] 

            self.added.emit(batchArray) 

    def labChanged(self):        
        
        whichLab = self.sender().currentIndex()              

        if whichLab == 0:     
                self.listDeliveryDate.setDate(datetime.datetime.now() + datetime.timedelta(days=self.timeNormal))         
        elif whichLab == 1:            
            self.listDeliveryDate.setDate(datetime.datetime.now() + datetime.timedelta(days=self.timeDensity))   
        elif whichLab == 2:            
            self.listDeliveryDate.setDate(datetime.datetime.now() + datetime.timedelta(days=self.timeMechanics))                               
        elif whichLab == 3:            
            self.listDeliveryDate.setDate(datetime.datetime.now() + datetime.timedelta(days=self.timeReach))
                           
    def closeEvent(self, event):
        
        if self.closeMenu == True:
            event.accept()
            self.finished.emit()
        else:
            event.ignore()

class SettingsWindow(QWidget):
    added = pyqtSignal(list)
    finished = pyqtSignal()

    def __init__(self, sortBy, timenormal, timedensity, timemechanics, timereach, usageFactor, maintenanceDay):
        super().__init__()   

        self.closeMenu = True
        self.sortBy = sortBy
        self.timeNormal = timenormal
        self.timeDensity = timedensity
        self.timeMechanics = timemechanics
        self.timeReach = timereach
        self.usageFactor = usageFactor
        self.maintenanceDay = maintenanceDay
    
        self.setWindowTitle('Einstellungen')
        self._createGUI()

    def _createGUI(self):        
        layout2 = QVBoxLayout()
        layout3 = QVBoxLayout()
        layout4 = QHBoxLayout() 
        layout5 = QVBoxLayout() 
        layout6 = QHBoxLayout()

        
        self.sortByBox = QComboBox()                 
        self.sortByBox.addItems(['Produktionsbeginn', 'Produktionsende', 'Abholung'])        
        self.sortByBox.setEditable(True) 
        self.sortByBox.setCurrentIndex(self.sortBy)
        self.sortByBox.InsertPolicy.InsertAlphabetically        
        self.sortByBox.setFixedWidth(140)  
        self.sortByBox.currentIndexChanged.connect(self.enableSaveButton)

        rx = QRegularExpression("\\d{1,2}")
        self.timenormalLine = QLineEdit() 
        self.timenormalLine.setFixedWidth(50) 
        self.timenormalLine.setMaxLength(2)
        self.timenormalLine.setText(str(self.timeNormal))
        self.timenormalLine.setValidator(QRegularExpressionValidator(rx, self)) 
        self.timenormalLine.textChanged.connect(self.enableSaveButton)     
        
        self.timedensityLine = QLineEdit() 
        self.timedensityLine.setFixedWidth(50) 
        self.timedensityLine.setMaxLength(2)
        self.timedensityLine.setText(str(self.timeDensity))
        self.timedensityLine.setValidator(QRegularExpressionValidator(rx, self)) 
        self.timedensityLine.textChanged.connect(self.enableSaveButton)
        
        self.timemechanicsLine = QLineEdit() 
        self.timemechanicsLine.setFixedWidth(50) 
        self.timemechanicsLine.setMaxLength(2)
        self.timemechanicsLine.setText(str(self.timeMechanics))
        self.timemechanicsLine.setValidator(QRegularExpressionValidator(rx, self)) 
        self.timemechanicsLine.textChanged.connect(self.enableSaveButton)
        
        self.timereachLine = QLineEdit() 
        self.timereachLine.setFixedWidth(50) 
        self.timereachLine.setMaxLength(2)
        self.timereachLine.setText(str(self.timeReach))
        self.timereachLine.setValidator(QRegularExpressionValidator(rx, self)) 
        self.timereachLine.textChanged.connect(self.enableSaveButton) 

        self.usageFactorLine = QLineEdit() 
        self.usageFactorLine.setFixedWidth(50) 
        self.usageFactorLine.setMaxLength(2)
        self.usageFactorLine.setText(str(self.usageFactor))
        self.usageFactorLine.setValidator(QRegularExpressionValidator(rx, self))
        self.usageFactorLine.textChanged.connect(self.enableSaveButton) 

        self.maintenanceBox = QComboBox()                 
        self.maintenanceBox.addItems(['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag'])        
        self.maintenanceBox.setEditable(True) 
        self.maintenanceBox.setCurrentIndex(self.maintenanceDay)       
        self.maintenanceBox.setFixedWidth(140)  
        self.maintenanceBox.currentIndexChanged.connect(self.enableSaveButton) 

        self.labelSort = QLabel('Sortieren nach ')   
        self.labelNormal = QLabel('Vorlauf Produktion ')
        self.labelDenisty = QLabel('Vorlauf Dichte-Messung ')
        self.labelMechanics = QLabel('Vorlauf Mechanik-Messung ')
        self.labelReach = QLabel('Vorlauf REACh-Messung ')
        self.labelUsageFactor = QLabel('Additiv-Verbrauch [%] ')
        self.labelMaintenaceDay = QLabel('Wartungstag ')

        self.closeButton = QPushButton('Schließen')
        self.closeButton.setFixedWidth(80)
        self.closeButton.clicked.connect(self.close)              

        self.addButton = QPushButton('Speichern')
        self.addButton.setFixedWidth(80)
        self.addButton.clicked.connect(self.saveSettings) 
        self.addButton.setShortcut("Return")    
        self.addButton.setEnabled(False)

        layout2.addWidget(self.sortByBox) 
        layout2.addWidget(self.timenormalLine)
        layout2.addWidget(self.timedensityLine)
        layout2.addWidget(self.timemechanicsLine)
        layout2.addWidget(self.timereachLine) 
        layout2.addWidget(self.usageFactorLine)  
        layout2.addWidget(self.maintenanceBox) 

        layout3.addWidget(self.labelSort)  
        layout3.addWidget(self.labelNormal)
        layout3.addWidget(self.labelDenisty)
        layout3.addWidget(self.labelMechanics)
        layout3.addWidget(self.labelReach)  
        layout3.addWidget(self.labelUsageFactor)
        layout3.addWidget(self.labelMaintenaceDay)
      
        layout4.addLayout(layout3)
        layout4.addLayout(layout2) 

        layout6.addWidget(self.addButton)
        layout6.addWidget(self.closeButton)
        layout6.addStretch()

        layout5.addLayout(layout4)
        layout5.addLayout(layout6)
              

        self.setLayout(layout5)  

    def enableSaveButton(self):        
        self.addButton.setEnabled(True)     

    def saveSettings(self):    
                 
        settingsToSave = [self.sortByBox.currentIndex(), self.timenormalLine.text(), self.timedensityLine.text( ), self.timemechanicsLine.text(), self.timereachLine.text(), self.usageFactorLine.text(), self.maintenanceBox.currentIndex()] 
        self.addButton.setEnabled(False)
        self.added.emit(settingsToSave)     
                   
    def closeEvent(self, event):
        
        if self.closeMenu == True:
            event.accept()
            self.finished.emit()
        else:
            event.ignore()

class EditDataWindow(QWidget):
    added = pyqtSignal(object)
    finished = pyqtSignal()

    def __init__(self, mode, articleList, additiveList, customerList):
        super().__init__() 
        self.w = None
        self.closeMenu = True
        self.mode = mode
        self.articleList = articleList
        self.additiveList = additiveList
        self.customerList = customerList
        self.imagePath = os.path.dirname(__file__)
        match self.mode:
            case 0:
                self.setWindowTitle('Kunden-Liste')               
            case 1:
                self.setWindowTitle('Artikel-Liste') 
            case 2:
                self.setWindowTitle('Additive-Liste')
        self._createGUI()

    def _createGUI(self):
        match self.mode:
            case 0:
                buttonText = 'Kunden'              
            case 1:
                buttonText = 'Artikel'
            case 2:                
                buttonText = 'Additive'


        self.menubar = QToolBar()              

        self.addItem = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'plus-solid.svg')), buttonText + ' hinzufügen  (Strg + A)', self)        
        self.addItem.triggered.connect(lambda: self.editEntry(0))  
        self.addItem.setShortcut("Ctrl+A")
        self.menubar.addAction(self.addItem)

        if not self.mode == 2:
            self.editItem = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'pen-solid.svg')), buttonText + ' ändern  (Strg + E)', self)        
            self.editItem.triggered.connect(lambda: self.editEntry(1))           
            self.editItem.setShortcut("Ctrl+E")
            self.menubar.addAction(self.editItem)


        layout1 = QVBoxLayout()        
        layout1.setMenuBar(self.menubar)

        menuContent = QWidget()
        menuContent.setLayout(layout1)

        layout2 = QVBoxLayout()         
        layout5 = QVBoxLayout() 
        layout6 = QHBoxLayout()
        
        self.listData = QTableWidget()
        self.listData.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.listData.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.listData.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)   
        self.listData.doubleClicked.connect(lambda: self.editEntry(1))     

        match self.mode:
            case 0:                
                tableHorizontalHeaders = ['Kunde'] 
                self.listData.verticalHeader().setVisible(False)
                self.listData.setFixedWidth(220) 
                self.listData.setFixedHeight(600)  
                self.listData.setColumnCount(1)
                self.listData.horizontalHeader().resizeSection(0, 200)
                self.listData.setHorizontalHeaderLabels(tableHorizontalHeaders)  
                
                for key in range(len(self.customerList)): 
                    self.listData.insertRow(key)            
                    self.listData.setItem(key, 0, QTableWidgetItem(self.customerList[key]))          

            case 1:                              
                tableHorizontalHeaders = ['Artikel-Nr.', 'Bezeichnung', 'Additive']                
                self.listData.verticalHeader().setVisible(False)
                self.listData.setFixedWidth(800) 
                self.listData.setFixedHeight(600)  
                self.listData.setColumnCount(3)  
                self.listData.setHorizontalHeaderLabels(tableHorizontalHeaders)  
                self.listData.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents) 
                self.listData.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents) 
                self.listData.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents) 

                keyRow = 0
                for key in self.articleList:                     
                     
                    self.listData.insertRow(keyRow)              

                    self.listData.setItem(keyRow, 0, QTableWidgetItem(self.articleList[key][1]))
                    self.listData.setItem(keyRow, 1, QTableWidgetItem(self.articleList[key][2])) 

                    additiveString = ''                          
                    for keys, value in self.articleList[key][3].items(): 
                        additiveConcentration = ast.literal_eval(value)
                            
                        additiveString = additiveString + str(keys) + ': ' + str(additiveConcentration[0]) + '; '

                    self.listData.setItem(keyRow, 2, QTableWidgetItem(additiveString[:-2]))
                    keyRow = keyRow + 1
                
                
            case 2:                
                tableHorizontalHeaders = ['Additiv-Nr.', 'Hersteller-Bezeichnung', 'Zweck']
                self.listData.verticalHeader().setVisible(False)
                self.listData.setFixedWidth(1030) 
                self.listData.setFixedHeight(600)  
                self.listData.setColumnCount(3)   
                self.listData.setHorizontalHeaderLabels(tableHorizontalHeaders)
                self.listData.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents) 
                self.listData.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents) 
                self.listData.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents) 

                keyRow = 0
                for key in self.additiveList: 
                    self.listData.insertRow(keyRow)               

                    self.listData.setItem(keyRow, 0, QTableWidgetItem(self.additiveList[key][0]))
                    self.listData.setItem(keyRow, 1, QTableWidgetItem(self.additiveList[key][1]))
                    self.listData.setItem(keyRow, 2, QTableWidgetItem(self.additiveList[key][2]))
                    keyRow = keyRow + 1
                

        self.closeButton = QPushButton('Schließen')
        self.closeButton.setFixedWidth(80)
        self.closeButton.clicked.connect(self.close)  

        self.addButton = QPushButton('Speichern')
        self.addButton.setFixedWidth(80)
        self.addButton.clicked.connect(self.sendSaveData) 
        self.addButton.setShortcut("Return")  
        self.addButton.setEnabled(False)      

        layout2.addWidget(self.listData)

        layout6.addWidget(self.addButton)
        layout6.addWidget(self.closeButton)
        layout6.addStretch()

        layout5.addWidget(menuContent)
        layout5.addLayout(layout2)
        layout5.addLayout(layout6)
        layout5.addStretch()
              

        self.setLayout(layout5)        

    def saveData(self, changedList):
        self.changedList = changedList 
        self.addButton.setEnabled(True) 
        match self.mode:
            case 0:  
                match self.changedList[0]:
                    case 0:
                        rowCount = self.listData.rowCount()
                        self.listData.insertRow(rowCount)
                        self.listData.setItem(rowCount, 0, QTableWidgetItem(self.changedList[3]))    
                         
                    case  1:                         
                        self.listData.setItem(self.changedList[2], 0, QTableWidgetItem(self.changedList[3]))    
                                     
            case 1:
                match self.changedList[0]:
                    case 0:
                        rowCount = self.listData.rowCount()
                        self.listData.insertRow(rowCount)
                        self.listData.setItem(rowCount, 0, QTableWidgetItem(self.changedList[3]))   
                        self.listData.setItem(rowCount, 1, QTableWidgetItem(self.changedList[4]))
                        additiveString = ''                          
                        for keys, value in self.changedList[5].items():
                            
                            additiveConcentration = ast.literal_eval(value)
                                
                            additiveString = additiveString + str(keys) + ': ' + str(additiveConcentration[0]) + '; '

                        self.listData.setItem(rowCount, 2, QTableWidgetItem(additiveString[:-2]))


                    case  1:                
                        self.listData.setItem(self.changedList[2], 0, QTableWidgetItem(self.changedList[3]))   
                        self.listData.setItem(self.changedList[2], 1, QTableWidgetItem(self.changedList[4]))
                        additiveString = ''                          
                        for keys, value in self.changedList[5].items():
                            
                            additiveConcentration = ast.literal_eval(value)
                                
                            additiveString = additiveString + str(keys) + ': ' + str(additiveConcentration[0]) + '; '

                        self.listData.setItem(self.changedList[2], 2, QTableWidgetItem(additiveString[:-2]))
                        
                                                
            case 2:                 
                match self.changedList[0]:                    
                    case 0:
                        rowCount = self.listData.rowCount()
                        self.listData.insertRow(rowCount)
                        self.listData.setItem(rowCount, 0, QTableWidgetItem(self.changedList[3]))   
                        self.listData.setItem(rowCount, 1, QTableWidgetItem(self.changedList[4]))
                        self.listData.setItem(rowCount, 2, QTableWidgetItem(self.changedList[5]))

                                              
                    case  1:                         
                        self.listData.setItem(self.changedList[2], 0, QTableWidgetItem(self.changedList[3]))   
                        self.listData.setItem(self.changedList[2], 1, QTableWidgetItem(self.changedList[4]))
                        self.listData.setItem(self.changedList[2], 2, QTableWidgetItem(self.changedList[5])) 
                                             

    def sendSaveData(self):
        self.addButton.setEnabled(False)
        match self.mode:
            case 0:
                match self.changedList[0]:
                    case 0:
                        self.customerList.append(self.changedList[3])
                    case 1:
                        self.customerList[self.changedList[2]] = self.changedList[3]

                self.added.emit(self.changedList) 
            case 1:                
                match self.changedList[0]:                    
                    case 0:
                        helperList = [1, self.changedList[3], self.changedList[4], self.changedList[5], self.changedList[6]]

                        self.articleList[len(self.articleList)] = helperList
                    case 1:
                        self.articleList[self.changedList[2]][1] = self.changedList[3] 
                        self.articleList[self.changedList[2]][2] = self.changedList[4]
                        self.articleList[self.changedList[2]][3] = self.changedList[5]   
                        self.articleList[self.changedList[2]][4] = self.changedList[6] 
                self.added.emit(self.changedList)                 
            case 2:
                match self.changedList[0]:                    
                    case 0:
                        helperList = [self.changedList[3], self.changedList[4], self.changedList[5]]

                        self.additiveList[len(self.additiveList)] = helperList  
                    case 1:
                        self.additiveList[self.changedList[2]][0] = self.changedList[3] 
                        self.additiveList[self.changedList[2]][1] = self.changedList[4]
                        self.additiveList[self.changedList[2]][2] = self.changedList[5]   
                self.added.emit(self.changedList)

    def editEntry(self, addORedit): 
        
        self.editData = []
        match addORedit:
            case 0:
                match self.mode:
                        case 0:
                            self.editData = [addORedit, self.mode, '', '']
                            self.openSecondaryWindow()
                        
                        case 1:
                            self.editData = [addORedit, self.mode, '', '', '', '', '', self.additiveList]
                            self.openSecondaryWindow()
                        
                        case 2:
                            self.editData = [addORedit, self.mode, '', '', '', '']
                            self.openSecondaryWindow()

            case 1:
                if len(self.listData.selectionModel().selectedRows()) != 0:    
                    
                    match self.mode:
                        case 0: 
                            self.editData = [addORedit, self.mode, self.listData.selectionModel().selectedRows()[0].row(), self.customerList[self.listData.selectionModel().selectedRows()[0].row()]]   
                            self.openSecondaryWindow()        
                            
                        case 1:
                            self.editData = [addORedit, self.mode, self.listData.selectionModel().selectedRows()[0].row(), self.articleList[self.listData.selectionModel().selectedRows()[0].row()][1], self.articleList[self.listData.selectionModel().selectedRows()[0].row()][2], self.articleList[self.listData.selectionModel().selectedRows()[0].row()][3], self.articleList[self.listData.selectionModel().selectedRows()[0].row()][4], self.additiveList] 
                            self.openSecondaryWindow()   
                        case 2:
                            self.editData = [addORedit, self.mode, self.listData.selectionModel().selectedRows()[0].row(), self.additiveList[self.listData.selectionModel().selectedRows()[0].row()][0], self.additiveList[self.listData.selectionModel().selectedRows()[0].row()][1], self.additiveList[self.listData.selectionModel().selectedRows()[0].row()][2]]                  
                            self.openSecondaryWindow()   

    def openSecondaryWindow(self):
        
        self.listData.setDisabled(True)
        self.addItem.setDisabled(True)        
        self.editItem.setDisabled(True)        
        self.closeButton.setDisabled(True)        
        self.menubar.setDisabled(True)


        self.closeMenu = False

        if self.w is None:            
                      
            self.w = EditDataItemWindow(self.editData)         
            self.w.show()
            self.w.finished.connect(self.closeSecondaryWindow)            
            self.w.edited.connect(self.saveData)                

        else:
            self.w.close()
            self.w = None  

    def closeSecondaryWindow(self):
        self.w = None 
        self.listData.setDisabled(False)
        self.addItem.setDisabled(False)        
        self.editItem.setDisabled(False) 
        self.closeButton.setDisabled(False)
        self.menubar.setDisabled(False)
        
        self.closeMenu = True 
                   
    def closeEvent(self, event):
        
        if self.closeMenu == True:
            event.accept()
            self.finished.emit()
        else:
            event.ignore()

class EditDataItemWindow(QWidget):
    edited = pyqtSignal(list)
    finished = pyqtSignal()

    def __init__(self, editData):
        super().__init__()   

        self.closeMenu = True        
        self.editData = editData 
        self.addORedit = self.editData[0]
        self.mode = self.editData[1]  

        match self.addORedit:
            case 0:            
                self.setWindowTitle('Hinzufügen')
            case 1:
                self.setWindowTitle('Bearbeiten')

        self._createGUI()

    def _createGUI(self):

        layout1 = QVBoxLayout()
        layout2 = QVBoxLayout()
        layout3 = QHBoxLayout()
        layout4 = QHBoxLayout()
        layout5 = QVBoxLayout()

        
        match self.mode:
                case 0:              
                    
                    self.customerName = QLineEdit()
                    self.customerName.setFixedWidth(200)
                    match self.addORedit:
                        case 0:
                            self.customerName.setText('')
                        case 1:
                            self.customerName.setText(self.editData[3])

                    layout1.addWidget(self.customerName)

                    self.labelCustomer = QLabel('Kunde')

                    layout2.addWidget(self.labelCustomer)

                    layout3.addLayout(layout2)
                    layout3.addLayout(layout1)
                                     
   
                case 1:                   

                    rx = QRegularExpression("3\\d{1}\\.\\d{1,4}\\-\\d{1}")
                    self.articleNo = QLineEdit()
                    self.articleNo.setFixedWidth(200)
                    self.articleNo.setValidator(QRegularExpressionValidator(rx, self)) 
                    match self.addORedit:
                        case 0:
                            self.articleNo.setText('32.')
                        case 1:
                            self.articleNo.setText(self.editData[3])

                    self.articleName = QLineEdit()
                    self.articleName.setFixedWidth(200)
                    match self.addORedit:
                        case 0:
                            self.articleName.setText('')
                        case 1:
                            self.articleName.setText(self.editData[4])
                    self.attr1 = []
                    for additive in self.editData[7]:
                        self.attr1.append(str(self.editData[7][additive][0] + ' | ' + self.editData[7][additive][1]))

                    self.articleWhichExtruder = QCheckBox()
                    self.articleWhichExtruder.setTristate(True)
                    self.articleWhichExtruder.stateChanged.connect(self.stateChangedBox)
                    self.articleWhichExtruder.setText('Beide Extruder')
                    
                    match self.editData[6]:
                        case '':
                            boxState = Qt.CheckState.Unchecked
                        case '0':
                            boxState = Qt.CheckState.Unchecked
                        case '1':
                            boxState = Qt.CheckState.PartiallyChecked
                        case '2':
                            boxState = Qt.CheckState.Checked
                    
                        
                            
                    self.articleWhichExtruder.setCheckState(boxState)

                    self.tableAdditives = QTableWidget()   
                    tableHorizontalHeaders = ['Aktiv', 'Additiv', 'Konzentration']                
                    self.tableAdditives.verticalHeader().setVisible(False)
                    self.tableAdditives.setFixedWidth(600) 
                    self.tableAdditives.setFixedHeight(500)  
                    self.tableAdditives.setColumnCount(3)  
                    self.tableAdditives.setHorizontalHeaderLabels(tableHorizontalHeaders)  
                    self.tableAdditives.horizontalHeader().resizeSection(0, 38)     
                    self.tableAdditives.horizontalHeader().resizeSection(1, 150)  
                    self.tableAdditives.horizontalHeader().resizeSection(2, 150)

                    for row in range(10): 
                        self.tableAdditives.insertRow(row)             

                        self.activateAdditive = QCheckBox()                                             
                        self.tableAdditives.setCellWidget(row, 0, self.activateAdditive)                

                        self.articleAdditives = QComboBox()                             
                        self.articleAdditives.addItem('')            
                        self.articleAdditives.addItems(self.attr1)        
                        self.articleAdditives.setEditable(True) 
                        self.articleAdditives.InsertPolicy.InsertAlphabetically 
                        width = self.articleAdditives.minimumSizeHint().width()
                        self.articleAdditives.view().setMinimumWidth(width)        
                        self.articleAdditives.setFixedWidth(150)                         
                                         
                        self.tableAdditives.setCellWidget(row, 1, self.articleAdditives)

                        rx = QRegularExpression("\d{1,2}\,\d{1,2}")
                        self.concentrationAdditive = QLineEdit()  
                        self.concentrationAdditive.setValidator(QRegularExpressionValidator(rx, self))  

                        self.tableAdditives.setCellWidget(row, 2, self.concentrationAdditive)   

                    match self.addORedit:
                        case 1:
                            keyNumber = 0
                            for keys, value in self.editData[5].items():                   
                                
                                
                                additiveConcentration = ast.literal_eval(value)                                        
                                self.tableAdditives.cellWidget(keyNumber, 0).setChecked(int(additiveConcentration[1]))
                                self.tableAdditives.cellWidget(keyNumber, 1).setCurrentText(keys)
                                self.tableAdditives.cellWidget(keyNumber, 2).setText(str(additiveConcentration[0]))                         
                                
                    
                                keyNumber = keyNumber + 1
                                

                    layout1.addWidget(self.articleNo)
                    layout1.addWidget(self.articleName)
                    match self.mode:
                        case 1:
                            layout1.addWidget(self.articleWhichExtruder)
                    layout1.addWidget(self.tableAdditives)

                    self.labelArticleNo = QLabel('Artikel-Nr.')
                    self.labelArticleName = QLabel('Bezeichnung')
                    self.labelArticleAdditives = QLabel('Additive')                   

                    layout2.addWidget(self.labelArticleNo)
                    layout2.addWidget(self.labelArticleName)
                    layout2.addWidget(self.labelArticleAdditives)                    
                    layout2.addStretch()

                    layout3.addLayout(layout2)
                    layout3.addLayout(layout1)

                case 2:
                    rx = QRegularExpression("\d{1,2}\.\d{1,4}\-\d{1}")
                    self.additiveNo = QLineEdit()
                    self.additiveNo.setFixedWidth(200)
                    self.additiveNo.setValidator(QRegularExpressionValidator(rx, self)) 
                    match self.addORedit:
                        case 0:
                            self.additiveNo.setText('02.')
                        case 1:
                            self.additiveNo.setText(self.editData[3])

                    self.additiveName = QLineEdit()
                    self.additiveName.setFixedWidth(200)
                    match self.addORedit:
                        case 0:
                            self.additiveName.setText('')
                        case 1:
                            self.additiveName.setText(self.editData[4])

                    self.additiveDesig = QLineEdit()
                    self.additiveDesig.setFixedWidth(200)
                    match self.addORedit:
                        case 0:
                            self.additiveDesig.setText('')
                        case 1:
                            self.additiveDesig.setText(self.editData[5])

                    layout1.addWidget(self.additiveNo)
                    layout1.addWidget(self.additiveName)
                    layout1.addWidget(self.additiveDesig)

                    self.labelAdditiveNo = QLabel('Additive-Nr.')
                    self.labelAdditiveName = QLabel('Handelsname')
                    self.labelAdditiveDesig = QLabel('Bezeichnung')

                    layout2.addWidget(self.labelAdditiveNo)
                    layout2.addWidget(self.labelAdditiveName)
                    layout2.addWidget(self.labelAdditiveDesig)

                    layout3.addLayout(layout2)
                    layout3.addLayout(layout1)

        self.closeButton = QPushButton('Schließen')
        self.closeButton.setFixedWidth(80)
        self.closeButton.clicked.connect(self.close)              

        match self.addORedit:
            case 0:
                self.addButton = QPushButton('Hinzufügen')
                        
            case 1:
                self.addButton = QPushButton('Speichern')
        self.addButton.setFixedWidth(80)
        self.addButton.clicked.connect(self.saveEditData) 
        self.addButton.setShortcut("Return") 

        layout4.addWidget(self.addButton)
        layout4.addWidget(self.closeButton)
        layout4.addStretch()

        layout5.addLayout(layout3)
        layout5.addLayout(layout4)

        
        self.setLayout(layout5)

    def saveEditData(self):
        match self.mode:
                case 0:                     
                    match self.addORedit:
                        case 0:
                            self.editData[3] = self.customerName.text()                   
                        case 1:
                            self.editData[3] = self.customerName.text()
                    
                    if self.editData[3] == '':
                        popUpEmptyField = QMessageBox()
                        popUpEmptyField.setWindowTitle("Achtung!")
                        popUpEmptyField.setText("Das Textfeld wurde nicht ausgefüllt!")
                        popUpEmptyField.exec()
                    else:                    
                                                 
                        self.edited.emit(self.editData)
                        self.close()        
                case 1:
                    match self.addORedit:
                        case 0:                            
                            self.editData[3] = self.articleNo.text()
                            self.editData[4] = self.articleName.text()
                            articleAdditiveDict = {}
                            emptyConcentrationCheck = False
                            for row in range(10):
                                if not self.tableAdditives.cellWidget(row, 1).currentText() == '':
                                    if self.tableAdditives.cellWidget(row, 2).text() == '':                                        
                                        emptyConcentrationCheck = 1
                                        break
                                    articleAdditiveDict[self.tableAdditives.cellWidget(row, 1).currentText()] = str([self.tableAdditives.cellWidget(row, 2).text(), str(int(self.tableAdditives.cellWidget(row, 0).isChecked()))])
                                else:
                                    if not self.tableAdditives.cellWidget(row, 2).text() == '':
                                        emptyConcentrationCheck = 2
                                        break 

                            self.editData[5] = articleAdditiveDict  
                            match self.articleWhichExtruder.checkState():
                                case Qt.CheckState.Unchecked:
                                    self.editData[6] = '0'
                                case Qt.CheckState.PartiallyChecked:
                                    self.editData[6] = '1'
                                case Qt.CheckState.Checked:
                                    self.editData[6] = '2'
                        case 1:                            
                            self.editData[3] = self.articleNo.text()
                            self.editData[4] = self.articleName.text()
                            articleAdditiveDict = {}
                            emptyConcentrationCheck = False
                            for row in range(10):
                                if not self.tableAdditives.cellWidget(row, 1).currentText() == '':                                    
                                    if self.tableAdditives.cellWidget(row, 2).text() == '':                                        
                                        emptyConcentrationCheck = 1
                                        break
                                    articleAdditiveDict[self.tableAdditives.cellWidget(row, 1).currentText()] = str([self.tableAdditives.cellWidget(row, 2).text(), str(int(self.tableAdditives.cellWidget(row, 0).isChecked()))])
                                else:
                                    if not self.tableAdditives.cellWidget(row, 2).text() == '':
                                        emptyConcentrationCheck = 2
                                        break                                                                   


                            self.editData[5] = articleAdditiveDict 
                            match self.articleWhichExtruder.checkState():
                                case Qt.CheckState.Unchecked:
                                    self.editData[6] = '0'
                                case Qt.CheckState.PartiallyChecked:
                                    self.editData[6] = '1'
                                case Qt.CheckState.Checked:
                                    self.editData[6] = '2'
                                                       
                    if self.editData[3] == '' or self.editData[4] == '':
                        popUpEmptyField = QMessageBox()
                        popUpEmptyField.setWindowTitle("Achtung!")
                        popUpEmptyField.setText("Ein Textfeld wurde nicht ausgefüllt!")
                        popUpEmptyField.exec()
                    elif emptyConcentrationCheck:
                        popUpEmptyField = QMessageBox()
                        popUpEmptyField.setWindowTitle("Achtung!")
                        if emptyConcentrationCheck == 1:                            
                            popUpEmptyField.setText("Bei einem Additiv wurde keine Konzentration angegeben!")
                        else:
                            popUpEmptyField.setText("In einer leeren Reihe wurde eine Konzentration angegeben!")
                        popUpEmptyField.exec()
                    else: 
                        self.edited.emit(self.editData)
                        self.close()

                case 2:
                    match self.addORedit:
                        case 0:
                            self.editData[3] = self.additiveNo.text()
                            self.editData[4] = self.additiveName.text()
                            self.editData[5] = self.additiveDesig.text()
                            
                        case 1:
                            self.editData[3] = self.additiveNo.text()
                            self.editData[4] = self.additiveName.text()
                            self.editData[5] = self.additiveDesig.text()
                    if self.editData[3] == '' or self.editData[4] == '' or self.editData[5] == '':        
                        popUpEmptyField = QMessageBox()
                        popUpEmptyField.setWindowTitle("Achtung!")
                        popUpEmptyField.setText("Ein Textfeld wurde nicht ausgefüllt!")
                        popUpEmptyField.exec()
                    else: 
                        self.edited.emit(self.editData)
                        self.close()

    def stateChangedBox(self):
        state = self.articleWhichExtruder.checkState()
        if state == Qt.CheckState.Checked:
            self.articleWhichExtruder.setText("Nur Extruder 1")
        elif state == Qt.CheckState.PartiallyChecked:
            self.articleWhichExtruder.setText("Nur Extruder 2")
        else:
            self.articleWhichExtruder.setText("Beide Extruder")

    def closeEvent(self, event):
        
        if self.closeMenu == True:
            event.accept()
            self.finished.emit()
        else:
            event.ignore()

class MainWindow(QMainWindow):           

    def __init__(self):
        super().__init__()
        self.w = None
        self.closeMenu = True  
        self.workingOnShiftPlan = False
        self.changeArticleNoCheck = False
        self.setLoadedFile = False    
        self.dataXLSX = os.path.join(os.path.dirname(__file__), 'data', 'data.xlsx')  
        self.imagePath = os.path.dirname(__file__)

        self.dispoNoList = {}
        self.checkDispoNoSilo = {} 

        self.config = configparser.ConfigParser()
        self.config.read(os.path.join(self.imagePath, 'settings.ini'))

        self.saveFilePath = self.config['PATH']['lastsaved']   
        self.importFilePath = self.config['PATH']['lastimport'] 
        self.sortBy = int(self.config['SETTINGS']['sortby'])  

        self.timeNormal = int(self.config['SETTINGS']['timenormal']) 
        self.timeDensity = int(self.config['SETTINGS']['timedensity']) 
        self.timeMechanics = int(self.config['SETTINGS']['timemechanics']) 
        self.timeReach = int(self.config['SETTINGS']['timereach']) 
        self.maintenanceDay = int(self.config['SETTINGS']['maintenanceday'])

        self.usageFactor = int(self.config['USAGE']['factor']) 

        match self.sortBy:
            case 0:
                self.sortByColumn = 2
            case 1:
                self.sortByColumn = 3
            case 2:
                self.sortByColumn = 10     
        

        self.attrShift = ['F-S','S-N','N-F','N-W-S', 'W-S-N', 'F', 'S', 'N']
        self.attrPack = ['Bigbag','Oktabin','Silo','Homogenisierung']
        self.attrLab = ['-','Dichte','Mechanik','REACh']     ### Dichte+Mechanik?  ### Muster?    

        self.setWindowTitle("Produktionsplaner")        
        self.setAcceptDrops(False)
        self._createMenu() 
        self._createTabs()     
        self._createPlanerViewExtruder1()   
        self._createPlanerViewExtruder2()  
        self._createPlanerViewHomogenisation()     
        self._createPlanerViewSilo() 
        self._createTextArea()
        self._createMaster()   
        self._loadData() 
        
    def _createMenu(self):
        self.menubar = self.addToolBar('Menü')              

        self.openFileDialog = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'folder-open-solid.svg')), 'Laden (Strg + O)', self)        
        self.openFileDialog.triggered.connect(self.loadFile)
        self.openFileDialog.setShortcut("Ctrl+O")

        self.saveFile = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'floppy-disk-solid.svg')), 'Speichern  (Strg + S)', self)        
        self.saveFile.triggered.connect(self.performSaveFile)
        self.saveFile.setShortcut("Ctrl+S")
        self.saveFile.setDisabled(True)

        self.saveFileAs = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'file-solid.svg')), 'Speichern unter...  (Strg + Shift + S)', self)        
        self.saveFileAs.triggered.connect(self.performSaveFileAs)
        self.saveFileAs.setShortcut("Ctrl+Shift+S")

        self.importExcel = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'excel-import.png')), 'Speichern unter...  (Strg + I)', self)        
        self.importExcel.triggered.connect(self.importExcelFile)
        self.importExcel.setShortcut("Ctrl+I")

        self.addBatch = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'plus-solid.svg')), 'Charge hinzufügen  (Strg + A)', self)        
        self.addBatch.triggered.connect(lambda: self.openSecondaryWindow(0))
        self.addBatch.setShortcut("Ctrl+A") 

        self.generateSiloListsButton = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'silo-gen.png')), 'Silo-Listen erstellen  (Strg + G)', self)        
        self.generateSiloListsButton.triggered.connect(self.generateSiloLists)
        self.generateSiloListsButton.setShortcut("Ctrl+G")  

        self.generateAdditiveUsageButton = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'flask-gen.png')), 'Additive-Verbrauch bestimmen  (Strg + U)', self)        
        self.generateAdditiveUsageButton.triggered.connect(self.generateAdditiveUsage)
        self.generateAdditiveUsageButton.setShortcut("Ctrl+U")

        self.changeCustomers = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'truck-gear.png')), 'Kunden ansehen/ändern/hinzufügen (Strg + T)', self)        
        self.changeCustomers.triggered.connect(lambda: self.openSecondaryWindow(2))
        self.changeCustomers.setShortcut("Ctrl+T")
        
        self.changeArticles = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'tag-gear.png')), 'Artikel ansehen/ändern/hinzufügen (Strg + F)', self)        
        self.changeArticles.triggered.connect(lambda: self.openSecondaryWindow(3))
        self.changeArticles.setShortcut("Ctrl+F")

        self.changeAdditives = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'flask-gear.png')), 'Additive ansehen/ändern/hinzufügen (Strg + R)', self)        
        self.changeAdditives.triggered.connect(lambda: self.openSecondaryWindow(4))
        self.changeAdditives.setShortcut("Ctrl+R")
        
        self.changeSettings = QAction(QIcon(os.path.join(self.imagePath, 'assets', 'gear-solid.svg')), 'Einstellungen (Strg + E)', self)        
        self.changeSettings.triggered.connect(lambda: self.openSecondaryWindow(1))
        self.changeSettings.setShortcut("Ctrl+E") 

        self.menubar.addAction(self.openFileDialog)
        self.menubar.addAction(self.saveFile)
        self.menubar.addAction(self.saveFileAs) 
        self.menubar.addAction(self.importExcel)        
        self.menubar.addSeparator()
        self.menubar.addAction(self.addBatch)  
        self.menubar.addAction(self.generateSiloListsButton)
        self.menubar.addAction(self.generateAdditiveUsageButton)
        self.menubar.addSeparator()        
        self.menubar.addAction(self.changeCustomers)
        self.menubar.addAction(self.changeArticles)
        self.menubar.addAction(self.changeAdditives)
        self.menubar.addAction(self.changeSettings)

    def _createTabs(self):
        self.tabs = QTabWidget()
        self.tabExtruder1 = QWidget()
        self.tabExtruder2 = QWidget()
        self.tabHomogenisation = QWidget()
        self.tabSilo = QWidget() 

        self.tabs.addTab(self.tabExtruder1, 'Extruder 1')   
        self.tabs.addTab(self.tabExtruder2, 'Extruder 2') 
        self.tabs.addTab(self.tabHomogenisation, 'Homogenisierung') 
        self.tabs.addTab(self.tabSilo, 'Silo')  

        self.tabLayout = QVBoxLayout()
        self.tabLayout.addWidget(self.tabs)    
             
    def _createPlanerViewExtruder1(self):                     
        
        tableHorizontalHeaders = ['KW', 'Schichten', 'Beginn', 'Ende', 'Artikel-Nr.', 'Chargen-Nr.', 'Dispo.-Nr.', 'Kunde', 'Zusatz', 'Labor', 'Abholung', 't', 'Vorlauf', 'Bemerkung']

        self.tableBatchesExtruder1 = QTableWidget() 
        self.tableBatchesExtruder1.verticalHeader().setVisible(False)
        self.tableBatchesExtruder1.setFixedWidth(1450) 
        self.tableBatchesExtruder1.setFixedHeight(875)  
        self.tableBatchesExtruder1.setColumnCount(14)  
        self.tableBatchesExtruder1.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)  
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(0, 38)     
        self.tableBatchesExtruder1.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(2, 80)  
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(3, 80)  
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(4, 100)
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(5, 100) 
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(6, 100)
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(7, 150)     
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(8, 120)             
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(9, 80) 
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(10, 80)
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(11, 38) 
        self.tableBatchesExtruder1.horizontalHeader().setSectionResizeMode(12, QHeaderView.ResizeMode.ResizeToContents) 
        self.tableBatchesExtruder1.horizontalHeader().resizeSection(13, 350)         
        self.tableBatchesExtruder1.setHorizontalHeaderLabels(tableHorizontalHeaders)  

        self.sortExtruder1byDeliveryDateButton = QPushButton('Sortieren')
        self.sortExtruder1byDeliveryDateButton.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'arrow-down-short-wide-solid.svg')))
        self.sortExtruder1byDeliveryDateButton.setFixedWidth(100)
        self.sortExtruder1byDeliveryDateButton.clicked.connect(lambda: self.sortExtruderbyDeliveryDateButton(1))                   

        self.moveToExtruder2 = QPushButton('zu Extruder 2')
        self.moveToExtruder2.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'arrows-turn-to-dots-solid')))
        self.moveToExtruder2.setFixedWidth(100)
        self.moveToExtruder2.clicked.connect(lambda: self.moveBatchToExtruder(1))

        self.deleteBatchExtruder1 = QPushButton('Löschen')
        self.deleteBatchExtruder1.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'trash-solid.svg')))
        self.deleteBatchExtruder1.setFixedWidth(100)
        self.deleteBatchExtruder1.clicked.connect(lambda: self.deleteBatchFromListExtruder(1))

        self.moveRowUp1 = QPushButton('nach oben')
        self.moveRowUp1.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'up-long-solid.svg')))
        self.moveRowUp1.setFixedWidth(100)
        self.moveRowUp1.clicked.connect(lambda: self.moveBatchRowUp(1))

        self.moveRowDown1 = QPushButton('nach unten')
        self.moveRowDown1.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'down-long-solid.svg')))
        self.moveRowDown1.setFixedWidth(100)
        self.moveRowDown1.clicked.connect(lambda: self.moveBatchRowDown(1))

        self.createShiftPlan1 = QPushButton('Schichten')
        self.createShiftPlan1.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'user-clock-solid.svg')))
        self.createShiftPlan1.setFixedWidth(100)
        self.createShiftPlan1.clicked.connect(lambda: self.createShiftPlan(1))

        self.enumerateBatches1 = QPushButton('Nummerieren')
        self.enumerateBatches1.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'arrow-down-1-9-solid.svg')))
        self.enumerateBatches1.setFixedWidth(100)
        self.enumerateBatches1.clicked.connect(lambda: self.enumerateBatches(1))

        tabLayoutExtruder1 = QVBoxLayout()
        buttonsLayoutExtruder1 = QHBoxLayout() 
        
        buttonsLayoutExtruder1.addWidget(self.moveToExtruder2)  
        buttonsLayoutExtruder1.addWidget(self.sortExtruder1byDeliveryDateButton) 
        buttonsLayoutExtruder1.addWidget(self.moveRowUp1) 
        buttonsLayoutExtruder1.addWidget(self.moveRowDown1)
        buttonsLayoutExtruder1.addWidget(self.createShiftPlan1)
        buttonsLayoutExtruder1.addWidget(self.enumerateBatches1)
        buttonsLayoutExtruder1.addWidget(self.deleteBatchExtruder1) 
        buttonsLayoutExtruder1.addStretch()
        
        tabLayoutExtruder1.addLayout(buttonsLayoutExtruder1) 
        tabLayoutExtruder1.addWidget(self.tableBatchesExtruder1)                  
        tabLayoutExtruder1.addStretch()

        self.tabExtruder1.setLayout(tabLayoutExtruder1)  

    def _createPlanerViewExtruder2(self):  

        tableHorizontalHeaders = ['KW', 'Schichten', 'Beginn', 'Ende', 'Artikel-Nr.', 'Chargen-Nr.', 'Dispo.-Nr.', 'Kunde', 'Zusatz', 'Labor', 'Abholung', 't', 'Vorlauf', 'Bemerkung']    
                        
        self.tableBatchesExtruder2 = QTableWidget() 
        self.tableBatchesExtruder2.verticalHeader().setVisible(False) 
        self.tableBatchesExtruder2.setFixedWidth(1450) 
        self.tableBatchesExtruder2.setFixedHeight(875)    
        self.tableBatchesExtruder2.setColumnCount(14)  
        self.tableBatchesExtruder2.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)     
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(0, 38)     
        self.tableBatchesExtruder2.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(2, 80)  
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(3, 80)  
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(4, 100)  
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(5, 100) 
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(6, 100)
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(7, 150)     
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(8, 120)             
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(9, 80) 
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(10, 80)
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(11, 38) 
        self.tableBatchesExtruder2.horizontalHeader().setSectionResizeMode(12, QHeaderView.ResizeMode.ResizeToContents)   
        self.tableBatchesExtruder2.horizontalHeader().resizeSection(13, 350)  
        self.tableBatchesExtruder2.setHorizontalHeaderLabels(tableHorizontalHeaders)

        self.sortExtruder2byDeliveryDateButton = QPushButton('Sortieren')
        self.sortExtruder2byDeliveryDateButton.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'arrow-down-short-wide-solid.svg')))
        self.sortExtruder2byDeliveryDateButton.setFixedWidth(100)
        self.sortExtruder2byDeliveryDateButton.clicked.connect(lambda: self.sortExtruderbyDeliveryDateButton(2))                   

        self.moveToExtruder1 = QPushButton('zu Extruder 1')
        self.moveToExtruder1.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'arrows-turn-to-dots-solid')))
        self.moveToExtruder1.setFixedWidth(100)
        self.moveToExtruder1.clicked.connect(lambda: self.moveBatchToExtruder(2))

        self.deleteBatchExtruder2 = QPushButton('Löschen')
        self.deleteBatchExtruder2.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'trash-solid.svg')))
        self.deleteBatchExtruder2.setFixedWidth(100)
        self.deleteBatchExtruder2.clicked.connect(lambda: self.deleteBatchFromListExtruder(2))

        self.moveRowUp2 = QPushButton('nach oben')
        self.moveRowUp2.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'up-long-solid.svg')))
        self.moveRowUp2.setFixedWidth(100)
        self.moveRowUp2.clicked.connect(lambda: self.moveBatchRowUp(2))

        self.moveRowDown2 = QPushButton('nach unten')
        self.moveRowDown2.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'down-long-solid.svg')))
        self.moveRowDown2.setFixedWidth(100)
        self.moveRowDown2.clicked.connect(lambda: self.moveBatchRowDown(2))

        self.createShiftPlan2 = QPushButton('Schichten')
        self.createShiftPlan2.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'user-clock-solid.svg')))
        self.createShiftPlan2.setFixedWidth(100)
        self.createShiftPlan2.clicked.connect(lambda: self.createShiftPlan(2))

        self.enumerateBatches2 = QPushButton('Nummerieren')
        self.enumerateBatches2.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'arrow-down-1-9-solid.svg')))
        self.enumerateBatches2.setFixedWidth(100)
        self.enumerateBatches2.clicked.connect(lambda: self.enumerateBatches(2))

        tabLayoutExtruder2 = QVBoxLayout()
        buttonsLayoutExtruder2 = QHBoxLayout() 
        
        buttonsLayoutExtruder2.addWidget(self.moveToExtruder1)
        buttonsLayoutExtruder2.addWidget(self.sortExtruder2byDeliveryDateButton)  
        buttonsLayoutExtruder2.addWidget(self.moveRowUp2) 
        buttonsLayoutExtruder2.addWidget(self.moveRowDown2)         
        buttonsLayoutExtruder2.addWidget(self.createShiftPlan2)
        buttonsLayoutExtruder2.addWidget(self.enumerateBatches2)
        buttonsLayoutExtruder2.addWidget(self.deleteBatchExtruder2) 
        buttonsLayoutExtruder2.addStretch()
        
        tabLayoutExtruder2.addLayout(buttonsLayoutExtruder2) 
        tabLayoutExtruder2.addWidget(self.tableBatchesExtruder2)                  
        tabLayoutExtruder2.addStretch()

        self.tabExtruder2.setLayout(tabLayoutExtruder2)        
    
    def _createPlanerViewHomogenisation(self):                     
        
        tableHorizontalHeaders = ['KW', 'Schichten', 'Beginn', 'Ende', 'Artikel-Nr.', 'Chargen-Nr.', 'Dispo.-Nr.', 'Kunde', 'Zusatz', 'Labor', 'Abholung', 't', 'Vorlauf', 'Bemerkung']

        self.tableBatchesHomogenisation = QTableWidget() 
        self.tableBatchesHomogenisation.verticalHeader().setVisible(False)
        self.tableBatchesHomogenisation.setFixedWidth(1450) 
        self.tableBatchesHomogenisation.setFixedHeight(875)  
        self.tableBatchesHomogenisation.setColumnCount(14)  
        self.tableBatchesHomogenisation.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)   
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(0, 38)     
        self.tableBatchesHomogenisation.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(2, 80)  
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(3, 80)  
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(4, 100)
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(5, 100) 
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(6, 100)
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(7, 150)     
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(8, 120)             
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(9, 80) 
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(10, 80)
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(11, 38) 
        self.tableBatchesHomogenisation.horizontalHeader().setSectionResizeMode(12, QHeaderView.ResizeMode.ResizeToContents)   
        self.tableBatchesHomogenisation.horizontalHeader().resizeSection(13, 350)         
        self.tableBatchesHomogenisation.setHorizontalHeaderLabels(tableHorizontalHeaders)  

        self.sortHomogenisationbyDeliveryDateButton = QPushButton('Sortieren')
        self.sortHomogenisationbyDeliveryDateButton.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'arrow-down-short-wide-solid.svg')))
        self.sortHomogenisationbyDeliveryDateButton.setFixedWidth(100)
        self.sortHomogenisationbyDeliveryDateButton.clicked.connect(lambda: self.sortExtruderbyDeliveryDateButton(3))                

        self.deleteBatchHomogenisation = QPushButton('Löschen')
        self.deleteBatchHomogenisation.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'trash-solid.svg')))
        self.deleteBatchHomogenisation.setFixedWidth(100)
        self.deleteBatchHomogenisation.clicked.connect(lambda: self.deleteBatchFromListExtruder(3))

        self.moveRowUp3 = QPushButton('nach oben')
        self.moveRowUp3.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'up-long-solid.svg')))
        self.moveRowUp3.setFixedWidth(100)
        self.moveRowUp3.clicked.connect(lambda: self.moveBatchRowUp(3))

        self.moveRowDown3 = QPushButton('nach unten')
        self.moveRowDown3.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'down-long-solid.svg')))
        self.moveRowDown3.setFixedWidth(100)
        self.moveRowDown3.clicked.connect(lambda: self.moveBatchRowDown(3))

        self.createShiftPlan3 = QPushButton('Schichten')
        self.createShiftPlan3.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'user-clock-solid.svg')))
        self.createShiftPlan3.setFixedWidth(100)
        self.createShiftPlan3.clicked.connect(lambda: self.createShiftPlan(3))

        self.printPlans = QPushButton('Plan drucken')
        self.printPlans.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'print-solid.svg')))       
        self.printPlans.clicked.connect(lambda: self.handlePrint(1))
        self.printPlans.setShortcut("Ctrl+P")    

        tabLayoutHomogenisation = QVBoxLayout()
        buttonsLayoutHomogenisation = QHBoxLayout() 

        buttonsLayoutHomogenisation.addWidget(self.sortHomogenisationbyDeliveryDateButton) 
        buttonsLayoutHomogenisation.addWidget(self.moveRowUp3) 
        buttonsLayoutHomogenisation.addWidget(self.moveRowDown3)
        buttonsLayoutHomogenisation.addWidget(self.createShiftPlan3)
        buttonsLayoutHomogenisation.addWidget(self.deleteBatchHomogenisation) 
        buttonsLayoutHomogenisation.addWidget(self.printPlans)
        buttonsLayoutHomogenisation.addStretch()
        
        tabLayoutHomogenisation.addLayout(buttonsLayoutHomogenisation) 
        tabLayoutHomogenisation.addWidget(self.tableBatchesHomogenisation)                  
        tabLayoutHomogenisation.addStretch()

        self.tabHomogenisation.setLayout(tabLayoutHomogenisation)

    def _createPlanerViewSilo(self):                     
        
        tableHorizontalHeaders = ['KW', 'Schichten', 'Beginn', 'Ende', 'Artikel-Nr.', 'Chargen-Nr.', 'Dispo.-Nr.', 'Kunde', 'Zusatz', 'Labor', 'Abholung', 't', 'Vorlauf', 'Bemerkung']

        self.tableBatchesSilo = QTableWidget() 
        self.tableBatchesSilo.verticalHeader().setVisible(False)
        self.tableBatchesSilo.setFixedWidth(1450) 
        self.tableBatchesSilo.setFixedHeight(875)  
        self.tableBatchesSilo.setColumnCount(14)  
        self.tableBatchesSilo.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)   
        self.tableBatchesSilo.horizontalHeader().resizeSection(0, 38)     
        self.tableBatchesSilo.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.tableBatchesSilo.horizontalHeader().resizeSection(2, 80)  
        self.tableBatchesSilo.horizontalHeader().resizeSection(3, 80)  
        self.tableBatchesSilo.horizontalHeader().resizeSection(4, 100)
        self.tableBatchesSilo.horizontalHeader().resizeSection(5, 100) 
        self.tableBatchesSilo.horizontalHeader().resizeSection(6, 100)
        self.tableBatchesSilo.horizontalHeader().resizeSection(7, 150)     
        self.tableBatchesSilo.horizontalHeader().resizeSection(8, 120)             
        self.tableBatchesSilo.horizontalHeader().resizeSection(9, 80) 
        self.tableBatchesSilo.horizontalHeader().resizeSection(10, 80)
        self.tableBatchesSilo.horizontalHeader().resizeSection(11, 38) 
        self.tableBatchesSilo.horizontalHeader().setSectionResizeMode(12, QHeaderView.ResizeMode.ResizeToContents)  
        self.tableBatchesSilo.horizontalHeader().resizeSection(13, 350)        
        self.tableBatchesSilo.setHorizontalHeaderLabels(tableHorizontalHeaders)  

        self.sortSilobyDeliveryDateButton = QPushButton('Sortieren')
        self.sortSilobyDeliveryDateButton.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'arrow-down-short-wide-solid.svg')))
        self.sortSilobyDeliveryDateButton.setFixedWidth(100)
        self.sortSilobyDeliveryDateButton.clicked.connect(lambda: self.sortExtruderbyDeliveryDateButton(4))                

        self.deleteBatchSilo = QPushButton('Löschen')
        self.deleteBatchSilo.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'trash-solid.svg')))
        self.deleteBatchSilo.setFixedWidth(100)
        self.deleteBatchSilo.clicked.connect(lambda: self.deleteBatchFromListExtruder(4))

        self.moveRowUp4 = QPushButton('nach oben')
        self.moveRowUp4.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'up-long-solid.svg')))
        self.moveRowUp4.setFixedWidth(100)
        self.moveRowUp4.clicked.connect(lambda: self.moveBatchRowUp(4))

        self.moveRowDown4 = QPushButton('nach unten')
        self.moveRowDown4.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'down-long-solid.svg')))
        self.moveRowDown4.setFixedWidth(100)
        self.moveRowDown4.clicked.connect(lambda: self.moveBatchRowDown(4))

        self.createShiftPlan4 = QPushButton('Schichten')
        self.createShiftPlan4.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'user-clock-solid.svg')))
        self.createShiftPlan4.setFixedWidth(100)
        self.createShiftPlan4.clicked.connect(lambda: self.createShiftPlan(4))

        self.printPlans = QPushButton('Plan drucken')
        self.printPlans.setIcon(QIcon(os.path.join(self.imagePath, 'assets', 'print-solid.svg')))       
        self.printPlans.clicked.connect(lambda: self.handlePrint(2))
        self.printPlans.setShortcut("Ctrl+P")  

        tabLayoutSilo = QVBoxLayout()
        buttonsLayoutSilo = QHBoxLayout() 

        buttonsLayoutSilo.addWidget(self.sortSilobyDeliveryDateButton) 
        buttonsLayoutSilo.addWidget(self.moveRowUp4) 
        buttonsLayoutSilo.addWidget(self.moveRowDown4)
        buttonsLayoutSilo.addWidget(self.createShiftPlan4)
        buttonsLayoutSilo.addWidget(self.deleteBatchSilo) 
        buttonsLayoutSilo.addWidget(self.printPlans)
        buttonsLayoutSilo.addStretch()
        
        tabLayoutSilo.addLayout(buttonsLayoutSilo) 
        tabLayoutSilo.addWidget(self.tableBatchesSilo)                  
        tabLayoutSilo.addStretch()

        self.tabSilo.setLayout(tabLayoutSilo)

    def _createTextArea(self):
        self.additiveUsageText = QTextEdit()
        self.additiveUsageText.setReadOnly(True)
        self.additiveUsageText.setLineWrapMode(QTextEdit.LineWrapMode.NoWrap)
        self.additiveUsageText.setFixedWidth(350)

    def _createMaster(self):       
        

        tabWidget = QWidget()   
        tabWidget.setLayout(self.tabLayout)

        guiWidget = QHBoxLayout()  
        guiWidget.addWidget(tabWidget) 
        guiWidget.addWidget(self.additiveUsageText) 

        masterWidget = QWidget()   
        masterWidget.setLayout(guiWidget)          

        self.setCentralWidget(masterWidget)          

    def _loadData(self):      

        self.articleNoList = [] 
        self.customerList = []    


        wb = load_workbook(filename=self.dataXLSX)
        sheets = wb.sheetnames

        sheetsNo = 4

        articleListHelp = {}
        additiveListHelp = {}
        self.colorList = {}        

        for sheet in range(sheetsNo):
            ws = wb[sheets[sheet]]   
            match sheet:
                    case 0: 
                        rowItem = 0                              
                        for row in ws.iter_rows(values_only=True):         

                            articleListHelp[rowItem] = [row[0], row[1], row[2], ast.literal_eval(row[3]), row[4]]
                            rowItem = rowItem + 1

                        
                        self.articleList = dict(sorted(articleListHelp.items(), key=lambda item: item[1][1]))                                                               

                    case 1:
                        rowItem = 0       
                        for row in ws.iter_rows(values_only=True):         

                            additiveListHelp[rowItem] = [row[0], row[1], row[2]]
                            rowItem = rowItem + 1

                        self.additiveList = dict(sorted(additiveListHelp.items(), key=lambda item: item[1][0]))
                        
                    
                    case 2:                             
                        for row in ws.iter_rows(values_only=True):         

                            self.customerList.append(row[0])                            

                        self.customerList.sort()

                    case 3:
                              
                        for row in ws.iter_rows(values_only=True):         

                            self.colorList[row[0]] = [row[1]]
                            
        self.dictArticleColors = {}  
        self.articleExtruderList = {}       
        for key in self.articleList: 

            self.dictArticleColors[self.articleList[key][1]] = self.articleList[key][2][-5:][:2]
            self.articleNoList.append(self.articleList[key][1]) 
            self.articleExtruderList[self.articleList[key][1]] = self.articleList[key][4]
                                                  
    def saveData(self, changedList):
        
        mode = changedList[1] 
        
        tableList = [self.tableBatchesExtruder1, self.tableBatchesExtruder2, self.tableBatchesSilo, self.tableBatchesHomogenisation ] 
   
        match mode:
            case 0:                                                
                for table in range(len(tableList)):
                    whichTable = tableList[table] 

                    
                    if whichTable.rowCount() != 0:                                                                      
                            
                        for row in range(whichTable.rowCount()):

                            if table == 0 or table == 1:
                                index = whichTable.cellWidget(row, 7).currentIndex()
                                whichTable.cellWidget(row, 7).clear()
                                whichTable.cellWidget(row, 7).addItem('')
                                whichTable.cellWidget(row, 7).addItems(self.customerList)  
                                whichTable.cellWidget(row, 7).setCurrentIndex(index)                                                      
                                              
            case 1:
                for table in range(len(tableList)):
                    whichTable = tableList[table]  

                    self.articleNoList = []
                    self.dictArticleColors = {}
                    self.articleExtruderList = {}         
                    for key in self.articleList: 

                        self.dictArticleColors[self.articleList[key][1]] = self.articleList[key][2][-5:][:2] 
                        self.articleNoList.append(self.articleList[key][1]) 
                        self.articleExtruderList[self.articleList[key][1]] = self.articleList[key][4]                            
                    
                    if whichTable.rowCount() != 0: 
                        self.changeArticleNoCheck = True
                        for row in range(whichTable.rowCount()): 
                            print(whichTable.rowCount())
                            if table == 0 or table == 1:                                                                       
                                
                                index = whichTable.cellWidget(row, 4).currentIndex()
                                whichTable.cellWidget(row, 4).clear()
                                whichTable.cellWidget(row, 4).addItems(self.articleNoList) 
                                for article in range(len(self.articleNoList)):                       
                                    whichIcon = self.colorList[self.dictArticleColors[self.articleNoList[article]]][0]+'.png'               
                                    whichTable.cellWidget(row, 4).setItemIcon(article, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))   
                                whichTable.cellWidget(row, 4).setCurrentIndex(index)
                        self.changeArticleNoCheck = False

       
        wb = Workbook() 

        ws = wb.active  
        ws.title = 'Artikel'
        wb.create_sheet('Additive')
        wb.create_sheet('Kunden')
        wb.create_sheet('Farben')

        sheets = wb.sheetnames

        sheetsNo = 4
        

        for sheet in range(sheetsNo):
            saveTableData = []
            saveRow = []

            ws = wb[sheets[sheet]]
            match sheet:
                case 0:                     
                    for key in self.articleList:        
                        saveRow = [self.articleList[key][0], self.articleList[key][1], self.articleList[key][2], str(self.articleList[key][3]), str(self.articleList[key][4])]  
                        saveTableData.append(saveRow) 
                         
                    for writeRow in saveTableData:
                        ws.append(writeRow)           
                                
                case 1:           
                    for key in self.additiveList:             
                        saveRow = [self.additiveList[key][0], self.additiveList[key][1], self.additiveList[key][2]]  
                        saveTableData.append(saveRow)                     
                         
                    for writeRow in saveTableData:
                        ws.append(writeRow) 
                case 2:                     
                    for writeRow in self.customerList:
                        saveRow = [writeRow]   
                        saveTableData.append(saveRow) 

                    for writeRow in saveTableData:
                        ws.append(writeRow)

                case 3:           
                    for key, value in self.colorList.items():             
                        saveRow = [key, value[0]]  
                        saveTableData.append(saveRow)                     
                         
                    for writeRow in saveTableData:
                        ws.append(writeRow) 
                                       
         
        wb.save(self.dataXLSX)

        if self.setLoadedFile == True:
            self.performSaveFile() 
                
    def openSecondaryWindow(self, window):
        
        self.tableBatchesExtruder1.setDisabled(True)
        self.tableBatchesExtruder2.setDisabled(True)        
        self.moveToExtruder2.setDisabled(True)
        self.moveToExtruder1.setDisabled(True) 
        self.sortExtruder1byDeliveryDateButton.setDisabled(True)
        self.deleteBatchExtruder1.setDisabled(True)  
        self.moveRowUp1.setDisabled(True)
        self.sortExtruder2byDeliveryDateButton.setDisabled(True)
        self.deleteBatchExtruder2.setDisabled(True)  
        self.moveRowUp2.setDisabled(True)
        self.tabs.setDisabled(True)
        self.menubar.setDisabled(True)


        self.closeMenu = False

        if self.w is None:  
            match window:
                case 0:          
                    self.w = AddBatchWindow(self.customerList, self.articleNoList, self.attrPack, self.attrLab, self.timeNormal, self.timeDensity, self.timeMechanics, self.timeReach, self.colorList, self.dictArticleColors, self.dispoNoList)         
                    self.w.show()
                    self.w.finished.connect(self.closeSecondaryWindow)
                    self.w.added.connect(self.addBatchesToList)
                case 1:
                    self.w = SettingsWindow(self.sortBy, self.timeNormal, self.timeDensity, self.timeMechanics, self.timeReach, self.usageFactor, self.maintenanceDay)         
                    self.w.show()
                    self.w.finished.connect(self.closeSecondaryWindow)
                    self.w.added.connect(self.writeSettingsToIni)
                case 2:                    
                    self.w = EditDataWindow(0, self.articleList, self.additiveList, self.customerList)         
                    self.w.show()
                    self.w.finished.connect(self.closeSecondaryWindow)
                    self.w.added.connect(self.saveData)

                case 3:
                    self.w = EditDataWindow(1, self.articleList, self.additiveList, self.customerList)         
                    self.w.show()
                    self.w.finished.connect(self.closeSecondaryWindow)
                    self.w.added.connect(self.saveData)

                case 4:  
                    self.w = EditDataWindow(2, self.articleList, self.additiveList, self.customerList)         
                    self.w.show()
                    self.w.finished.connect(self.closeSecondaryWindow)
                    self.w.added.connect(self.saveData) 

        else:
            self.w.close()
            self.w = None  

    def closeSecondaryWindow(self):
        self.w = None 
        self.tableBatchesExtruder1.setDisabled(False)
        self.tableBatchesExtruder2.setDisabled(False)        
        self.moveToExtruder2.setDisabled(False)
        self.moveToExtruder1.setDisabled(False)
        self.sortExtruder1byDeliveryDateButton.setDisabled(False)
        self.deleteBatchExtruder1.setDisabled(False) 
        self.moveRowUp1.setDisabled(False)
        self.sortExtruder2byDeliveryDateButton.setDisabled(False)
        self.deleteBatchExtruder2.setDisabled(False)
        self.moveRowUp2.setDisabled(False)
        self.tabs.setDisabled(False)
        self.menubar.setDisabled(False)
        
        self.closeMenu = True        

    def addBatchesToList(self, addBatchArray): 
        self.tableBatchesExtruder1.blockSignals(True)
        self.tableBatchesExtruder2.blockSignals(True)
        
        self.saveFile.setEnabled(True)

        deliveryDate = datetime.datetime.strptime(addBatchArray[10], '%d.%m.%Y')
        
        if addBatchArray[9] == 0: 
            productionDate = deliveryDate - datetime.timedelta(days=self.timeNormal)  
            
        elif addBatchArray[9] == 1:            
            productionDate = datetime.datetime.strptime(addBatchArray[10], '%d.%m.%Y') - datetime.timedelta(days=self.timeDensity)
            
        elif addBatchArray[9] == 2:            
            productionDate = datetime.datetime.strptime(addBatchArray[10], '%d.%m.%Y') - datetime.timedelta(days=self.timeMechanics)
                                 
        elif addBatchArray[9] == 3:            
            productionDate = datetime.datetime.strptime(addBatchArray[10], '%d.%m.%Y') - datetime.timedelta(days=self.timeReach)
            

        timeToDelivery = (deliveryDate - productionDate).days
       
        year =  addBatchArray[10][-2:]              
        
        calendarWeek = productionDate.strftime('%V')

        rx2 = QRegularExpression("3\\d{1}\\.\\d{1,4}\\-\\d{1}")
        rx3 = QRegularExpression("1-\\d{1,2}-\\d{1,3}")
        rx4 = QRegularExpression("\\d{1,2}")

        if addBatchArray[4] in self.articleExtruderList and self.articleExtruderList[addBatchArray[4]] == '1':
            whichTable = self.tableBatchesExtruder2
            table = 2
        else:
            whichTable = self.tableBatchesExtruder1
            table = 1

        rowPosition = whichTable.rowCount()

        self.dispoNoList[addBatchArray[6]] = rowPosition       

        whichTable.insertRow(rowPosition)

        item = 0
        for item in range(14):
            match item:
                case 0:                    
                    whichCalendarWeek = QLabel()
                    whichCalendarWeek.setText(calendarWeek)
                    whichCalendarWeek.setFixedWidth(40)
                    whichCalendarWeek.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    whichTable.setCellWidget(rowPosition, item, whichCalendarWeek)

                case 1:                                    
                    whichShift = QComboBox()
                    whichShift.addItems(self.attrShift)
                    whichShift.setProperty('row', rowPosition)
                    if table == 1:
                        whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(1))
                    else:
                        whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(2))
                    whichTable.setCellWidget(rowPosition, item, whichShift)

                case 2:
                    buttonProductionDate = QDateEdit()
                    buttonProductionDate.setFixedWidth(80)
                    buttonProductionDate.setDate(productionDate) 
                    buttonProductionDate.setProperty('row', rowPosition)       
                    if table == 1:
                        buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(1))
                    else:
                        buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(2))
                    whichTable.setCellWidget(rowPosition, item, buttonProductionDate) 

                case 3:
                    buttonProductionDate = QDateEdit()
                    buttonProductionDate.setFixedWidth(80)
                    buttonProductionDate.setDate(productionDate)
                    buttonProductionDate.setProperty('row', rowPosition) 
                    if table == 1:      
                        buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(1))
                    else:
                        buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(2))
                    whichTable.setCellWidget(rowPosition, item, buttonProductionDate)                    

                case 4:                                                    
                    whichArticle = QComboBox()
                    whichArticle.addItems(self.articleNoList)     
                    whichArticle.setCurrentText(addBatchArray[4])
                    whichArticle.setEditable(True)
                    whichArticle.setValidator(QRegularExpressionValidator(rx2, self)) 
                    if table == 1:      
                        whichArticle.currentIndexChanged.connect(lambda: self.articleChanged(1))
                    else:
                        whichArticle.currentIndexChanged.connect(lambda: self.articleChanged(2))
                    whichArticle.setProperty('row', rowPosition) 
                    for article in range(len(self.articleNoList)):                       
                        whichIcon = self.colorList[self.dictArticleColors[self.articleNoList[article]]][0]+'.png'               
                        whichArticle.setItemIcon(article, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))
                    
                    whichTable.setCellWidget(rowPosition, item, whichArticle)  

                case 5:                                      
                    newBatchNo = QPushButton()                                                                                      
                    if table == 1: 
                        newBatchNo.clicked.connect(lambda: self.changeBatchNo(1))
                        newBatchNo.setText('1-'+year+'-')
                    else:
                        newBatchNo.clicked.connect(lambda: self.changeBatchNo(2))
                        newBatchNo.setText('2-'+year+'-')
                    newBatchNo.setProperty('row', rowPosition)
                    newBatchNo.setFixedWidth(100)                     
                    whichTable.setCellWidget(rowPosition, item, newBatchNo) 

                case 6:                                         
                    newDispo = QPushButton()
                    newDispo.setText(addBatchArray[item])
                    newDispo.setFixedWidth(100)
                    if table == 1: 
                        newDispo.clicked.connect(lambda: self.changeDispoNo(1))
                    else:
                        newDispo.clicked.connect(lambda: self.changeDispoNo(2))
                    newDispo.setProperty('row', rowPosition) 
                    whichTable.setCellWidget(rowPosition, item, newDispo)    

                case 7:                                
                    whichCustomer = QComboBox()
                    whichCustomer.addItem(' ')
                    whichCustomer.addItems(self.customerList)
                    whichCustomer.setCurrentText(addBatchArray[7])
                    whichCustomer.setEditable(True)
                    whichCustomer.setProperty('row', rowPosition)
                    if table == 1: 
                        whichCustomer.currentIndexChanged.connect(lambda: self.customerChanged(1))
                    else:
                        whichCustomer.currentIndexChanged.connect(lambda: self.customerChanged(2)) 
                    whichTable.setCellWidget(rowPosition, item, whichCustomer)   

                case 8:                                    
                    whichPackaging = QComboBox()
                    whichPackaging.addItems(self.attrPack)
                    whichPackaging.setCurrentIndex(addBatchArray[8])
                    whichPackaging.setProperty('row', rowPosition)
                    if table == 1:
                        whichPackaging.currentIndexChanged.connect(lambda: self.packagingChanged(1))
                    else:
                        whichPackaging.currentIndexChanged.connect(lambda: self.packagingChanged(2))
                    whichTable.setCellWidget(rowPosition, item, whichPackaging)      

                case 9:                                    
                    whichLab = QComboBox()
                    whichLab.addItems(self.attrLab)
                    whichLab.setCurrentIndex(addBatchArray[9]) 
                    whichLab.setProperty('row', rowPosition)
                    whichLab.setFixedWidth(80)
                    if table == 1:
                        whichLab.currentIndexChanged.connect(lambda: self.labChanged(1))
                    else:
                        whichLab.currentIndexChanged.connect(lambda: self.labChanged(2))
                    whichTable.setCellWidget(rowPosition, item, whichLab)             

                case 10:
                    buttonDeliveryDate = QDateEdit()
                    buttonDeliveryDate.setFixedWidth(80)
                    buttonDeliveryDate.setDate(datetime.datetime.strptime(addBatchArray[item], '%d.%m.%Y'))     
                    buttonDeliveryDate.setProperty('row', rowPosition)
                    if table == 1:            
                        buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(1)) 
                    else:
                        buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(2))           
                    whichTable.setCellWidget(rowPosition, item, buttonDeliveryDate)  
                     
                case 11:
                    whichBatchSize = QLineEdit()
                    if addBatchArray[11] == '':
                        batchSize = '24'
                    else:
                        batchSize = addBatchArray[11]
                    whichBatchSize.setText(batchSize)
                    whichBatchSize.setEnabled(True)
                    whichBatchSize.setFixedWidth(38)
                    whichBatchSize.setValidator(QRegularExpressionValidator(rx4, self)) 
                    whichBatchSize.setProperty('row', rowPosition)
                    if table == 1:            
                        whichBatchSize.textChanged.connect(lambda: self.batchSizeChanged(1)) 
                    else:
                        whichBatchSize.textChanged.connect(lambda: self.batchSizeChanged(2)) 
                    whichTable.setCellWidget(rowPosition, item, whichBatchSize)

                case 12:
                    whichDeliveryDate = QLabel()
                    whichDeliveryDate.setText(str(timeToDelivery))                    
                    whichDeliveryDate.setFixedWidth(38)
                    whichDeliveryDate.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    whichTable.setCellWidget(rowPosition, item, whichDeliveryDate)

                    newTimeToDelivery = int(whichTable.cellWidget(rowPosition, 12).text())

                    if whichTable.cellWidget(rowPosition, 9).currentIndex() == 1 and newTimeToDelivery < self.timeDensity:            
                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')
                    elif whichTable.cellWidget(rowPosition, 9).currentIndex() == 2 and newTimeToDelivery < self.timeMechanics:            
                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                    elif whichTable.cellWidget(rowPosition, 9).currentIndex() == 3 and newTimeToDelivery < self.timeReach:            
                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                    elif newTimeToDelivery < self.timeNormal:            
                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')       
                    else:
                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: white')

                    if whichTable.cellWidget(rowPosition, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):             
                        whichTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: red')
                    else:
                        whichTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: white')            
                
                case 13:
                    whichComment = QLineEdit()
                    whichComment.setFixedWidth(350)
                    whichComment.textChanged.connect(self.enableSaving)
                    whichTable.setCellWidget(rowPosition, item, whichComment)  

        self.generateSiloLists()
        self.generateAdditiveUsage()
        self.tableBatchesExtruder1.blockSignals(False)
        self.tableBatchesExtruder2.blockSignals(False)

    def writeSettingsToIni(self, changedSettings): 
        self.tableBatchesExtruder1.blockSignals(True)
        self.tableBatchesExtruder2.blockSignals(True)

        self.sortBy = changedSettings[0] 
        match self.sortBy:
            case 0:
                self.sortByColumn = 2
            case 1:
                self.sortByColumn = 3
            case 2:
                self.sortByColumn = 10

        self.timeNormal = int(changedSettings[1])
        self.timeDensity = int(changedSettings[2])
        self.timeMechanics = int(changedSettings[3])
        self.timeReach = int(changedSettings[4]) 
        self.usageFactor = int(changedSettings[5])
        self.maintenanceDay = int(changedSettings[6])

        self.config['SETTINGS']['sortby'] = str(self.sortBy)
        self.config['SETTINGS']['timenormal'] = str(changedSettings[1]) 
        self.config['SETTINGS']['timedensity'] = str(changedSettings[2]) 
        self.config['SETTINGS']['timemechanics'] = str(changedSettings[3]) 
        self.config['SETTINGS']['timereach'] = str(changedSettings[4]) 
        self.config['USAGE']['factor'] = str(changedSettings[5]) 
        self.config['SETTINGS']['maintenanceday'] = str(changedSettings[6])

        with open(os.path.join(self.imagePath, 'settings.ini'), 'w') as configfile:
                self.config.write(configfile) 
            

        self.tableBatchesExtruder1.blockSignals(False)
        self.tableBatchesExtruder2.blockSignals(False)

    def closeEvent(self, event):

        if self.saveFile.isEnabled() == True:

            messageBox = QMessageBox().warning(self, 'Schließen ohne zu speichern?', 'Du hast ungespeicherte Änderungen.\nTrotzdem schließen?', buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)        
        
            if messageBox == QMessageBox.StandardButton.Yes:                
                event.accept()
            else:                
                event.ignore()

        else:
            if self.closeMenu == True:
                event.accept()         

            else:
                event.ignore()
                     
    def moveBatchToExtruder(self, table):
        self.tableBatchesExtruder1.blockSignals(True)
        self.tableBatchesExtruder2.blockSignals(True)  
        
        self.saveFile.setEnabled(True)          

        if table == 1:            
            whichTable = self.tableBatchesExtruder1
            otherTable = self.tableBatchesExtruder2
        else:           
            whichTable = self.tableBatchesExtruder2
            otherTable = self.tableBatchesExtruder1 

        rx = QRegularExpression("SP\\d{1,7}[a-zA-Z]")
        rx2 = QRegularExpression("3\\d{1}\\.\\d{1,4}\\-\\d{1}")
        rx3 = QRegularExpression("1-\\d{1,2}-\\d{1,3}")
        rx4 = QRegularExpression("\\d{1,2}")
                    
        if len(whichTable.selectionModel().selectedRows()) != 0:
            
            rowList = []
            
            for row in whichTable.selectionModel().selectedRows():
                rowList.append(row.row())
                
            rowList.sort() 

            for row in rowList:
                if not whichTable.cellWidget(row, 4).currentText() == '32.':
                    if (self.articleExtruderList[whichTable.cellWidget(row, 4).currentText()] == '1' and table == 2) or (self.articleExtruderList[whichTable.cellWidget(row, 4).currentText()] == '2' and table == 1):

                        invalidArticle = whichTable.cellWidget(row, 4).currentText()
                        if table == 1:
                            invalidExtruder = 'Extruder 2'
                        else:
                            invalidExtruder = 'Extruder 1'

                        messageBox = QMessageBox().warning(self, "Achtung!", "Der Artikel " + invalidArticle + " kann oder darf nicht mit " + invalidExtruder + " produziert werden!\nZum anderen Extruder verschieben?", buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)        
    
                        if not messageBox == QMessageBox.StandardButton.Yes:  
                            return                       

            for item in rowList:
                
                rowPosition = otherTable.rowCount()
                otherTable.insertRow(rowPosition)

                rowItem = 0
                for rowItem in range(14):                  

                    match rowItem:
                        case 0:                    
                            whichCalendarWeek = QLabel()
                            whichCalendarWeek.setText(whichTable.cellWidget(item, rowItem).text())
                            whichCalendarWeek.setFixedWidth(40)
                            whichCalendarWeek.setAlignment(Qt.AlignmentFlag.AlignCenter)
                            otherTable.setCellWidget(rowPosition, rowItem, whichCalendarWeek)

                        case 1:                                            
                            whichShift = QComboBox()
                            whichShift.addItems(self.attrShift)
                            whichShift.setCurrentIndex(whichTable.cellWidget(item, rowItem).currentIndex())
                            if table == 1:
                                whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(2))
                            else:
                                whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(1))
                            whichShift.setProperty('row', rowPosition)                          
                            otherTable.setCellWidget(rowPosition, rowItem, whichShift)                       

                        case 2:                                  
                            buttonProductionDate = QDateEdit()
                            buttonProductionDate.setFixedWidth(80)
                            buttonProductionDate.setDate(datetime.datetime.strptime(whichTable.cellWidget(item, rowItem).text(), '%d.%m.%Y')) 
                            buttonProductionDate.setProperty('row', rowPosition)
                            if table == 1:       
                                buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(2)) 
                            else: 
                                buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(1))                                            
                            otherTable.setCellWidget(rowPosition, rowItem, buttonProductionDate) 

                        case 3:                                  
                            buttonProductionDate = QDateEdit()
                            buttonProductionDate.setFixedWidth(80)
                            buttonProductionDate.setDate(datetime.datetime.strptime(whichTable.cellWidget(item, rowItem).text(), '%d.%m.%Y'))
                            buttonProductionDate.setProperty('row', rowPosition)
                            if table == 1:       
                                buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(2)) 
                            else: 
                                buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(1))               
                            otherTable.setCellWidget(rowPosition, rowItem, buttonProductionDate)                            

                        case 4:                                
                            whichArticle = QComboBox()                            
                            whichArticle.addItems(self.articleNoList)
                            whichArticle.setValidator(QRegularExpressionValidator(rx2, self))
                            whichArticle.setCurrentIndex(whichTable.cellWidget(item, rowItem).currentIndex())
                            whichArticle.setEditable(True)
                            whichArticle.setProperty('row', rowPosition)
                            if table == 1:      
                                whichArticle.currentIndexChanged.connect(lambda: self.articleChanged(2))
                            else:
                                whichArticle.currentIndexChanged.connect(lambda: self.articleChanged(1))                            

                            for article in range(len(self.articleNoList)):                       
                                whichIcon = self.colorList[self.dictArticleColors[self.articleNoList[article]]][0]+'.png'               
                                whichArticle.setItemIcon(article, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))


                            if table == 1:
                                whichArticle.currentIndexChanged.connect(lambda: self.enableSaving)
                            else:
                                whichArticle.currentIndexChanged.connect(lambda: self.enableSaving)
                            otherTable.setCellWidget(rowPosition, rowItem, whichArticle)
                            
                        case 5:                                         
                            newBatchNo = QPushButton() 
                            if table == 1:                          
                                newBatchNo.setText('2-'+ whichTable.cellWidget(item, 10).text()[-2:] + '-' )
                            else:
                                newBatchNo.setText('1-'+ whichTable.cellWidget(item, 10).text()[-2:] + '-' )                            
                            newBatchNo.setFixedWidth(100)
                            if table == 1: 
                                newBatchNo.clicked.connect(lambda: self.changeBatchNo(2))
                            else:
                                newBatchNo.clicked.connect(lambda: self.changeBatchNo(1))
                            newBatchNo.setProperty('row', rowPosition)
                            otherTable.setCellWidget(rowPosition, rowItem, newBatchNo) 

                        case 6:                                                                
                            newDispo = QPushButton()
                            newDispo.setText(whichTable.cellWidget(item, rowItem).text())
                            newDispo.setFixedWidth(100)
                            if table == 1: 
                                newDispo.clicked.connect(lambda: self.changeDispoNo(2))
                            else:
                                newDispo.clicked.connect(lambda: self.changeDispoNo(1))
                            newDispo.setProperty('row', rowPosition) 
                            otherTable.setCellWidget(rowPosition, rowItem, newDispo)       

                        case 7:                                
                            whichCustomer = QComboBox()
                            whichCustomer.addItem(' ')
                            whichCustomer.addItems(self.customerList)                                
                            whichCustomer.setCurrentIndex(whichTable.cellWidget(item, rowItem).currentIndex())
                            whichCustomer.setEditable(True)
                            whichCustomer.setProperty('row', rowPosition)
                            if table == 1: 
                                whichCustomer.currentIndexChanged.connect(lambda: self.customerChanged(2))
                            else:
                                whichCustomer.currentIndexChanged.connect(lambda: self.customerChanged(1))
                            otherTable.setCellWidget(rowPosition, rowItem, whichCustomer) 

                        case 8:                
                            whichPackaging = QComboBox()
                            whichPackaging.addItems(self.attrPack)
                            whichPackaging.setProperty('row', rowPosition)
                            whichPackaging.setCurrentIndex(whichTable.cellWidget(item, rowItem).currentIndex())
                            if table == 1:
                                whichPackaging.currentIndexChanged.connect(lambda: self.packagingChanged(2))
                            else:
                                whichPackaging.currentIndexChanged.connect(lambda: self.packagingChanged(1))
                            otherTable.setCellWidget(rowPosition, rowItem, whichPackaging)      

                        case 9:               
                            whichLab = QComboBox()
                            whichLab.addItems(self.attrLab)
                            whichLab.setCurrentIndex(whichTable.cellWidget(item, rowItem).currentIndex())
                            whichLab.setProperty('row', rowPosition)
                            whichLab.setFixedWidth(80)
                            if table == 1:       
                                whichLab.currentIndexChanged.connect(lambda: self.labChanged(2)) 
                            else: 
                                whichLab.currentIndexChanged.connect(lambda: self.labChanged(1))                            
                            otherTable.setCellWidget(rowPosition, rowItem, whichLab)                           
                                
                        case 10:                                  
                            buttonDeliveryDate = QDateEdit()
                            buttonDeliveryDate.setFixedWidth(80)
                            buttonDeliveryDate.setDate(datetime.datetime.strptime(whichTable.cellWidget(item, rowItem).text(), '%d.%m.%Y')) 
                            if table == 1:       
                                buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(2)) 
                            else: 
                                buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(1))   
                            buttonDeliveryDate.setProperty('row', rowPosition)            
                            otherTable.setCellWidget(rowPosition, rowItem, buttonDeliveryDate)

                        case 11:
                            whichBatchSize = QLineEdit()
                            whichBatchSize.setText(whichTable.cellWidget(item, rowItem).text())
                            whichBatchSize.setEnabled(True)
                            whichBatchSize.setFixedWidth(38)
                            whichBatchSize.setValidator(QRegularExpressionValidator(rx4, self))
                            whichBatchSize.setProperty('row', rowPosition)
                            if table == 1:            
                                whichBatchSize.textChanged.connect(lambda: self.batchSizeChanged(2)) 
                            else:
                                whichBatchSize.textChanged.connect(lambda: self.batchSizeChanged(1))
                            otherTable.setCellWidget(rowPosition, rowItem, whichBatchSize)

                        case 12:
                            whichDeliveryDate = QLabel()
                            whichDeliveryDate.setText(whichTable.cellWidget(item, rowItem).text())
                            whichDeliveryDate.setFixedWidth(38)
                            whichDeliveryDate.setAlignment(Qt.AlignmentFlag.AlignCenter)
                            otherTable.setCellWidget(rowPosition, rowItem, whichDeliveryDate) 

                            newTimeToDelivery = int(otherTable.cellWidget(rowPosition, rowItem).text() ) 

                            if otherTable.cellWidget(rowPosition, 9).currentIndex() == 1 and newTimeToDelivery < self.timeDensity:            
                                otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')
                            elif otherTable.cellWidget(rowPosition, 9).currentIndex() == 2 and newTimeToDelivery < self.timeMechanics:            
                                otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                            elif otherTable.cellWidget(rowPosition, 9).currentIndex() == 3 and newTimeToDelivery < self.timeReach:            
                                otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                            elif newTimeToDelivery < self.timeNormal:            
                                otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')       
                            else:
                                otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: white')    

                            if otherTable.cellWidget(rowPosition, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):             
                                otherTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: red')
                            else:
                                otherTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: white')      

                        case 13:
                            whichComment = QLineEdit() 
                            whichComment.setText(whichTable.cellWidget(item, rowItem).text())
                            whichComment.setFixedWidth(350)
                            whichComment.textChanged.connect(self.enableSaving)
                            otherTable.setCellWidget(rowPosition, rowItem, whichComment)          
            
                if table == 1 or table == 2:
                    self.dispoNoList[otherTable.cellWidget(rowPosition, 6).text()] = rowPosition 

            rowList.sort(reverse=True)
            for item in rowList:    
                 whichTable.removeRow(item)
            
            for row in range(whichTable.rowCount()):
                whichTable.cellWidget(row, 1).setProperty('row', row) 
                whichTable.cellWidget(row, 2).setProperty('row', row)           
                whichTable.cellWidget(row, 3).setProperty('row', row)
                whichTable.cellWidget(row, 4).setProperty('row', row)
                whichTable.cellWidget(row, 5).setProperty('row', row)
                whichTable.cellWidget(row, 6).setProperty('row', row)
                whichTable.cellWidget(row, 7).setProperty('row', row)
                whichTable.cellWidget(row, 8).setProperty('row', row)
                whichTable.cellWidget(row, 9).setProperty('row', row) 
                whichTable.cellWidget(row, 10).setProperty('row', row) 
                whichTable.cellWidget(row, 11).setProperty('row', row) 

        self.tableBatchesExtruder1.blockSignals(False)
        self.tableBatchesExtruder2.blockSignals(False)

    def deleteBatchFromListExtruder(self, table):
        self.tableBatchesExtruder1.blockSignals(True)
        self.tableBatchesExtruder2.blockSignals(True)
        
        self.saveFile.setEnabled(True)

        if self.workingOnShiftPlan == False:
            match table:
                case 1:            
                    whichTable = self.tableBatchesExtruder1            
                case 2:           
                    whichTable = self.tableBatchesExtruder2
                case 3:
                    whichTable = self.tableBatchesHomogenisation
                case 4:
                    whichTable = self.tableBatchesSilo
                

            if len(whichTable.selectionModel().selectedRows()) != 0:            

                rowList = []
                for row in whichTable.selectionModel().selectedRows():
                    rowList.append(row.row())

                rowList.sort(reverse=True)            

                for item in rowList: 
                    if table == 1 or table == 2:            
                        self.dispoNoList.pop(whichTable.cellWidget(item, 6).text())
                    else:
                        self.checkDispoNoSilo.pop(whichTable.cellWidget(item, 6).text())
                    whichTable.removeRow(item)

            for row in range(whichTable.rowCount()):
                whichTable.cellWidget(row, 1).setProperty('row', row) 
                whichTable.cellWidget(row, 2).setProperty('row', row)           
                whichTable.cellWidget(row, 3).setProperty('row', row)
                whichTable.cellWidget(row, 5).setProperty('row', row)
                whichTable.cellWidget(row, 6).setProperty('row', row)
                whichTable.cellWidget(row, 7).setProperty('row', row)
                whichTable.cellWidget(row, 8).setProperty('row', row)
                whichTable.cellWidget(row, 9).setProperty('row', row) 
                whichTable.cellWidget(row, 10).setProperty('row', row) 
                whichTable.cellWidget(row, 11).setProperty('row', row)
                if table == 3 or table == 4:
                    self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()] = row                

        self.tableBatchesExtruder1.blockSignals(False)
        self.tableBatchesExtruder2.blockSignals(False)
    
    def productionStartDateChangedInTable(self, table):
        if self.workingOnShiftPlan == False:

            self.saveFile.setEnabled(True)
            self.tableBatchesExtruder1.blockSignals(True)
            self.tableBatchesExtruder2.blockSignals(True)
            self.tableBatchesHomogenisation.blockSignals(True)
            self.tableBatchesSilo.blockSignals(True)
            changedDate = self.sender()            
            row = changedDate.property('row')                

            match table:
                case 1:            
                    whichTable = self.tableBatchesExtruder1            
                case 2:           
                    whichTable = self.tableBatchesExtruder2
                case 3:
                    whichTable = self.tableBatchesHomogenisation
                case 4:
                    whichTable = self.tableBatchesSilo   


            whichShift = whichTable.cellWidget(row, 1).currentText()
            if whichShift == 'N-W-S' or whichShift == 'S-N' or whichShift == 'N-F':          

                whichTable.cellWidget(row, 3).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y') + datetime.timedelta(days=1))  
            else:
                whichTable.cellWidget(row, 3).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'))

            newKW = datetime.datetime.strptime(whichTable.cellWidget(row, 3).date().toString('dd.MM.yyyy'), '%d.%m.%Y').strftime('%V')
            whichTable.cellWidget(row, 0).setText(newKW)

            newTimeToDelivery = (datetime.datetime.strptime(whichTable.cellWidget(row, 10).date().toString('dd.MM.yyyy'), '%d.%m.%Y') - datetime.datetime.strptime(whichTable.cellWidget(row, 3).date().toString('dd.MM.yyyy'), '%d.%m.%Y')).days
            whichTable.cellWidget(row, 12).setText(str(newTimeToDelivery))  

            if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo and (table == 1 or table == 2):

                try:
                    if whichTable.cellWidget(row, 8).currentText() == 'Silo' and self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 2).date() <  whichTable.cellWidget(row, 2).date():
                        self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 2).setDate(changedDate.date())  
                except:                    
                    pass                       
                                      
                
                try:
                    if whichTable.cellWidget(row, 8).currentText() == 'Homogenisierung'  and self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 2).date() <  whichTable.cellWidget(row, 2).date():
                        self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 2).setDate(changedDate.date())
                except:                    
                    pass             
                    

            whichActualLab = 0
            if table == 1 or table == 2:
                whichActualLab = whichTable.cellWidget(row, 9).currentIndex()  
            else:
                match whichTable.cellWidget(row, 9).text():
                    case 'Dichte':
                        whichActualLab = 1 
                    case 'Mechanik':
                        whichActualLab = 2 
                    case 'REACh':
                        whichActualLab = 3           

            if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')
            elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
            elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')  
            elif newTimeToDelivery < self.timeNormal:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')  
            else:
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: white')

            if whichTable.cellWidget(row, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):             
                whichTable.cellWidget(row, 3).setStyleSheet('background-color: red')
            else:
                whichTable.cellWidget(row, 3).setStyleSheet('background-color: white')           
            
            self.tableBatchesExtruder1.blockSignals(False)
            self.tableBatchesExtruder2.blockSignals(False) 
            self.tableBatchesHomogenisation.blockSignals(False)
            self.tableBatchesSilo.blockSignals(False)

    def productionEndDateChangedInTable(self, table):
        if self.workingOnShiftPlan == False:            
            
            self.saveFile.setEnabled(True)
            self.tableBatchesExtruder1.blockSignals(True)
            self.tableBatchesExtruder2.blockSignals(True)
            changedDate = self.sender()
            row = changedDate.property('row')
            newKW = datetime.datetime.strptime(changedDate.date().toString('dd.MM.yyyy'), '%d.%m.%Y').strftime('%V')       

            match table:
                case 1:            
                    whichTable = self.tableBatchesExtruder1            
                case 2:           
                    whichTable = self.tableBatchesExtruder2
                case 3:
                    whichTable = self.tableBatchesHomogenisation
                case 4:
                    whichTable = self.tableBatchesSilo
            
            newTimeToDelivery = (datetime.datetime.strptime(whichTable.cellWidget(row, 10).date().toString('dd.MM.yyyy'), '%d.%m.%Y') - datetime.datetime.strptime(changedDate.date().toString('dd.MM.yyyy'), '%d.%m.%Y')).days

            whichTable.cellWidget(row, 0).setText(newKW)

            whichShift = whichTable.cellWidget(row, 1).currentText()
            if whichShift == 'N-W-S' or whichShift == 'S-N' or whichShift == 'N-F':          

                whichTable.cellWidget(row, 2).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y') - datetime.timedelta(days=1))  
            else:
                whichTable.cellWidget(row, 2).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'))

            whichTable.cellWidget(row, 12).setText(str(newTimeToDelivery))

            if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo and (table == 1 or table == 2):

                try:
                    if whichTable.cellWidget(row, 8).currentText() == 'Silo' and self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 3).date() < whichTable.cellWidget(row, 3).date():
                        self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 3).setDate(changedDate.date())   

                except:
                    pass               
                
                try:
                    if whichTable.cellWidget(row, 8).currentText() == 'Homogenisierung'  and self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 3).date() < whichTable.cellWidget(row, 3).date():
                        self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 3).setDate(changedDate.date())
                except:
                    pass

            elif whichTable.cellWidget(row, 6).text() in self.dispoNoList and (table == 3 or table == 4):
                
                if int(self.dispoNoList[whichTable.cellWidget(row, 6).text()]) <= self.tableBatchesExtruder1.rowCount() and self.tableBatchesExtruder1.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6) and self.tableBatchesExtruder1.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                    otherTable = self.tableBatchesExtruder1
                    
                elif int(self.dispoNoList[whichTable.cellWidget(row, 6).text()]) <= self.tableBatchesExtruder2.rowCount() and self.tableBatchesExtruder2.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6) and self.tableBatchesExtruder2.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                    otherTable = self.tableBatchesExtruder2
                                                   

                if otherTable.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 3).date() > whichTable.cellWidget(row, 3).date():                    

                    dispoInUse = QMessageBox()
                    dispoInUse.setWindowTitle("Achtung!")
                    dispoInUse.setText("Das Datum für die Silierung ist vor dem Produktionsdatum!")
                    dispoInUse.exec()
 
            whichActualLab = 0
            if table == 1 or table == 2:
                whichActualLab = whichTable.cellWidget(row, 9).currentIndex()  
            else:
                match whichTable.cellWidget(row, 9).text():
                    case 'Dichte':
                        whichActualLab = 1 
                    case 'Mechanik':
                        whichActualLab = 2 
                    case 'REACh':
                        whichActualLab = 3
                    
            
            if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')
            elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
            elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
            elif newTimeToDelivery < self.timeNormal:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')       
            else:
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: white')

            if whichTable.cellWidget(row, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):             
                whichTable.cellWidget(row, 3).setStyleSheet('background-color: red')
            else:
                whichTable.cellWidget(row, 3).setStyleSheet('background-color: white')            
            
            self.tableBatchesExtruder1.blockSignals(False)
            self.tableBatchesExtruder2.blockSignals(False) 

    def deliveryDateChangedInTable(self, table):
        self.tableBatchesExtruder1.blockSignals(True)
        self.tableBatchesExtruder2.blockSignals(True)   
        if self.workingOnShiftPlan == False:  
            
            self.saveFile.setEnabled(True)  
            changedDate = self.sender()
            row = changedDate.property('row')                   

            match table:
                case 1:            
                    whichTable = self.tableBatchesExtruder1            
                case 2:           
                    whichTable = self.tableBatchesExtruder2
                case 3:
                    whichTable = self.tableBatchesHomogenisation
                case 4:
                    whichTable = self.tableBatchesSilo
            
            newTimeToDelivery = (datetime.datetime.strptime(changedDate.date().toString('dd.MM.yyyy'), '%d.%m.%Y') - datetime.datetime.strptime(whichTable.cellWidget(row, 3).date().toString('dd.MM.yyyy'), '%d.%m.%Y')).days

            whichTable.cellWidget(row, 12).setText(str(newTimeToDelivery))

            if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo and (table == 1 or table == 2):

                try:
                    if whichTable.cellWidget(row, 8).currentText() == 'Silo':
                        self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 10).setDate(changedDate.date())                    
                except:
                    pass
                
                try:    
                    if whichTable.cellWidget(row, 8).currentText() == 'Homogenisierung':
                        self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 10).setDate(changedDate.date())
                except:
                    pass
                
            whichActualLab = 0
            if table == 1 or table == 2:
                whichActualLab = whichTable.cellWidget(row, 9).currentIndex()  
            else:
                match whichTable.cellWidget(row, 9).text():
                    case 'Dichte':
                        whichActualLab = 1 
                    case 'Mechanik':
                        whichActualLab = 2 
                    case 'REACh':
                        whichActualLab = 3

            if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')
            elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
            elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
            elif newTimeToDelivery < self.timeNormal:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')       
            else:
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: white')
        
        self.tableBatchesExtruder1.blockSignals(False)
        self.tableBatchesExtruder2.blockSignals(False) 
    
    def enableSaving(self):
        self.saveFile.setEnabled(True)

    def changeDispoNo(self, table):
        if self.workingOnShiftPlan == False:
            self.saveFile.setEnabled(True)
            changedDispo = self.sender()            
            rowDispo = changedDispo.property('row') 
            oldDispo = changedDispo.text() 

            if table == 1:            
                whichTable = self.tableBatchesExtruder1
            else:           
                whichTable = self.tableBatchesExtruder2

            rx = QRegularExpression("SP\\d{1,7}[a-zA-Z]")
            msgbox = QInputDialog()
            msgbox.setInputMode(QInputDialog.InputMode.TextInput)
            msgbox.setWindowTitle('Dispo-Nr. ändern')
            msgbox.setLabelText('Neue Dispo-Nr.:') 
            msgbox.setTextValue(oldDispo)    
            msgbox.setCancelButtonText('Abbrechen')
            lineedit = msgbox.findChild(QLineEdit)
            lineedit.setValidator(QRegularExpressionValidator(rx, self))
            lineedit.setMaxLength(10)
            #msgbox.exec()
            
            if msgbox.exec() and lineedit.text() != '':
                if lineedit.text() in self.dispoNoList.keys() and lineedit.text() != oldDispo:
                    dispoInUse = QMessageBox()
                    dispoInUse.setWindowTitle("Achtung!")
                    dispoInUse.setText("Dispo-Nr. konnte nicht geändert werden. Die angegebene Dispo-Nr. ist bereits vergeben!")
                    dispoInUse.exec()

                else:
                    changedDispo.setText(lineedit.text())
                    self.dispoNoList.pop(oldDispo)
                    self.dispoNoList[lineedit.text()] = rowDispo

                    if oldDispo in self.checkDispoNoSilo and (table == 1 or table == 2):

                        if whichTable.cellWidget(rowDispo, 8).currentText() == 'Silo':
                            self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[oldDispo], 6).setText(lineedit.text()) 
                            self.checkDispoNoSilo[lineedit.text()] = self.checkDispoNoSilo[oldDispo]
                            self.checkDispoNoSilo.pop(oldDispo)
                                            
                        
                        elif whichTable.cellWidget(rowDispo, 8).currentText() == 'Homogenisierung':
                            self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[oldDispo], 6).setText(lineedit.text())
                            self.checkDispoNoSilo[lineedit.text()] = self.checkDispoNoSilo[oldDispo]
                            self.checkDispoNoSilo.pop(oldDispo) 

    def changeBatchNo(self, table):
        if self.workingOnShiftPlan == False:
            self.saveFile.setEnabled(True)
            changedBatch = self.sender()            
            rowBatch = changedBatch.property('row') 
            oldBatch = changedBatch.text() 

            if table == 1:            
                whichTable = self.tableBatchesExtruder1
                rx = QRegularExpression("1-\\d{1,2}-\\d{1,3}[a-zA-Z]")
            else:           
                whichTable = self.tableBatchesExtruder2
                rx = QRegularExpression("2-\\d{1,2}-\\d{1,3}[a-zA-Z]")

            rowCount =  whichTable.rowCount()
            batchNoInUse = []
            for row in range(rowCount):
                if (not whichTable.cellWidget(row, 5).text() in batchNoInUse) and (not len(batchNoInUse) <= 5):
                    batchNoInUse.append(whichTable.cellWidget(row, 5).text())
            
            msgbox = QInputDialog()
            msgbox.setInputMode(QInputDialog.InputMode.TextInput)
            msgbox.setWindowTitle('Chargen-Nr. ändern')
            msgbox.setLabelText('Neue Chargen-Nr.:') 
            msgbox.setTextValue(oldBatch)    
            msgbox.setCancelButtonText('Abbrechen')
            lineedit = msgbox.findChild(QLineEdit)
            lineedit.setValidator(QRegularExpressionValidator(rx, self))
            lineedit.setMaxLength(9)
            #msgbox.exec()
            
            if msgbox.exec() and lineedit.text() != '':
                if lineedit.text() in batchNoInUse and lineedit.text() != oldBatch:
                    dispoInUse = QMessageBox()
                    dispoInUse.setWindowTitle("Achtung!")
                    dispoInUse.setText("Chargen-Nr. konnte nicht geändert werden.\nDie angegebene Chargen-Nr. ist bereits vergeben!")
                    dispoInUse.exec()

                else:
                    changedBatch.setText(lineedit.text())             

                    if whichTable.cellWidget(rowBatch, 6).text() in self.checkDispoNoSilo and (table == 1 or table == 2):

                        if whichTable.cellWidget(rowBatch, 8).currentText() == 'Silo':
                            self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowBatch, 6).text()], 5).setText(lineedit.text())                      
                                            
                        
                        elif whichTable.cellWidget(rowBatch, 8).currentText() == 'Homogenisierung':                        
                            if table == 1:
                                if len(whichTable.cellWidget(rowBatch, 5).text()) == 9:
                                    newNoHomogenisation = whichTable.cellWidget(rowBatch, 5).text()[:-1][-6:] + '.1' + whichTable.cellWidget(rowBatch, 5).text()[-1]
                                else:
                                    newNoHomogenisation = whichTable.cellWidget(rowBatch, 5).text()[-6:] + '.1'
                                self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowBatch, 6).text()], 5).setText(newNoHomogenisation)
                            else:
                                if len(whichTable.cellWidget(rowBatch, 5).text()) == 9:
                                    newNoHomogenisation = whichTable.cellWidget(rowBatch, 5).text()[:-1][-6:] + '.2' + whichTable.cellWidget(rowBatch, 5).text()[-1]
                                else:
                                    newNoHomogenisation = whichTable.cellWidget(rowBatch, 5).text()[-6:] + '.2'
                                self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowBatch, 6).text()], 5).setText(newNoHomogenisation)

    def customerChanged(self, table):
        if self.workingOnShiftPlan == False:
            self.saveFile.setEnabled(True)
            changedCustomer = self.sender()            
            rowCustomer = changedCustomer.property('row') 

            if table == 1:            
                whichTable = self.tableBatchesExtruder1            
            else:           
                whichTable = self.tableBatchesExtruder2

            if whichTable.cellWidget(rowCustomer, 6).text() in self.checkDispoNoSilo and (table == 1 or table == 2):

                if whichTable.cellWidget(rowCustomer, 8).currentText() == 'Silo':
                    self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowCustomer, 6).text()], 7).setText(changedCustomer.currentText())                      
                                    
                
                elif whichTable.cellWidget(rowCustomer, 8).currentText() == 'Homogenisierung':                        
                    self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowCustomer, 6).text()], 7).setText(changedCustomer.currentText())

    def articleChanged(self, table):
        if self.workingOnShiftPlan == False and self.changeArticleNoCheck == False:
            self.saveFile.setEnabled(True)
            changedArticle = self.sender()            
            rowArticle = changedArticle.property('row')       

            if table == 1:            
                whichTable = self.tableBatchesExtruder1            
            else:           
                whichTable = self.tableBatchesExtruder2     
            
            if whichTable.cellWidget(rowArticle, 6).text() in self.checkDispoNoSilo and (table == 1 or table == 2):

                if whichTable.cellWidget(rowArticle, 8).currentText() == 'Silo':
                    
                    self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowArticle, 6).text()], 4).clear()
                    self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowArticle, 6).text()], 4).addItem(changedArticle.currentText())

                    if whichTable.cellWidget(rowArticle, 4).currentText() in self.articleNoList:                                  
                                                
                        whichIcon = self.colorList[self.dictArticleColors[whichTable.cellWidget(rowArticle, 4).currentText()]][0]+'.png'                   
                        self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowArticle, 6).text()], 4).setItemIcon(0, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))                     
                                    
                
                elif whichTable.cellWidget(rowArticle, 8).currentText() == 'Homogenisierung':                        
                    self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowArticle, 6).text()], 4).clear()
                    self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowArticle, 6).text()], 4).addItem(changedArticle.currentText())

                    if whichTable.cellWidget(rowArticle, 4).currentText() in self.articleNoList:                                  
                                                
                        whichIcon = self.colorList[self.dictArticleColors[whichTable.cellWidget(rowArticle, 4).currentText()]][0]+'.png'                   
                        self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowArticle, 6).text()], 4).setItemIcon(0, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon)))) 

            if not whichTable.cellWidget(rowArticle, 4).currentText() == '32.' and not whichTable.cellWidget(rowArticle, 4).currentText() == '':
                if (self.articleExtruderList[whichTable.cellWidget(rowArticle, 4).currentText()] == '1' and table == 1) or (self.articleExtruderList[whichTable.cellWidget(rowArticle, 4).currentText()] == '2' and table == 2):

                    invalidArticle = whichTable.cellWidget(rowArticle, 4).currentText()
                    if table == 1:
                        invalidExtruder = 'Extruder 1'
                    else:
                        invalidExtruder = 'Extruder 2'
                    messageBox = QMessageBox().warning(self, "Achtung!", "Der Artikel " + invalidArticle + " kann oder darf nicht mit " + invalidExtruder + " produziert werden!\nZum anderen Extruder verschieben?", buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)        
        
                    if messageBox == QMessageBox.StandardButton.Yes:               

                        dispoNo = whichTable.cellWidget(rowArticle, 6).text()                  
                        if table == 1:                            
                            self.moveBatchToExtruder(1)
                            otherTable = self.tableBatchesExtruder2
                        else:                            
                            self.moveBatchToExtruder(2)
                            otherTable = self.tableBatchesExtruder1

                        newRow = self.dispoNoList[dispoNo]

                        if otherTable.cellWidget(newRow, 6).text() in self.checkDispoNoSilo and (table == 1 or table == 2):

                            if otherTable.cellWidget(newRow, 8).currentText() == 'Silo':
                                self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[otherTable.cellWidget(newRow, 6).text()], 5).setText(otherTable.cellWidget(newRow, 5).text())                      
                                                
                            
                            elif otherTable.cellWidget(newRow, 8).currentText() == 'Homogenisierung':                        
                                if table == 1:
                                    newNoHomogenisation = otherTable.cellWidget(newRow, 5).text()[-6:] + '.2'
                                    self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[otherTable.cellWidget(newRow, 6).text()], 5).setText(newNoHomogenisation)
                                else:
                                    newNoHomogenisation = otherTable.cellWidget(newRow, 5).text()[-6:] + '.1'
                                    self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[otherTable.cellWidget(newRow, 6).text()], 5).setText(newNoHomogenisation)
                    else:
                        return 
                    
    def shiftChanged(self, table): 
        if self.workingOnShiftPlan == False: 
            
            self.saveFile.setEnabled(True)
            self.tableBatchesExtruder1.blockSignals(True)
            self.tableBatchesExtruder2.blockSignals(True) 
            row = self.sender().property('row')             
            whichShift = self.sender().currentText()  

            match table:
                case 1:            
                    whichTable = self.tableBatchesExtruder1            
                case 2:           
                    whichTable = self.tableBatchesExtruder2
                case 3:
                    whichTable = self.tableBatchesHomogenisation
                case 4:
                    whichTable = self.tableBatchesSilo 

            newTimeToDelivery = int(whichTable.cellWidget(row, 12).text())        

            if whichShift == 'N-W-S' or whichShift == 'S-N' or whichShift == 'N-F' or whichShift == 'N' or whichShift == 'W-S-N':         

                whichTable.cellWidget(row, 3).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y') + datetime.timedelta(days=1))  
                newTimeToDelivery = newTimeToDelivery - 1
                whichTable.cellWidget(row, 12).setText(str(newTimeToDelivery))

            else:
                whichTable.cellWidget(row, 3).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'))  

            if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo and (table == 1 or table == 2):

                if whichTable.cellWidget(row, 8).currentText() == 'Silo' and self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 3).date() < whichTable.cellWidget(row, 3).date():
                    self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 3).setDate(whichTable.cellWidget(row, 3).date())                    
                
                elif whichTable.cellWidget(row, 8).currentText() == 'Homogenisierung'  and self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 3).date() < whichTable.cellWidget(row, 3).date():
                    self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 3).setDate(whichTable.cellWidget(row, 3).date())
                    

            elif whichTable.cellWidget(row, 6).text() in self.dispoNoList and (table == 3 or table == 4):
                
                if int(self.dispoNoList[whichTable.cellWidget(row, 6).text()]) <= self.tableBatchesExtruder1.rowCount() and self.tableBatchesExtruder1.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6) and self.tableBatchesExtruder1.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                    otherTable = self.tableBatchesExtruder1
                    
                elif int(self.dispoNoList[whichTable.cellWidget(row, 6).text()]) <= self.tableBatchesExtruder2.rowCount() and self.tableBatchesExtruder2.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6) and self.tableBatchesExtruder2.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                    otherTable = self.tableBatchesExtruder2
                                                   

                if otherTable.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 3).date() > whichTable.cellWidget(row, 3).date():                    

                    dispoInUse = QMessageBox()
                    dispoInUse.setWindowTitle("Achtung!")
                    dispoInUse.setText("Das Datum für die Silierung ist vor dem Produktionsdatum!")
                    dispoInUse.exec()

            whichActualLab = 0
            if table == 1 or table == 2:
                whichActualLab = whichTable.cellWidget(row, 9).currentIndex()  
            else:
                match whichTable.cellWidget(row, 9).text():
                    case 'Dichte':
                        whichActualLab = 1 
                    case 'Mechanik':
                        whichActualLab = 2 
                    case 'REACh':
                        whichActualLab = 3          

            if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')
            elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
            elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
            elif newTimeToDelivery < self.timeNormal:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')       
            else:
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: white')


            if whichTable.cellWidget(row, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):            
                whichTable.cellWidget(row, 3).setStyleSheet('background-color: red')
            else:
                whichTable.cellWidget(row, 3).setStyleSheet('background-color: white')


            self.tableBatchesExtruder1.blockSignals(False)
            self.tableBatchesExtruder2.blockSignals(False)    

    def labChanged(self, table):            
        self.tableBatchesExtruder1.blockSignals(True)
        self.tableBatchesExtruder2.blockSignals(True)
        
        self.saveFile.setEnabled(True)

        if self.workingOnShiftPlan == False:

            whichLab = self.sender().currentIndex()
            row = self.sender().property('row')                    

            match table:
                case 1:            
                    whichTable = self.tableBatchesExtruder1            
                case 2:           
                    whichTable = self.tableBatchesExtruder2
                case 3:
                    whichTable = self.tableBatchesHomogenisation
                case 4:
                    whichTable = self.tableBatchesSilo          
            

            if whichLab == 0:     
                whichTable.cellWidget(row, 3).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y') - datetime.timedelta(days=self.timeNormal))         
            elif whichLab == 1:            
                whichTable.cellWidget(row, 3).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y') - datetime.timedelta(days=self.timeDensity))   
            elif whichLab == 2:            
                whichTable.cellWidget(row, 3).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y') - datetime.timedelta(days=self.timeMechanics))                               
            elif whichLab == 3:            
                whichTable.cellWidget(row, 3).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y') - datetime.timedelta(days=self.timeReach))

            newTimeToDelivery = (datetime.datetime.strptime(whichTable.cellWidget(row, 10).date().toString('dd.MM.yyyy'), '%d.%m.%Y') - datetime.datetime.strptime(whichTable.cellWidget(row, 3).date().toString('dd.MM.yyyy'), '%d.%m.%Y')).days
            
            newKW = datetime.datetime.strptime(whichTable.cellWidget(row, 3).date().toString('dd.MM.yyyy'), '%d.%m.%Y').strftime('%V')       
            whichTable.cellWidget(row, 0).setText(newKW)

            whichShift = whichTable.cellWidget(row, 1).currentText()

            if whichShift == 'N-W-S' or whichShift == 'S-N' or whichShift == 'N-F':          

                whichTable.cellWidget(row, 2).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y') - datetime.timedelta(days=1))  
            else:
                whichTable.cellWidget(row, 2).setDate(datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'))

            whichTable.cellWidget(row, 12).setText(str(newTimeToDelivery))

            whichActualLab = 0
            if table == 1 or table == 2:
                whichActualLab = whichTable.cellWidget(row, 9).currentIndex()  
            else:
                match whichTable.cellWidget(row, 9).text():
                    case 'Dichte':
                        whichActualLab = 1 
                    case 'Mechanik':
                        whichActualLab = 2 
                    case 'REACh':
                        whichActualLab = 3 

            if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')
            elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
            elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
            elif newTimeToDelivery < self.timeNormal:            
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')       
            else:
                whichTable.cellWidget(row, 12).setStyleSheet('background-color: white')

            if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo and (table == 1 or table == 2):

                if whichTable.cellWidget(row, 8).currentText() == 'Silo':
                    self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 9).setText(self.sender().currentText())                      
                                    
                elif whichTable.cellWidget(row, 8).currentText() == 'Homogenisierung': 
                    self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 9).setText(self.sender().currentText()) 
                    
        self.tableBatchesExtruder1.blockSignals(False)
        self.tableBatchesExtruder2.blockSignals(False) 
   
    def packagingChanged(self, table):
        if self.workingOnShiftPlan == False:        
                
            self.saveFile.setEnabled(True)  

            changedPackaging = self.sender()            
            row = changedPackaging.property('row') 
            selectedPackaging = changedPackaging.currentText() 

            if selectedPackaging == 'Homogenisierung': 

                if table == 1:            
                    whichTable = self.tableBatchesExtruder1
                    
                    if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo:
                        try:
                            self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text()                                
                        except:                                             
                            return
                        else:
                            if self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                otherTable = self.tableBatchesHomogenisation
                                silo = 1
                                                    
                            else:                            
                                return
                    else:
                        self.generateSiloLists()
                        self.sortExtruderbyDeliveryDateButton(3)
                        return
                else:           
                    whichTable = self.tableBatchesExtruder2
                    if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo:
                        try:
                            self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text()                      
                        except:
                            return
                        else:
                            if self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                otherTable = self.tableBatchesHomogenisation
                                silo = 1                            
                            else:
                                return
                    else:
                        self.generateSiloLists()
                        self.sortExtruderbyDeliveryDateButton(3)
                        return

            elif selectedPackaging == 'Silo':   

                if table == 1:            
                    whichTable = self.tableBatchesExtruder1
                    if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo:
                        try:
                            self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text()                
                        except:
                            return
                        else:
                            if self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                otherTable = self.tableBatchesSilo
                                silo = 0                            
                            else:
                                return
                    else:
                        self.generateSiloLists()
                        self.sortExtruderbyDeliveryDateButton(4)
                        return
                else:           
                    whichTable = self.tableBatchesExtruder2
                    if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo:
                        try:
                            self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text()                      
                        except:
                            return
                        else:
                            if self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                otherTable = self.tableBatchesSilo                            
                                silo = 0
                            else:
                                return
                    else:
                        self.generateSiloLists()
                        self.sortExtruderbyDeliveryDateButton(4)
                        return
                            
            elif selectedPackaging == 'Bigbag':   

                if table == 1:     
                    whichTable = self.tableBatchesExtruder1
                    if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo:
                        try:
                            self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text()
                            if self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():                                                     
                                silo = 3                            
                            else:                                
                                raise Exception('wrong table')                                        
                        except:
                            try:
                                self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text()                                                              
                            except:                                                                             
                                return
                            else:
                                if self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                    otherTable = self.tableBatchesHomogenisation                                    
                                    silo = 2                      
                            
                    else:                        
                        return
                else:           
                    whichTable = self.tableBatchesExtruder2
                    if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo:
                        try:
                            self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() 
                            if self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():                            
                                silo = 3                            
                            else:
                                raise Exception('wrong table')                     
                        except:
                            try:
                                self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text()                                
                            except:                                             
                                return
                            else:
                                if self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                    otherTable = self.tableBatchesHomogenisation
                                    silo = 2                      
                            
                    else:
                        return
            elif selectedPackaging == 'Oktabin':   

                if table == 1:            
                    whichTable = self.tableBatchesExtruder1
                    if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo:
                        try:
                            self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() 
                            if self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():                            
                                silo = 5                            
                            else:
                                raise Exception('wrong table')              
                        except:
                            try:
                                self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text()                                
                            except:                                             
                                return
                            else:
                                if self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                    otherTable = self.tableBatchesHomogenisation
                                    silo = 4                     
                            
                    else:
                        return
                else:           
                    whichTable = self.tableBatchesExtruder2
                    if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo:
                        try:
                            self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text()
                            if self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():                            
                                silo = 5                            
                            else:
                                raise Exception('wrong table')                     
                        except:
                            try:
                                self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text()                                
                            except:                                             
                                return
                            else:
                                if self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                    otherTable = self.tableBatchesHomogenisation
                                    silo = 4                      
                            
                    else:
                        return                 
            else:
                return 
                        
            if (silo == 1 and selectedPackaging == 'Homogenisierung'):

                dispoNo = whichTable.cellWidget(row, 4).currentText() 

                messageBox = QMessageBox().warning(self, "Silo geändert!", "Die Charge " + dispoNo + " ist aktuell eine Siloabholung.\nSoll sie stattdessen homogenisiert werden?", buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)        
        
                if not messageBox == QMessageBox.StandardButton.Yes:   
                    whichTable.cellWidget(row, 8).setCurrentText('Silo')             
                    return
                
            if (silo == 0 and selectedPackaging == 'Silo'):                   
                dispoNo = whichTable.cellWidget(row, 4).currentText() 

                messageBox = QMessageBox().warning(self, "Silo geändert!", "Die Charge " + dispoNo + " wird aktuell homogenisiert.\nSoll sie stattdessen siliert werden?", buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)        
        
                if not messageBox == QMessageBox.StandardButton.Yes: 
                    whichTable.cellWidget(row, 8).setCurrentText('Homogenisierung')               
                    return 
                
            if (silo == 2 and selectedPackaging == 'Bigbag'):                   
                dispoNo = whichTable.cellWidget(row, 4).currentText() 

                messageBox = QMessageBox().warning(self, "Silo geändert!", "Die Charge " + dispoNo + " ist aktuell eine Siloabholung.\nSoll sie stattdessen im Bigbag ausgeliefert werden?", buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)        
        
                if messageBox == QMessageBox.StandardButton.Yes: 
                    self.tableBatchesSilo.removeRow(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()])  
                    self.checkDispoNoSilo.pop(whichTable.cellWidget(row, 6).text())      

                    for row in range(self.tableBatchesSilo.rowCount()):
                        self.tableBatchesSilo.cellWidget(row, 1).setProperty('row', row) 
                        self.tableBatchesSilo.cellWidget(row, 2).setProperty('row', row)           
                        self.tableBatchesSilo.cellWidget(row, 3).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 4).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 5).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 6).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 7).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 8).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 9).setProperty('row', row) 
                        self.tableBatchesSilo.cellWidget(row, 10).setProperty('row', row)  
                        self.tableBatchesSilo.cellWidget(row, 11).setProperty('row', row)            
                        self.checkDispoNoSilo[self.tableBatchesSilo.cellWidget(row, 6).text()] = row             
                    return
                else:
                    whichTable.cellWidget(row, 8).setCurrentText('Silo')
                    return 
                
            if (silo == 3 and selectedPackaging == 'Bigbag'):                   
                dispoNo = whichTable.cellWidget(row, 4).currentText() 

                messageBox = QMessageBox().warning(self, "Silo geändert!", "Die Charge " + dispoNo + " wird aktuell homogenisiert.\nSoll sie stattdessen im Bigbag ausgeliefert werden?", buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)        
        
                if messageBox == QMessageBox.StandardButton.Yes: 
                    self.tableBatchesHomogenisation.removeRow(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()])             
                    self.checkDispoNoSilo.pop(whichTable.cellWidget(row, 6).text()) 

                    for row in range(self.tableBatchesHomogenisation.rowCount()):                    
                        self.tableBatchesHomogenisation.cellWidget(row, 1).setProperty('row', row) 
                        self.tableBatchesHomogenisation.cellWidget(row, 2).setProperty('row', row)           
                        self.tableBatchesHomogenisation.cellWidget(row, 3).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 4).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 5).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 6).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 7).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 8).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 9).setProperty('row', row) 
                        self.tableBatchesHomogenisation.cellWidget(row, 10).setProperty('row', row) 
                        self.tableBatchesHomogenisation.cellWidget(row, 11).setProperty('row', row)           
                        self.checkDispoNoSilo[self.tableBatchesHomogenisation.cellWidget(row, 6).text()] = row          
                    return
                else:
                    whichTable.cellWidget(row, 8).setCurrentText('Homogenisierung')
                    return
            if (silo == 4 and selectedPackaging == 'Oktabin'):                   
                dispoNo = whichTable.cellWidget(row, 4).currentText() 

                messageBox = QMessageBox().warning(self, "Silo geändert!", "Die Charge " + dispoNo + " ist aktuell eine Siloabholung.\nSoll sie stattdessen im Oktabin ausgeliefert werden?", buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)        
        
                if messageBox == QMessageBox.StandardButton.Yes: 
                    self.tableBatchesSilo.removeRow(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()])  
                    self.checkDispoNoSilo.pop(whichTable.cellWidget(row, 6).text())      

                    for row in range(self.tableBatchesSilo.rowCount()):
                        self.tableBatchesSilo.cellWidget(row, 1).setProperty('row', row) 
                        self.tableBatchesSilo.cellWidget(row, 2).setProperty('row', row)           
                        self.tableBatchesSilo.cellWidget(row, 3).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 4).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 5).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 6).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 7).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 8).setProperty('row', row)
                        self.tableBatchesSilo.cellWidget(row, 9).setProperty('row', row) 
                        self.tableBatchesSilo.cellWidget(row, 10).setProperty('row', row) 
                        self.tableBatchesSilo.cellWidget(row, 11).setProperty('row', row)            
                        self.checkDispoNoSilo[self.tableBatchesSilo.cellWidget(row, 6).text()] = row             
                    return
                else:
                    whichTable.cellWidget(row, 8).setCurrentText('Silo')
                    return 
                
            if (silo == 5 and selectedPackaging == 'Oktabin'):                   
                dispoNo = whichTable.cellWidget(row, 4).currentText() 

                messageBox = QMessageBox().warning(self, "Silo geändert!", "Die Charge " + dispoNo + " wird aktuell homogenisiert.\nSoll sie stattdessen im Oktabin ausgeliefert werden?", buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)        
        
                if messageBox == QMessageBox.StandardButton.Yes: 
                    self.tableBatchesHomogenisation.removeRow(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()])             
                    self.checkDispoNoSilo.pop(whichTable.cellWidget(row, 6).text()) 

                    for row in range(self.tableBatchesHomogenisation.rowCount()):                    
                        self.tableBatchesHomogenisation.cellWidget(row, 1).setProperty('row', row) 
                        self.tableBatchesHomogenisation.cellWidget(row, 2).setProperty('row', row)           
                        self.tableBatchesHomogenisation.cellWidget(row, 3).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 4).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 5).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 6).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 7).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 8).setProperty('row', row)
                        self.tableBatchesHomogenisation.cellWidget(row, 9).setProperty('row', row) 
                        self.tableBatchesHomogenisation.cellWidget(row, 10).setProperty('row', row)  
                        self.tableBatchesHomogenisation.cellWidget(row, 11).setProperty('row', row)          
                        self.checkDispoNoSilo[self.tableBatchesHomogenisation.cellWidget(row, 6).text()] = row          
                    return
                else:
                    whichTable.cellWidget(row, 8).setCurrentText('Homogenisierung')
                    return
                    
            rowPosition = otherTable.rowCount()
            otherTable.insertRow(rowPosition)                       

            attrSiloDelivery = ['Silo', 'Dino'] 

            rowItem = 0
            for rowItem in range(14):             
                match rowItem:                    
                    case 0:                    
                        whichCalendarWeek = QLabel()
                        whichCalendarWeek.setText(whichTable.cellWidget(row, rowItem).text())
                        whichCalendarWeek.setFixedWidth(40)
                        whichCalendarWeek.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        otherTable.setCellWidget(rowPosition, rowItem, whichCalendarWeek)

                    case 1:                                            
                        whichShift = QComboBox()
                        whichShift.addItems(self.attrShift)
                        whichShift.setCurrentIndex(0)
                        if silo == 0:
                            whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(4))
                        else:
                            whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(3))                                              
                        otherTable.setCellWidget(rowPosition, rowItem, whichShift)                       

                    case 2:                                  
                        buttonProductionDate = QDateEdit()
                        buttonProductionDate.setFixedWidth(80)
                        buttonProductionDate.setDate(datetime.datetime.strptime(whichTable.cellWidget(row, rowItem+1).text(), '%d.%m.%Y'))                     
                        if silo == 0:       
                            buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(4)) 
                        else: 
                            buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(3))                                            
                        otherTable.setCellWidget(rowPosition, rowItem, buttonProductionDate) 

                    case 3:                                  
                        buttonProductionDate = QDateEdit()
                        buttonProductionDate.setFixedWidth(80)
                        buttonProductionDate.setDate(datetime.datetime.strptime(whichTable.cellWidget(row, rowItem).text(), '%d.%m.%Y'))                    
                        if silo == 0:       
                            buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(4)) 
                        else: 
                            buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(3))               
                        otherTable.setCellWidget(rowPosition, rowItem, buttonProductionDate)                            

                    case 4:                                
                        whichArticle = QComboBox()  
                        whichArticle.addItem(whichTable.cellWidget(row, rowItem).currentText())                                      
                        whichArticle.setCurrentText(whichTable.cellWidget(row, rowItem).currentText())                                   
                        if whichTable.cellWidget(row, rowItem).currentText() in self.articleNoList:                                    
                            
                            whichIcon = self.colorList[self.dictArticleColors[whichTable.cellWidget(row, rowItem).currentText()]][0]+'.png'                   
                            whichArticle.setItemIcon(0, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))
                    
                        otherTable.setCellWidget(rowPosition, rowItem, whichArticle)
                        
                    case 5:                                         
                        newBatchNo = QLabel()
                        if silo == 0:
                            newBatchNo.setText(whichTable.cellWidget(row, rowItem).text()) 
                        else:                                     
                            if table == 1:                                
                                if len(whichTable.cellWidget(row, rowItem).text()) == 9:                                    
                                    newNoHomogenisation = whichTable.cellWidget(row, rowItem).text()[:-1][-6:] + '.1' + whichTable.cellWidget(row, rowItem).text()[-1]
                                else:                                    
                                    newNoHomogenisation = whichTable.cellWidget(row, rowItem).text()[-6:] + '.1'
                                newBatchNo.setText(str(newNoHomogenisation))
                            else:                                
                                if len(whichTable.cellWidget(row, rowItem).text()) == 9:
                                    newNoHomogenisation = whichTable.cellWidget(row, rowItem).text()[:-1][-6:] + '.2' + whichTable.cellWidget(row, rowItem).text()[-1]
                                else:
                                    newNoHomogenisation = whichTable.cellWidget(row, rowItem).text()[-6:] + '.2'
                                newBatchNo.setText(str(newNoHomogenisation))
                                
                        newBatchNo.setFixedWidth(100)
                        newBatchNo.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        otherTable.setCellWidget(rowPosition, rowItem, newBatchNo) 

                    case 6:                                         
                        newDispo = QLabel()
                        newDispo.setText(whichTable.cellWidget(row, rowItem).text())
                        newDispo.setFixedWidth(100)
                        newDispo.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        otherTable.setCellWidget(rowPosition, rowItem, newDispo)    

                    case 7:                                
                        whichCustomer = QLabel()                                                                        
                        whichCustomer.setText(whichTable.cellWidget(row, rowItem).currentText())                                  
                        otherTable.setCellWidget(rowPosition, rowItem, whichCustomer) 

                    case 8: 
                        if silo == 0:
                            whichPackaging = QComboBox()
                            whichPackaging.addItems(attrSiloDelivery)  
                            whichPackaging.setCurrentIndex(0)                                      
                        else:
                            whichPackaging = QLabel()
                            whichPackaging.setText('Homogenisierung')                        
                            whichPackaging.setAlignment(Qt.AlignmentFlag.AlignCenter)
                            
                        otherTable.setCellWidget(rowPosition, rowItem, whichPackaging)      

                    case 9:               
                        whichLab = QLabel()                                        
                        whichLab.setText(whichTable.cellWidget(row, rowItem).currentText())
                        whichLab.setFixedWidth(80)
                        whichLab.setAlignment(Qt.AlignmentFlag.AlignCenter)                                                              
                        otherTable.setCellWidget(rowPosition, rowItem, whichLab)                           
                            
                    case 10:                                  
                        buttonDeliveryDate = QDateEdit()
                        buttonDeliveryDate.setFixedWidth(80)
                        buttonDeliveryDate.setDate(datetime.datetime.strptime(whichTable.cellWidget(row, rowItem).text(), '%d.%m.%Y')) 
                        if silo == 0:       
                            buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(4)) 
                        else: 
                            buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(3))        
                        buttonDeliveryDate.setDisabled(True)       
                        otherTable.setCellWidget(rowPosition, rowItem, buttonDeliveryDate)

                    case 11:
                        whichBatchSize = QLabel()
                        whichBatchSize.setText(whichTable.cellWidget(row, rowItem).text())
                        whichBatchSize.setEnabled(True)
                        whichBatchSize.setFixedWidth(38)
                        whichBatchSize.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        otherTable.setCellWidget(rowPosition, rowItem, whichBatchSize)

                    case 12:
                        whichDeliveryDate = QLabel()
                        whichDeliveryDate.setText(whichTable.cellWidget(row, rowItem).text())                                        
                        whichDeliveryDate.setFixedWidth(38)
                        whichDeliveryDate.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        otherTable.setCellWidget(rowPosition, rowItem, whichDeliveryDate) 
                        
                        newTimeToDelivery = int(otherTable.cellWidget(rowPosition, rowItem).text() ) 

                        whichActualLab = 0
                        match otherTable.cellWidget(rowPosition, 9).text():
                            case 'Dichte':
                                whichActualLab = 1 
                            case 'Mechanik':
                                whichActualLab = 2 
                            case 'REACh':
                                whichActualLab = 3 

                        if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                            otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')
                        elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                            otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                        elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                            otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                        elif newTimeToDelivery < self.timeNormal:            
                            otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')       
                        else:
                            otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: white')    

                        if otherTable.cellWidget(rowPosition, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):             
                            otherTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: red')
                        else:
                            otherTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: white')                    
                        
                    case 13:
                        whichComment = QLineEdit() 
                        whichComment.setText(whichTable.cellWidget(row, rowItem).text())
                        whichComment.setFixedWidth(350)
                        otherTable.setCellWidget(rowPosition, rowItem, whichComment)          
            
            if silo == 1:
                self.tableBatchesSilo.removeRow(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()])  
                self.sortExtruderbyDeliveryDateButton(3)             
            else:
                self.tableBatchesHomogenisation.removeRow(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()]) 
                self.sortExtruderbyDeliveryDateButton(4)   
            
            for row in range(self.tableBatchesHomogenisation.rowCount()):                    
                self.tableBatchesHomogenisation.cellWidget(row, 1).setProperty('row', row) 
                self.tableBatchesHomogenisation.cellWidget(row, 2).setProperty('row', row)           
                self.tableBatchesHomogenisation.cellWidget(row, 3).setProperty('row', row)
                self.tableBatchesHomogenisation.cellWidget(row, 4).setProperty('row', row)
                self.tableBatchesHomogenisation.cellWidget(row, 5).setProperty('row', row)
                self.tableBatchesHomogenisation.cellWidget(row, 6).setProperty('row', row)
                self.tableBatchesHomogenisation.cellWidget(row, 7).setProperty('row', row)
                self.tableBatchesHomogenisation.cellWidget(row, 8).setProperty('row', row)
                self.tableBatchesHomogenisation.cellWidget(row, 9).setProperty('row', row) 
                self.tableBatchesHomogenisation.cellWidget(row, 10).setProperty('row', row) 
                self.tableBatchesHomogenisation.cellWidget(row, 11).setProperty('row', row)            
                self.checkDispoNoSilo[self.tableBatchesHomogenisation.cellWidget(row, 6).text()] = row
            for row in range(self.tableBatchesSilo.rowCount()):
                self.tableBatchesSilo.cellWidget(row, 1).setProperty('row', row) 
                self.tableBatchesSilo.cellWidget(row, 2).setProperty('row', row)           
                self.tableBatchesSilo.cellWidget(row, 3).setProperty('row', row)
                self.tableBatchesSilo.cellWidget(row, 4).setProperty('row', row)
                self.tableBatchesSilo.cellWidget(row, 5).setProperty('row', row)
                self.tableBatchesSilo.cellWidget(row, 6).setProperty('row', row)
                self.tableBatchesSilo.cellWidget(row, 7).setProperty('row', row)
                self.tableBatchesSilo.cellWidget(row, 8).setProperty('row', row)
                self.tableBatchesSilo.cellWidget(row, 9).setProperty('row', row) 
                self.tableBatchesSilo.cellWidget(row, 10).setProperty('row', row)   
                self.tableBatchesSilo.cellWidget(row, 11).setProperty('row', row)          
                self.checkDispoNoSilo[self.tableBatchesSilo.cellWidget(row, 6).text()] = row  

    def batchSizeChanged(self, table):
        if self.workingOnShiftPlan == False:
            self.saveFile.setEnabled(True)
            changedBatchSize = self.sender()            
            rowBatchSize = changedBatchSize.property('row') 

            if table == 1:            
                whichTable = self.tableBatchesExtruder1            
            else:           
                whichTable = self.tableBatchesExtruder2

            if whichTable.cellWidget(rowBatchSize, 6).text() in self.checkDispoNoSilo and (table == 1 or table == 2):

                if whichTable.cellWidget(rowBatchSize, 8).currentText() == 'Silo':
                    self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowBatchSize, 6).text()], 11).setText(changedBatchSize.text())                      
                                    
                
                elif whichTable.cellWidget(rowBatchSize, 8).currentText() == 'Homogenisierung':                        
                    self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(rowBatchSize, 6).text()], 11).setText(changedBatchSize.text())
     
    def sortExtruderbyDeliveryDateButton(self, table):
        
        self.saveFile.setEnabled(True)

        self.workingOnShiftPlan = True
        match table:
                case 1:            
                    whichTable = self.tableBatchesExtruder1            
                case 2:           
                    whichTable = self.tableBatchesExtruder2
                case 3:
                    whichTable = self.tableBatchesHomogenisation
                case 4:
                    whichTable = self.tableBatchesSilo
        

        saveTableDataHelp = {}
        
        for row in range(whichTable.rowCount()): 
            if table == 1 or table == 2:            
                saveTableDataHelp[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentIndex(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).currentIndex(), whichTable.cellWidget(row, 8).currentIndex(), whichTable.cellWidget(row, 9).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
            elif table == 4:
                saveTableDataHelp[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentIndex(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).currentIndex(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
            else:
                saveTableDataHelp[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentIndex(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).text(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]

        saveTableData = dict(sorted(saveTableDataHelp.items(), key=lambda item: item[1][self.sortByColumn]))    
          
        row = 0
        for key in saveTableData:  
            
            for rowItem in range(14):               
                
                match rowItem:
                    case 0:                  

                        whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

                    case 1:                                           
                        
                        whichTable.cellWidget(row, rowItem).setProperty('row', row)                          
                        whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])                      

                    case 2:
                        whichTable.cellWidget(row, rowItem).setProperty('row', row) 
                        whichTable.cellWidget(row, rowItem).setDate(saveTableData[key][rowItem])

                    case 3:                                  
                         
                        whichTable.cellWidget(row, rowItem).setProperty('row', row)                  
                        whichTable.cellWidget(row, rowItem).setDate(saveTableData[key][rowItem])  

                        if whichTable.cellWidget(row, rowItem).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):            
                            whichTable.cellWidget(row, rowItem).setStyleSheet('background-color: red')
                        else:
                            whichTable.cellWidget(row, rowItem).setStyleSheet('background-color: white')                      

                    case 4:                               
                        whichTable.cellWidget(row, rowItem).setProperty('row', row)
                        whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])
                            
                    case 5:                                       
                        whichTable.cellWidget(row, rowItem).setProperty('row', row)
                        whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])
                        

                    case 6:                                         
                        whichTable.cellWidget(row, rowItem).setProperty('row', row)
                        whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])      

                    case 7:                                
                        if table == 1 or table == 2:                                                        
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)
                            whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])    
                        else:
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem]) 

                    case 8:                
                        if table == 1 or table == 2 or table == 4:
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)   
                            whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])
                        else:
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

   
                    case 9: 
                        if table == 1 or table == 2:    
                            whichTable.cellWidget(row, rowItem).setProperty('row', row) 
                            whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])
                        else:
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])
                        
                                                          
                    case 10:                                  
                           
                        whichTable.cellWidget(row, rowItem).setProperty('row', row)            
                        whichTable.cellWidget(row, rowItem).setDate(saveTableData[key][rowItem])

                    case 11:
                        whichTable.cellWidget(row, rowItem).setProperty('row', row)
                        whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

                    case 12:
                        
                        whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

                        newTimeToDelivery = int(saveTableData[key][rowItem])

                        whichActualLab = 0
                        if table == 1 or table == 2:
                            whichActualLab = whichTable.cellWidget(row, 9).currentIndex()  
                        else:
                            match whichTable.cellWidget(row, 9).text():
                                case 'Dichte':
                                    whichActualLab = 1 
                                case 'Mechanik':
                                    whichActualLab = 2 
                                case 'REACh':
                                    whichActualLab = 3
                        
                        if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                            whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')
                        elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                            whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
                        elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                            whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
                        elif newTimeToDelivery < self.timeNormal:            
                            whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')       
                        else:
                            whichTable.cellWidget(row, 12).setStyleSheet('background-color: white')

                        if whichTable.cellWidget(row, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):             
                            whichTable.cellWidget(row, 3).setStyleSheet('background-color: red')
                        else:
                            whichTable.cellWidget(row, 3).setStyleSheet('background-color: white')
                    case 13:
                        whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])
            
            if table == 1 or table == 2:
                self.dispoNoList[whichTable.cellWidget(row, 6).text()] = row 
            
            elif table == 3 or table == 4:
                self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()] = row    
                                 
            row = row + 1
        self.workingOnShiftPlan = False

    def moveBatchRowUp(self, table):
        self.workingOnShiftPlan = True
        
        self.saveFile.setEnabled(True)

        match table:
            case 1:            
                whichTable = self.tableBatchesExtruder1            
            case 2:           
                whichTable = self.tableBatchesExtruder2
            case 3:
                whichTable = self.tableBatchesHomogenisation
            case 4:
                whichTable = self.tableBatchesSilo

        if len(whichTable.selectionModel().selectedRows()) != 0:            

            rowList = []            
            for rowTable in whichTable.selectionModel().selectedRows():
                rowList.append(rowTable.row())

            rowList.sort()
            
            rowCount = whichTable.rowCount()

            firstRowtoOverwrite = rowList[0] - 1

            if firstRowtoOverwrite < 0:
                firstRowtoOverwrite = 0

            saveTableDataSelected = {}
            saveTableDataNotSelected = {}

            for row in range(rowCount): 
                if row in rowList:
                    if table == 1 or table == 2:            
                        saveTableDataSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentIndex(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).currentIndex(), whichTable.cellWidget(row, 8).currentIndex(), whichTable.cellWidget(row, 9).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                    elif table == 4:
                        saveTableDataSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).currentIndex(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                    else:
                        saveTableDataSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).text(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                        
                else:
                    if row > firstRowtoOverwrite or row == firstRowtoOverwrite:
                        if table == 1 or table == 2:            
                            saveTableDataNotSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentIndex(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).currentIndex(), whichTable.cellWidget(row, 8).currentIndex(), whichTable.cellWidget(row, 9).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                        elif table == 4:
                            saveTableDataNotSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).currentIndex(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                        else:
                            saveTableDataNotSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).text(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]

            saveTableData = saveTableDataSelected | saveTableDataNotSelected    
            
            whichTable.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
            whichTable.clearSelection()
            row = firstRowtoOverwrite
            selectRows = firstRowtoOverwrite + len(rowList) - 1
            for key in saveTableData:

                if row <= selectRows:                                       
                    whichTable.selectRow(row)                    

                for rowItem in range(14):                 
                    
                    match rowItem:
                        case 0:                  

                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

                        case 1:                                           
                            
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)                          
                            whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])                      

                        case 2:

                            whichTable.cellWidget(row, rowItem).setProperty('row', row) 
                            whichTable.cellWidget(row, rowItem).setDate(saveTableData[key][rowItem])

                        case 3:                                  
                            
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)                  
                            whichTable.cellWidget(row, rowItem).setDate(saveTableData[key][rowItem])

                        case 4: 

                            if table == 1 or table == 2:
                                whichTable.cellWidget(row, rowItem).setProperty('row', row)
                                whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])
                            else:   
                                whichTable.cellWidget(row, rowItem).clear()                             
                                whichTable.cellWidget(row, rowItem).addItem(saveTableData[key][rowItem])                                      
                                whichTable.cellWidget(row, rowItem).setCurrentText(saveTableData[key][rowItem])                                   
                                if saveTableData[key][rowItem] in self.articleNoList:                                    
                                                
                                    whichIcon = self.colorList[self.dictArticleColors[saveTableData[key][rowItem]]][0]+'.png'                   
                                    whichTable.cellWidget(row, rowItem).setItemIcon(0, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))
   
                                
                        case 5:                                         
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])
                            

                        case 6:                                         
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])      

                        case 7:                                
                            if table == 1 or table == 2:                                
                                whichTable.cellWidget(row, rowItem).setProperty('row', row)                             
                                whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])    
                            else:
                                whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem]) 

                        case 8:                
                            if table == 1 or table == 2 or table == 4:
                                whichTable.cellWidget(row, rowItem).setProperty('row', row)  
                                whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])
                            else:
                                whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

    
                        case 9: 
                            if table == 1 or table == 2:    
                                whichTable.cellWidget(row, rowItem).setProperty('row', row) 
                                whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])
                            else:
                                whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])
                                                            
                        case 10:                                  
                            
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)            
                            whichTable.cellWidget(row, rowItem).setDate(saveTableData[key][rowItem])

                        case 11:
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

                        case 12:
                            
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

                            newTimeToDelivery = int(saveTableData[key][rowItem])

                            whichActualLab = 0
                            if table == 1 or table == 2:
                                whichActualLab = whichTable.cellWidget(row, 9).currentIndex()  
                            else:
                                match whichTable.cellWidget(row, 9).text():
                                    case 'Dichte':
                                        whichActualLab = 1 
                                    case 'Mechanik':
                                        whichActualLab = 2 
                                    case 'REACh':
                                        whichActualLab = 3
                            
                            if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')
                            elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
                            elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
                            elif newTimeToDelivery < self.timeNormal:            
                                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')       
                            else:
                                whichTable.cellWidget(row, 12).setStyleSheet('background-color: white')

                            if whichTable.cellWidget(row, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):             
                                whichTable.cellWidget(row, 3).setStyleSheet('background-color: red')
                            else:
                                whichTable.cellWidget(row, 3).setStyleSheet('background-color: white')

                        case 13:
                                                     
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem]) 

                if table == 1 or table == 2:
                    self.dispoNoList[whichTable.cellWidget(row, 6).text()] = row 
                elif table == 3 or table == 4:
                    self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()] = row    
                 
                row = row + 1
            whichTable.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)

        self.workingOnShiftPlan = False

    def moveBatchRowDown(self, table):
        self.workingOnShiftPlan = True

        self.saveFile.setEnabled(True)

        match table:
            case 1:            
                whichTable = self.tableBatchesExtruder1            
            case 2:           
                whichTable = self.tableBatchesExtruder2
            case 3:
                whichTable = self.tableBatchesHomogenisation
            case 4:
                whichTable = self.tableBatchesSilo

        if len(whichTable.selectionModel().selectedRows()) != 0:            

            rowList = []            
            for rowTable in whichTable.selectionModel().selectedRows():
                rowList.append(rowTable.row())

            rowList.sort()
                
            rowCount = whichTable.rowCount()

            firstRowtoOverwrite = rowList[0]
            if len(rowList) < 2:
                lastRowtoOverwrite = firstRowtoOverwrite + 1
            else:
                lastRowtoOverwrite = rowList[-1] + 1

            if firstRowtoOverwrite < 0:
                firstRowtoOverwrite = 0

            saveTableDataSelected = {}
            saveTableDataNotSelected = {}

            for row in range(rowCount): 
                if row in rowList:
                    if table == 1 or table == 2:            
                        saveTableDataSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentIndex(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).currentIndex(), whichTable.cellWidget(row, 8).currentIndex(), whichTable.cellWidget(row, 9).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                    elif table == 4:
                        saveTableDataSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).currentIndex(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                    else:
                        saveTableDataSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).text(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                else:
                    if row >= firstRowtoOverwrite and row <= lastRowtoOverwrite:
                        if table == 1 or table == 2:            
                            saveTableDataNotSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentIndex(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).currentIndex(), whichTable.cellWidget(row, 8).currentIndex(), whichTable.cellWidget(row, 9).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                        elif table == 4:
                            saveTableDataNotSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).currentIndex(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                        else:
                            saveTableDataNotSelected[row] = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentIndex(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y'), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).text(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y'), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                        
            saveTableData = saveTableDataNotSelected | saveTableDataSelected    
                
            whichTable.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
            whichTable.clearSelection()
            row = firstRowtoOverwrite
            selectRows = lastRowtoOverwrite - len(rowList) + 1
            for key in saveTableData:

                if row >= selectRows and row <= lastRowtoOverwrite:
                    whichTable.selectRow(row)  

                for rowItem in range(14):                 
                        
                    match rowItem:
                        case 0:                  

                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

                        case 1:                                           
                                
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)                          
                            whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])                      

                        case 2:

                            whichTable.cellWidget(row, rowItem).setProperty('row', row) 
                            whichTable.cellWidget(row, rowItem).setDate(saveTableData[key][rowItem])

                        case 3:                                  
                                
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)                  
                            whichTable.cellWidget(row, rowItem).setDate(saveTableData[key][rowItem])

                        case 4:                                
                                
                            if table == 1 or table == 2:
                                whichTable.cellWidget(row, rowItem).setProperty('row', row)
                                whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])
                            else:   
                                whichTable.cellWidget(row, rowItem).clear()                             
                                whichTable.cellWidget(row, rowItem).addItem(saveTableData[key][rowItem])                                      
                                whichTable.cellWidget(row, rowItem).setCurrentText(saveTableData[key][rowItem])                                   
                                if saveTableData[key][rowItem] in self.articleNoList:                                    
                                                
                                    whichIcon = self.colorList[self.dictArticleColors[saveTableData[key][rowItem]]][0]+'.png'                   
                                    whichTable.cellWidget(row, rowItem).setItemIcon(0, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))
                                    
                        case 5:                                         
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])
                                

                        case 6:                                         
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)    
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])      

                        case 7:                                
                            if table == 1 or table == 2:  
                                whichTable.cellWidget(row, rowItem).setProperty('row', row)                             
                                whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])    
                            else:
                                whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem]) 

                        case 8:                
                            if table == 1 or table == 2 or table == 4: 
                                whichTable.cellWidget(row, rowItem).setProperty('row', row)  
                                whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])
                            else:
                                whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

    
                        case 9: 
                            if table == 1 or table == 2:    
                                whichTable.cellWidget(row, rowItem).setProperty('row', row) 
                                whichTable.cellWidget(row, rowItem).setCurrentIndex(saveTableData[key][rowItem])
                            else:
                                whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])
                                                                
                        case 10:                                  
                                
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)            
                            whichTable.cellWidget(row, rowItem).setDate(saveTableData[key][rowItem])

                        case 11:
                            whichTable.cellWidget(row, rowItem).setProperty('row', row)
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

                        case 12:
                                
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])

                            newTimeToDelivery = int(saveTableData[key][rowItem])

                            whichActualLab = 0
                            if table == 1 or table == 2:
                                whichActualLab = whichTable.cellWidget(row, 9).currentIndex()  
                            else:
                                match whichTable.cellWidget(row, 9).text():
                                    case 'Dichte':
                                        whichActualLab = 1 
                                    case 'Mechanik':
                                        whichActualLab = 2 
                                    case 'REACh':
                                        whichActualLab = 3
                            
                            if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')
                            elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
                            elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red') 
                            elif newTimeToDelivery < self.timeNormal:            
                                whichTable.cellWidget(row, 12).setStyleSheet('background-color: red')       
                            else:
                                whichTable.cellWidget(row, 12).setStyleSheet('background-color: white')

                            if whichTable.cellWidget(row, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):             
                                whichTable.cellWidget(row, 3).setStyleSheet('background-color: red')
                            else:
                                whichTable.cellWidget(row, 3).setStyleSheet('background-color: white')
                        
                        case 13:
                            
                            whichTable.cellWidget(row, rowItem).setText(saveTableData[key][rowItem])  
                if table == 1 or table == 2:
                    self.dispoNoList[whichTable.cellWidget(row, 6).text()] = row 
                
                elif table == 3 or table == 4:
                    self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()] = row   
                        
                row = row + 1
            whichTable.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.workingOnShiftPlan = False

    def createShiftPlan(self, table):
        self.workingOnShiftPlan = True

        self.saveFile.setEnabled(True)

        match self.maintenanceDay:
            case 0:
                maintenanceShiftOn = 'Monday'
            case 1:
                maintenanceShiftOn = 'Tuesday'
            case 2:
                maintenanceShiftOn = 'Wednesday'
            case 3:
                maintenanceShiftOn = 'Thursday'
            case 4:
                maintenanceShiftOn = 'Friday'

        match table:
            case 1:            
                whichTable = self.tableBatchesExtruder1            
            case 2:           
                whichTable = self.tableBatchesExtruder2
            case 3:
                whichTable = self.tableBatchesHomogenisation
            case 4:
                whichTable = self.tableBatchesSilo

        if whichTable.rowCount() != 0:  

            nextShift = whichTable.cellWidget(0, 1).currentIndex() 
            nextStartDay = whichTable.cellWidget(0, 3).text() 
            if table == 4:
                if whichTable.rowCount() <= 1:
                    nextDelivery = whichTable.cellWidget(0, 8).currentIndex()
                else:
                    nextDelivery = whichTable.cellWidget(1, 8).currentIndex()
               
        for row in range(whichTable.rowCount()):
            
            if table == 1 or table ==2:

                match nextShift:
                    case 0:
                        if (row + 1) < whichTable.rowCount():

                            if int(float(whichTable.cellWidget(row+1, 11).text() or 24)) <= 12:
                                whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                                whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))

                                nextStartDay = whichTable.cellWidget(row+1, 3).text()                         

                                whichTable.cellWidget(row+1, 1).setCurrentIndex(7)
                                nextShift = 7 

                            else:                            
                                whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                                whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))

                                nextStartDay = whichTable.cellWidget(row+1, 3).text() 

                                if datetime.datetime.strptime(nextStartDay, '%d.%m.%Y').strftime('%A') != maintenanceShiftOn:
                                    whichTable.cellWidget(row+1, 1).setCurrentIndex(2)                         
                                    nextShift = 2 
                                else:
                                    whichTable.cellWidget(row+1, 1).setCurrentIndex(3)                         
                                    nextShift = 3 


                    case 1:
                        if (row + 1) < whichTable.rowCount(): 

                            if int(float(whichTable.cellWidget(row+1, 11).text() or 24)) <= 12:
                                whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                                whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))

                                nextStartDay = whichTable.cellWidget(row+1, 3).text()                         

                                whichTable.cellWidget(row+1, 1).setCurrentIndex(5)
                                nextShift = 5 

                            else:                      

                                whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                                whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))

                                nextStartDay = whichTable.cellWidget(row+1, 3).text()                        

                                if datetime.datetime.strptime(nextStartDay, '%d.%m.%Y').strftime('%A') != maintenanceShiftOn:
                                    whichTable.cellWidget(row+1, 1).setCurrentIndex(0)                         
                                    nextShift = 0 
                                else:
                                    whichTable.cellWidget(row+1, 1).setCurrentIndex(4)                                                     
                                    nextShift = 4  

                    case 2:
                        if (row + 1) < whichTable.rowCount():   

                            if int(float(whichTable.cellWidget(row+1, 11).text() or 24)) <= 12:
                                whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                                whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))

                                nextStartDay = whichTable.cellWidget(row+1, 3).text()                         

                                whichTable.cellWidget(row+1, 1).setCurrentIndex(6)
                                nextShift = 6 

                            else:                     

                                whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                                whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))

                                nextStartDay = whichTable.cellWidget(row+1, 3).text()                         

                                whichTable.cellWidget(row+1, 1).setCurrentIndex(1)
                                nextShift = 1
                            
                    
                    case 3:
                        if (row + 1) < whichTable.rowCount():                       

                            whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                            whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))

                            nextStartDay = whichTable.cellWidget(row+1, 3).text()
                            
                            whichTable.cellWidget(row+1, 1).setCurrentIndex(2)                         
                            nextShift = 2

                    case 4:
                        if (row + 1) < whichTable.rowCount():                       

                            whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                            whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))

                            nextStartDay = whichTable.cellWidget(row+1, 3).text()
                            
                            whichTable.cellWidget(row+1, 1).setCurrentIndex(0)                         
                            nextShift = 0 

                    case 5:
                        if (row + 1) < whichTable.rowCount():                       

                            whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                            whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))

                            nextStartDay = whichTable.cellWidget(row+1, 3).text()                         

                            whichTable.cellWidget(row+1, 1).setCurrentIndex(1)
                            nextShift = 1

                    case 6:
                        if (row + 1) < whichTable.rowCount():                       

                            whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                            whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))

                            nextStartDay = whichTable.cellWidget(row+1, 3).text()                         

                            whichTable.cellWidget(row+1, 1).setCurrentIndex(2)
                            nextShift = 2

                    case 7:
                        if (row + 1) < whichTable.rowCount():                       

                            whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                            whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))

                            nextStartDay = whichTable.cellWidget(row+1, 3).text()                         

                            whichTable.cellWidget(row+1, 1).setCurrentIndex(0)
                            nextShift = 0
            
            if table == 3:
                if (row + 1) < whichTable.rowCount():                
                    
                    if int(self.dispoNoList[whichTable.cellWidget(row, 6).text()]) <= self.tableBatchesExtruder1.rowCount() and self.tableBatchesExtruder1.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6) and self.tableBatchesExtruder1.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                        otherTable = self.tableBatchesExtruder1
                    
                    elif int(self.dispoNoList[whichTable.cellWidget(row, 6).text()]) <= self.tableBatchesExtruder2.rowCount() and self.tableBatchesExtruder2.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6) and self.tableBatchesExtruder2.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                        otherTable = self.tableBatchesExtruder2
                                                    

                    if otherTable.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 3).date() < whichTable.cellWidget(row, 3).date():
                        whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))
                        whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))
                        nextStartDay = whichTable.cellWidget(row+1, 3).text()
                    else:
                        whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(otherTable.cellWidget(self.dispoNoList[whichTable.cellWidget(row+1, 6).text()], 2).text(), '%d.%m.%Y') + datetime.timedelta(days=1))
                        whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(otherTable.cellWidget(self.dispoNoList[whichTable.cellWidget(row+1, 6).text()], 3).text(), '%d.%m.%Y') + datetime.timedelta(days=1))
                        nextStartDay = whichTable.cellWidget(row+1, 3).text()


            if table == 4:                
                match nextDelivery:
                    case 0:
                        if (row + 1) < whichTable.rowCount():

                            if int(self.dispoNoList[whichTable.cellWidget(row, 6).text()]) <= self.tableBatchesExtruder1.rowCount() and self.tableBatchesExtruder1.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6) and self.tableBatchesExtruder1.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                otherTable = self.tableBatchesExtruder1
                            
                            elif int(self.dispoNoList[whichTable.cellWidget(row, 6).text()]) <= self.tableBatchesExtruder2.rowCount() and self.tableBatchesExtruder2.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6) and self.tableBatchesExtruder2.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                otherTable = self.tableBatchesExtruder2

                            if otherTable.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 3).date() < whichTable.cellWidget(row, 3).date():
                                whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))
                                whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y') + datetime.timedelta(days=1))
                                nextStartDay = whichTable.cellWidget(row+1, 3).text()
                            else:
                                whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(otherTable.cellWidget(self.dispoNoList[whichTable.cellWidget(row+1, 6).text()], 2).text(), '%d.%m.%Y') + datetime.timedelta(days=1))
                                whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(otherTable.cellWidget(self.dispoNoList[whichTable.cellWidget(row+1, 6).text()], 3).text(), '%d.%m.%Y') + datetime.timedelta(days=1))
                                nextStartDay = whichTable.cellWidget(row+1, 3).text()


                            if (row + 2) < whichTable.rowCount():                                
                                nextDelivery = whichTable.cellWidget(row+2, 8).currentIndex()
                            else:
                                nextDelivery = whichTable.cellWidget(row+1, 8).currentIndex() 
                            


                    case 1:    
                        if (row + 1) < whichTable.rowCount():

                            if int(self.dispoNoList[whichTable.cellWidget(row, 6).text()]) <= self.tableBatchesExtruder1.rowCount() and self.tableBatchesExtruder1.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6) and self.tableBatchesExtruder1.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                otherTable = self.tableBatchesExtruder1
                            
                            elif int(self.dispoNoList[whichTable.cellWidget(row, 6).text()]) <= self.tableBatchesExtruder2.rowCount() and self.tableBatchesExtruder2.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6) and self.tableBatchesExtruder2.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 6).text() == whichTable.cellWidget(row, 6).text():
                                otherTable = self.tableBatchesExtruder2

                            if otherTable.cellWidget(self.dispoNoList[whichTable.cellWidget(row, 6).text()], 3).date() < whichTable.cellWidget(row, 3).date():
                                whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))
                                whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(nextStartDay, '%d.%m.%Y'))                            
                                nextStartDay = whichTable.cellWidget(row+1, 3).text()
                            else:
                                whichTable.cellWidget(row+1, 2).setDate(datetime.datetime.strptime(otherTable.cellWidget(self.dispoNoList[whichTable.cellWidget(row+1, 6).text()], 2).text(), '%d.%m.%Y'))
                                whichTable.cellWidget(row+1, 3).setDate(datetime.datetime.strptime(otherTable.cellWidget(self.dispoNoList[whichTable.cellWidget(row+1, 6).text()], 3).text(), '%d.%m.%Y'))
                                nextStartDay = whichTable.cellWidget(row+1, 3).text()


                            if (row + 2) < whichTable.rowCount():                                
                                nextDelivery = whichTable.cellWidget(row+2, 8).currentIndex()
                            else:
                                nextDelivery = whichTable.cellWidget(row+1, 8).currentIndex()    
                                


            if (row + 1) < whichTable.rowCount(): 
                newKW = datetime.datetime.strptime(whichTable.cellWidget(row+1, 3).date().toString('dd.MM.yyyy'), '%d.%m.%Y').strftime('%V')  
                whichTable.cellWidget(row+1, 0).setText(newKW)
                newTimeToDelivery = (datetime.datetime.strptime(whichTable.cellWidget(row+1, 10).date().toString('dd.MM.yyyy'), '%d.%m.%Y') - datetime.datetime.strptime(whichTable.cellWidget(row+1, 3).date().toString('dd.MM.yyyy'), '%d.%m.%Y')).days
                whichTable.cellWidget(row+1, 12).setText(str(newTimeToDelivery))
                
                whichActualLab = 0
                if table == 1 or table == 2:
                    whichActualLab = whichTable.cellWidget(row+1, 9).currentIndex()  
                else:
                    match whichTable.cellWidget(row+1, 9).text():
                        case 'Dichte':
                            whichActualLab = 1 
                        case 'Mechanik':
                            whichActualLab = 2 
                        case 'REACh':
                            whichActualLab = 3
                        
                if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                    whichTable.cellWidget(row+1, 12).setStyleSheet('background-color: red')
                elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                    whichTable.cellWidget(row+1, 12).setStyleSheet('background-color: red') 
                elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                    whichTable.cellWidget(row+1, 12).setStyleSheet('background-color: red') 
                elif newTimeToDelivery < self.timeNormal:            
                    whichTable.cellWidget(row+1, 12).setStyleSheet('background-color: red')       
                else:
                    whichTable.cellWidget(row+1, 12).setStyleSheet('background-color: white') 

                if whichTable.cellWidget(row+1, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):            
                    whichTable.cellWidget(row+1, 3).setStyleSheet('background-color: red')
                else:
                    whichTable.cellWidget(row+1, 3).setStyleSheet('background-color: white')

                
            
        self.workingOnShiftPlan = False               
    
    def enumerateBatches(self, table):
        
        self.saveFile.setEnabled(True)

        if table == 1:            
            whichTable = self.tableBatchesExtruder1
        else:           
            whichTable = self.tableBatchesExtruder2

        if whichTable.rowCount() != 0:            

            rowList = []
            for row in range(whichTable.rowCount()): 

                stringLength = len(whichTable.cellWidget(row, 5).text())
                if stringLength == 9:
                    numberWithoutAtoZ = whichTable.cellWidget(row, 5).text()[:-1]  
                else:
                    numberWithoutAtoZ = whichTable.cellWidget(row, 5).text()                                

                if 0 <= stringLength <= 5:
                    rowList.append(0)
                elif 5 < stringLength <= 6:
                    rowList.append(int(numberWithoutAtoZ[-1:]))                  
                elif 6 < stringLength <= 7:
                    rowList.append(int(numberWithoutAtoZ[-2:]))                  
                else:                
                    rowList.append(int(numberWithoutAtoZ[-3:]))                     

            rowList.sort(reverse=True) 
            
            highestBatchNo = rowList[0] + 1   
                        
                                        
            for row in range(whichTable.rowCount()):
                changeBatchNo = False

                stringLength = len(whichTable.cellWidget(row, 5).text())

                if stringLength <= 5 and highestBatchNo < 10:  
                    newBatchNo = whichTable.cellWidget(row, 5).text() + '00' + str(highestBatchNo)
                    whichTable.cellWidget(row, 5).setText(newBatchNo)                         
                    highestBatchNo = highestBatchNo + 1
                    changeBatchNo = True
                elif stringLength <= 5 and highestBatchNo < 100:
                    newBatchNo = whichTable.cellWidget(row, 5).text() + '0' + str(highestBatchNo)
                    whichTable.cellWidget(row, 5).setText(newBatchNo)                        
                    highestBatchNo = highestBatchNo + 1
                    changeBatchNo = True
                elif stringLength <= 5 and highestBatchNo < 1000:
                    newBatchNo = whichTable.cellWidget(row, 5).text() + str(highestBatchNo)
                    whichTable.cellWidget(row, 5).setText(newBatchNo)                        
                    highestBatchNo = highestBatchNo + 1  
                    changeBatchNo = True

                if whichTable.cellWidget(row, 6).text() in self.checkDispoNoSilo and (table == 1 or table == 2) and changeBatchNo == True:

                    if whichTable.cellWidget(row, 8).currentText() == 'Silo':
                        self.tableBatchesSilo.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 5).setText(newBatchNo)                      
                                        
                    
                    elif whichTable.cellWidget(row, 8).currentText() == 'Homogenisierung':                        
                        if table == 1:
                            newNoHomogenisation = newBatchNo[-6:] + '.1'
                            self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 5).setText(newNoHomogenisation)
                        else:
                            newNoHomogenisation = newBatchNo[-6:] + '.2'
                            self.tableBatchesHomogenisation.cellWidget(self.checkDispoNoSilo[whichTable.cellWidget(row, 6).text()], 5).setText(newNoHomogenisation)                 

    def loadFile(self):
        self.tableBatchesExtruder1.blockSignals(True)
        self.tableBatchesExtruder2.blockSignals(True) 
        
        fileName = QFileDialog.getOpenFileName(self, 'Speichern unter...', self.saveFilePath, 'Excel-Dateien (*.xlsx)' )   
        attrSiloDelivery = ['Silo', 'Dino']        
        if fileName != ([], '') and fileName != ('', ''):           
            
            self.saveFilePath = os.path.join(os.path.dirname(__file__), (fileName[0]))            

            self.config['PATH']['LastSaved'] = self.saveFilePath 
            with open(os.path.join(self.imagePath, 'settings.ini'), 'w') as configfile:
                self.config.write(configfile) 

            shutil.copy(self.saveFilePath, os.path.join(os.path.dirname(__file__), os.path.splitext(os.path.basename(fileName[0]))[0]+'.backup'))

            wb = load_workbook(filename=self.saveFilePath)
            sheets = wb.sheetnames

            sheetsNo = 4 ### modify for homogenisation and silo

            for sheet in range(sheetsNo):                

                ws = wb[sheets[sheet]]                

                match sheet:
                    case 0:            
                        whichTable = self.tableBatchesExtruder1            
                    case 1:           
                        whichTable = self.tableBatchesExtruder2
                    case 2:
                        whichTable = self.tableBatchesHomogenisation
                    case 3:
                        whichTable = self.tableBatchesSilo                                   

                rx = QRegularExpression("SP\\d{1,7}[a-zA-Z]")
                rx2 = QRegularExpression("3\\d{1}\\.\\d{1,4}\\-\\d{1}")
                rx3 = QRegularExpression("1-\\d{1,2}-\\d{1,3}")
                rx4 = QRegularExpression("\\d{1,2}")

                rowPosition = whichTable.rowCount()
                for row in ws.iter_rows(values_only=True):
                    if ((sheet == 0 or sheet == 1) and row[6] not in self.dispoNoList.keys()) or ((sheet == 2 or sheet == 3) and row[6] not in self.checkDispoNoSilo.keys()):

                        whichTable.insertRow(rowPosition)   

                        for rowItem in range(14):                       

                            match rowItem:                            
                                case 0:                    
                                    whichCalendarWeek = QLabel()
                                    whichCalendarWeek.setText(row[rowItem])                                
                                    whichCalendarWeek.setFixedWidth(40)
                                    whichCalendarWeek.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                    whichTable.setCellWidget(rowPosition, rowItem, whichCalendarWeek)                                

                                case 1:                                            
                                    whichShift = QComboBox()
                                    whichShift.addItems(self.attrShift)
                                    whichShift.setCurrentText(row[rowItem])
                                    if sheet == 0:
                                        whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(1))
                                    elif sheet == 1:
                                        whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(2))
                                    elif sheet == 2:
                                        whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(3))
                                    elif sheet == 3:
                                        whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(4))        
                                    whichShift.setProperty('row', rowPosition)                          
                                    whichTable.setCellWidget(rowPosition, rowItem, whichShift)                       

                                case 2:                                  
                                    buttonProductionDate = QDateEdit()
                                    buttonProductionDate.setFixedWidth(80)
                                    buttonProductionDate.setDate(row[rowItem]) 
                                    buttonProductionDate.setProperty('row', rowPosition)                                 
                                    if sheet == 0:
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(1))
                                    elif sheet == 1:
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(2))
                                    elif sheet == 2:
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(3))
                                    elif sheet == 3:
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(4))                                                                          
                                    whichTable.setCellWidget(rowPosition, rowItem, buttonProductionDate) 

                                case 3:                                  
                                    buttonProductionDate = QDateEdit()
                                    buttonProductionDate.setFixedWidth(80)
                                    buttonProductionDate.setDate(row[rowItem])
                                    buttonProductionDate.setProperty('row', rowPosition)
                                    if sheet == 0:
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(1))
                                    elif sheet == 1:
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(2))
                                    elif sheet == 2:
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(3))
                                    elif sheet == 3:
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(4))               
                                    whichTable.setCellWidget(rowPosition, rowItem, buttonProductionDate)

                                    if row[rowItem].strftime('%Y.%m.%d') <= datetime.datetime.now().strftime('%Y.%m.%d'):           
                                        whichTable.cellWidget(rowPosition, rowItem).setStyleSheet('background-color: red')
                                    else:
                                        whichTable.cellWidget(rowPosition, rowItem).setStyleSheet('background-color: white')

                                case 4: 
                                    if sheet == 0 or sheet == 1:                               
                                        whichArticle = QComboBox()                                    
                                        whichArticle.addItems(self.articleNoList)
                                        whichArticle.setValidator(QRegularExpressionValidator(rx2, self))
                                        whichArticle.setCurrentText(row[rowItem])
                                        whichArticle.setEditable(True)
                                        whichArticle.setProperty('row', rowPosition)
                                        if sheet == 0:
                                            whichArticle.currentIndexChanged.connect(lambda: self.articleChanged(1)) 
                                        elif sheet == 1:
                                            whichArticle.currentIndexChanged.connect(lambda: self.articleChanged(2))
                                        for article in range(len(self.articleNoList)):                       
                                            whichIcon = self.colorList[self.dictArticleColors[self.articleNoList[article]]][0]+'.png'               
                                            whichArticle.setItemIcon(article, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))

                                    else:
                                        whichArticle = QComboBox()
                                        whichArticle.addItem(row[rowItem])
                                        whichArticle.setCurrentText(row[rowItem])
                                        whichArticle.setEditable(False)
                                        whichArticle.setProperty('row', rowPosition)                                      
                                        
                                        if whichArticle.currentText() in self.articleNoList:                                                                  

                                            whichIcon = self.colorList[self.dictArticleColors[row[rowItem]]][0]+'.png'               
                                            whichArticle.setItemIcon(0, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))                                
            
                                    whichTable.setCellWidget(rowPosition, rowItem, whichArticle)
                                    
                                case 5:
                                    if sheet == 0 or sheet == 1:                                         
                                        newBatchNo = QPushButton()                                                                                      
                                        if sheet == 0: 
                                            newBatchNo.clicked.connect(lambda: self.changeBatchNo(1))
                                        elif sheet == 1:
                                            newBatchNo.clicked.connect(lambda: self.changeBatchNo(2))
                                        newBatchNo.setProperty('row', rowPosition)
                                    else:
                                        newBatchNo = QLabel()
                                        newBatchNo.setAlignment(Qt.AlignmentFlag.AlignCenter)

                                    newBatchNo.setFixedWidth(100)
                                    newBatchNo.setText(row[rowItem])
                                    whichTable.setCellWidget(rowPosition, rowItem, newBatchNo) 

                                case 6: 
                                    if sheet == 0 or sheet == 1:
                                        newDispo = QPushButton()
                                        if sheet == 0: 
                                            newDispo.clicked.connect(lambda: self.changeDispoNo(1))
                                        elif sheet == 1:
                                            newDispo.clicked.connect(lambda: self.changeDispoNo(2))
                                        newDispo.setProperty('row', rowPosition)
                                        self.dispoNoList[row[rowItem]] = rowPosition

                                    else:
                                        newDispo = QLabel()
                                        newDispo.setFixedWidth(100)
                                        newDispo.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                        self.checkDispoNoSilo[row[rowItem]] = rowPosition
                                    
                                    newDispo.setFixedWidth(100) 
                                    newDispo.setText(row[rowItem])
                                    whichTable.setCellWidget(rowPosition, rowItem, newDispo)      

                                case 7: 
                                    if sheet == 0 or sheet == 1:                               
                                        whichCustomer = QComboBox()
                                        whichCustomer.addItem(' ')
                                        whichCustomer.addItems(self.customerList)                                
                                        whichCustomer.setCurrentText(row[rowItem])                                    
                                        whichCustomer.setEditable(True)
                                        whichCustomer.setProperty('row', rowPosition)
                                        if sheet == 0: 
                                            whichCustomer.currentIndexChanged.connect(lambda: self.customerChanged(1))
                                        elif sheet == 1:
                                            whichCustomer.currentIndexChanged.connect(lambda: self.customerChanged(2))  
                                    else:
                                        whichCustomer = QLabel()
                                        whichCustomer.setText(row[rowItem])
                                    
                                    whichTable.setCellWidget(rowPosition, rowItem, whichCustomer) 

                                case 8:
                                    if sheet == 0 or sheet == 1:                
                                        whichPackaging = QComboBox()
                                        whichPackaging.addItems(self.attrPack)
                                        whichPackaging.setCurrentText(row[rowItem])
                                        whichPackaging.setProperty('row', rowPosition)
                                        if sheet == 0:
                                            whichPackaging.currentIndexChanged.connect(lambda: self.packagingChanged(1))
                                        elif sheet == 1:
                                            whichPackaging.currentIndexChanged.connect(lambda: self.packagingChanged(2))
                                    elif sheet == 3:
                                        whichPackaging = QComboBox()
                                        whichPackaging.addItems(attrSiloDelivery)  
                                        whichPackaging.setCurrentIndex(0)  
                                        whichPackaging.currentIndexChanged.connect(self.enableSaving) 

                                    else:
                                        whichPackaging = QLabel()
                                        whichPackaging.setText('Homogenisierung')
                                        whichPackaging.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                    
                                    whichTable.setCellWidget(rowPosition, rowItem, whichPackaging)      

                                case 9: 
                                    if sheet == 0 or sheet == 1:                
                                        whichLab = QComboBox()
                                        whichLab.addItems(self.attrLab)
                                        whichLab.setCurrentText(row[rowItem])
                                        whichLab.setProperty('row', rowPosition)
                                    else:
                                        whichLab = QLabel()
                                        whichLab.setText(row[rowItem])
                                        whichLab.setAlignment(Qt.AlignmentFlag.AlignCenter)

                                    whichLab.setFixedWidth(80)                                 
                                    if sheet == 0:
                                        whichLab.currentIndexChanged.connect(lambda: self.labChanged(1))
                                    elif sheet == 1:
                                        whichLab.currentIndexChanged.connect(lambda: self.labChanged(2))
                                    
                                    whichTable.setCellWidget(rowPosition, rowItem, whichLab)                           
                                        
                                case 10:                                  
                                    buttonDeliveryDate = QDateEdit()
                                    buttonDeliveryDate.setFixedWidth(80)
                                    buttonDeliveryDate.setDate(row[rowItem])                                  
                                    if sheet == 0:
                                        buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(1))
                                    elif sheet == 1:
                                        buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(2))
                                    elif sheet == 2:
                                        buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(3))
                                        buttonDeliveryDate.setDisabled(True)
                                    elif sheet == 3:
                                        buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(4))
                                        buttonDeliveryDate.setDisabled(True)
                                    buttonDeliveryDate.setProperty('row', rowPosition)            
                                    whichTable.setCellWidget(rowPosition, rowItem, buttonDeliveryDate)

                                case 11:
                                    if sheet == 0 or sheet == 1:    
                                        whichBatchSize = QLineEdit()
                                        whichBatchSize.setText(row[rowItem])
                                        whichBatchSize.setEnabled(True)
                                        whichBatchSize.setValidator(QRegularExpressionValidator(rx4, self)) 
                                        whichBatchSize.setProperty('row', rowPosition)
                                        if sheet == 0:            
                                            whichBatchSize.textChanged.connect(lambda: self.batchSizeChanged(1)) 
                                        elif sheet == 1:
                                            whichBatchSize.textChanged.connect(lambda: self.batchSizeChanged(2))
                                    elif sheet == 2 or sheet == 3:
                                        whichBatchSize = QLabel()
                                    
                                    whichBatchSize.setFixedWidth(38)
                                    whichBatchSize.setText(row[rowItem])                                
                                    whichBatchSize.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                    whichTable.setCellWidget(rowPosition, rowItem, whichBatchSize)

                                case 12:
                                    whichDeliveryDate = QLabel()
                                    whichDeliveryDate.setText(row[rowItem])                                
                                    whichDeliveryDate.setFixedWidth(38)
                                    whichDeliveryDate.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                    whichTable.setCellWidget(rowPosition, rowItem, whichDeliveryDate)  

                                    newTimeToDelivery = int(row[rowItem]) 

                                    whichActualLab = 0
                                    if sheet == 0 or sheet == 1:
                                        whichActualLab = whichTable.cellWidget(rowPosition, 9).currentIndex()  
                                    else:
                                        match whichTable.cellWidget(rowPosition, 9).text():
                                            case 'Dichte':
                                                whichActualLab = 1 
                                            case 'Mechanik':
                                                whichActualLab = 2 
                                            case 'REACh':
                                                whichActualLab = 3 
                            
                                    if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')
                                    elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                                    elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                                    elif newTimeToDelivery < self.timeNormal:            
                                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')       
                                    else:
                                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: white')

                                    if whichTable.cellWidget(rowPosition, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):             
                                        whichTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: red')
                                    else:
                                        whichTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: white')
                                
                                case 13:                                
                                    whichComment = QLineEdit() 
                                    whichComment.setText(row[rowItem])
                                    whichComment.setFixedWidth(350)
                                    whichComment.textChanged.connect(self.enableSaving)
                                    whichTable.setCellWidget(rowPosition, rowItem, whichComment) 
                        rowPosition = rowPosition + 1 

            self.setLoadedFile = True
        
        self.generateAdditiveUsage()
        self.tableBatchesExtruder1.blockSignals(False)
        self.tableBatchesExtruder2.blockSignals(False) 
    
    def performSaveFileAs(self): 
        fileName = QFileDialog.getSaveFileName(self, 'Speichern unter...', self.saveFilePath, 'Excel-Dateien (*.xlsx)' )        
        if fileName != ([], '') and fileName != ('', ''):          
            wb = Workbook() 

            ws = wb.active  
            ws.title = 'Extruder 1'
            wb.create_sheet('Extruder 2')
            wb.create_sheet('Homogenisierung')
            wb.create_sheet('Silo')

            sheets = wb.sheetnames

            sheetsNo = 4 ### modify for homogenisation and silo

            for sheet in range(sheetsNo):

                ws = wb[sheets[sheet]]
                match sheet:
                    case 0:            
                        whichTable = self.tableBatchesExtruder1            
                    case 1:           
                        whichTable = self.tableBatchesExtruder2
                    case 2:
                        whichTable = self.tableBatchesHomogenisation
                    case 3:
                        whichTable = self.tableBatchesSilo 

                saveTableData = []
                
                saveRow = []        
                for row in range(whichTable.rowCount()): 
                    if sheet == 0 or sheet == 1:            
                        saveRow = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentText(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y').date(), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).currentText(), whichTable.cellWidget(row, 8).currentText(), whichTable.cellWidget(row, 9).currentText(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                    elif sheet == 3:
                        saveRow = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentText(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y').date(), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).currentText(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                    else:
                        saveRow = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentText(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y').date(), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).text(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                    saveTableData.append(saveRow)

                for writeRow in saveTableData:
                    ws.append(writeRow)        

            self.saveFilePath = os.path.join(os.path.dirname(__file__), (fileName[0]))      
            
            wb.save(self.saveFilePath)

            self.setLoadedFile = True

            self.config['PATH']['LastSaved'] = self.saveFilePath 
            with open(os.path.join(self.imagePath, 'settings.ini'), 'w') as configfile:
                self.config.write(configfile)
        
    def performSaveFile(self): 
        
        self.saveFile.setEnabled(False)       

        if self.saveFilePath == '' or self.setLoadedFile == False:            
            self.performSaveFileAs()            
        else:
            wb = Workbook() 

            ws = wb.active  
            ws.title = 'Extruder 1'
            wb.create_sheet('Extruder 2')
            wb.create_sheet('Homogenisierung')
            wb.create_sheet('Silo')

            sheets = wb.sheetnames

            sheetsNo = 4 ### modify for homogenisation and silo

            for sheet in range(sheetsNo):

                ws = wb[sheets[sheet]]
                match sheet:
                    case 0:            
                        whichTable = self.tableBatchesExtruder1            
                    case 1:           
                        whichTable = self.tableBatchesExtruder2
                    case 2:
                        whichTable = self.tableBatchesHomogenisation
                    case 3:
                        whichTable = self.tableBatchesSilo 

                saveTableData = []
                
                saveRow = []        
                for row in range(whichTable.rowCount()): 
                    if sheet == 0 or sheet == 1:            
                        saveRow = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentText(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y').date(), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).currentText(), whichTable.cellWidget(row, 8).currentText(), whichTable.cellWidget(row, 9).currentText(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                    elif sheet == 3:
                        saveRow = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentText(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y').date(), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).currentText(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                    else:
                        saveRow = [whichTable.cellWidget(row, 0).text(), whichTable.cellWidget(row, 1).currentText(), datetime.datetime.strptime(whichTable.cellWidget(row, 2).text(), '%d.%m.%Y').date(), datetime.datetime.strptime(whichTable.cellWidget(row, 3).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 4).currentText(), whichTable.cellWidget(row, 5).text(), whichTable.cellWidget(row, 6).text(), whichTable.cellWidget(row, 7).text(), whichTable.cellWidget(row, 8).text(), whichTable.cellWidget(row, 9).text(), datetime.datetime.strptime(whichTable.cellWidget(row, 10).text(), '%d.%m.%Y').date(), whichTable.cellWidget(row, 11).text(), whichTable.cellWidget(row, 12).text(), whichTable.cellWidget(row, 13).text()]
                    
                    saveTableData.append(saveRow)

                for writeRow in saveTableData:
                    ws.append(writeRow)
                
                wb.save(self.saveFilePath)    
    
    def generateSiloLists(self):

        self.saveFile.setEnabled(True)

        extruderList = [self.tableBatchesExtruder1, self.tableBatchesExtruder2 ]

        siloList = [self.tableBatchesSilo, self.tableBatchesHomogenisation ]

        attrSiloDelivery = ['Silo', 'Dino'] 

        for extruder in range(len(extruderList)):
            whichTable = extruderList[extruder] 

            saveHomogenisationTable = []
            saveSiloTable = []            

            if whichTable.rowCount() != 0:
            
                for row in range(whichTable.rowCount()):                     

                    if whichTable.cellWidget(row, 8).currentIndex() == 2:         
                    
                        saveSiloTable.append(row)
                    
                    elif whichTable.cellWidget(row, 8).currentIndex() == 3:            
                    
                        saveHomogenisationTable.append(row)             
                

                saveSiloTable.sort()
                saveHomogenisationTable.sort()               


                for silo in range(len(siloList)):
                    if silo == 0:    
                        whichSiloTable = saveSiloTable 
                    else:
                        whichSiloTable = saveHomogenisationTable                   
                
                for silo in range(len(siloList)):

                    if silo == 0:    
                        whichSiloTable = saveSiloTable 
                    else:
                        whichSiloTable = saveHomogenisationTable  

                    otherTable = siloList[silo]                   

                    for item in whichSiloTable:
                        
                        if whichTable.cellWidget(item, 6).text() not in self.checkDispoNoSilo.keys():
                        
                            rowPosition = otherTable.rowCount()
                            otherTable.insertRow(rowPosition)

                            rowItem = 0
                            for rowItem in range(14):                  

                                match rowItem:
                                    case 0:                    
                                        whichCalendarWeek = QLabel()
                                        whichCalendarWeek.setText(whichTable.cellWidget(item, rowItem).text())
                                        whichCalendarWeek.setFixedWidth(40)
                                        whichCalendarWeek.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                        otherTable.setCellWidget(rowPosition, rowItem, whichCalendarWeek)

                                    case 1:                                            
                                        whichShift = QComboBox()
                                        whichShift.addItems(self.attrShift)
                                        whichShift.setCurrentIndex(0)
                                        if silo == 0:
                                            whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(4))
                                        else:
                                            whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(3))
                                        whichShift.setProperty('row', rowPosition)                          
                                        otherTable.setCellWidget(rowPosition, rowItem, whichShift)                       

                                    case 2:                                  
                                        buttonProductionDate = QDateEdit()
                                        buttonProductionDate.setFixedWidth(80)
                                        buttonProductionDate.setDate(datetime.datetime.strptime(whichTable.cellWidget(item, rowItem+1).text(), '%d.%m.%Y')) 
                                        buttonProductionDate.setProperty('row', rowPosition)
                                        if silo == 0:       
                                            buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(4)) 
                                        else: 
                                            buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(3))                                            
                                        otherTable.setCellWidget(rowPosition, rowItem, buttonProductionDate) 

                                    case 3:                                  
                                        buttonProductionDate = QDateEdit()
                                        buttonProductionDate.setFixedWidth(80)
                                        buttonProductionDate.setDate(datetime.datetime.strptime(whichTable.cellWidget(item, rowItem).text(), '%d.%m.%Y'))
                                        buttonProductionDate.setProperty('row', rowPosition)
                                        if silo == 0:       
                                            buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(4)) 
                                        else: 
                                            buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(3))               
                                        otherTable.setCellWidget(rowPosition, rowItem, buttonProductionDate)                            

                                    case 4:                                
                                        whichArticle = QComboBox()  
                                        whichArticle.addItem(whichTable.cellWidget(item, rowItem).currentText())                                      
                                        whichArticle.setCurrentText(whichTable.cellWidget(item, rowItem).currentText()) 
                                        whichArticle.setProperty('row', rowPosition)                                  
                                        if whichTable.cellWidget(item, rowItem).currentText() in self.articleNoList:                                    
                                            
                                            whichIcon = self.colorList[self.dictArticleColors[whichTable.cellWidget(item, rowItem).currentText()]][0]+'.png'                   
                                            whichArticle.setItemIcon(0, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))
                                        otherTable.setCellWidget(rowPosition, rowItem, whichArticle)
                                        
                                    case 5:                                         
                                        newBatchNo = QLabel()
                                        if silo == 0:
                                            newBatchNo.setText(whichTable.cellWidget(item, rowItem).text()) 
                                        else:
                                            

                                            if extruder == 0:                                
                                                if len(whichTable.cellWidget(item, rowItem).text()) == 9:                                    
                                                    newNoHomogenisation = whichTable.cellWidget(item, rowItem).text()[:-1][-6:] + '.1' + whichTable.cellWidget(item, rowItem).text()[-1]
                                                else:                                    
                                                    newNoHomogenisation = whichTable.cellWidget(item, rowItem).text()[-6:] + '.1'
                                                newBatchNo.setText(str(newNoHomogenisation))
                                            else:                                
                                                if len(whichTable.cellWidget(item, rowItem).text()) == 9:
                                                    newNoHomogenisation = whichTable.cellWidget(item, rowItem).text()[:-1][-6:] + '.2' + whichTable.cellWidget(item, rowItem).text()[-1]
                                                else:
                                                    newNoHomogenisation = whichTable.cellWidget(item, rowItem).text()[-6:] + '.2'
                                        newBatchNo.setFixedWidth(100)
                                        newBatchNo.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                        otherTable.setCellWidget(rowPosition, rowItem, newBatchNo) 

                                    case 6:                                         
                                        newDispo = QLabel()
                                        newDispo.setText(whichTable.cellWidget(item, rowItem).text())
                                        newDispo.setFixedWidth(100)
                                        newDispo.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                        otherTable.setCellWidget(rowPosition, rowItem, newDispo)  
                                        self.checkDispoNoSilo[whichTable.cellWidget(item, rowItem).text()] = rowPosition    

                                    case 7:                                
                                        whichCustomer = QLabel()                                                                        
                                        whichCustomer.setText(whichTable.cellWidget(item, rowItem).currentText())                                     
                                        #whichPackaging.currentIndexChanged.connect(lambda: self.labChanged(1))
                                        otherTable.setCellWidget(rowPosition, rowItem, whichCustomer) 

                                    case 8: 
                                        if silo == 0:
                                            whichPackaging = QComboBox()
                                            whichPackaging.addItems(attrSiloDelivery)  
                                            whichPackaging.setCurrentIndex(0)                                      
                                        else:
                                            whichPackaging = QLabel()
                                            whichPackaging.setText('Homogenisierung')
                                            whichPackaging.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                        #whichPackaging.currentIndexChanged.connect(lambda: self.labChanged(1))
                                        otherTable.setCellWidget(rowPosition, rowItem, whichPackaging)      

                                    case 9:               
                                        whichLab = QLabel()                                        
                                        whichLab.setText(whichTable.cellWidget(item, rowItem).currentText())
                                        whichLab.setProperty('row', rowPosition)
                                        whichLab.setFixedWidth(80)
                                        whichLab.setAlignment(Qt.AlignmentFlag.AlignCenter)                                                              
                                        otherTable.setCellWidget(rowPosition, rowItem, whichLab)                           
                                            
                                    case 10:                                  
                                        buttonDeliveryDate = QDateEdit()
                                        buttonDeliveryDate.setFixedWidth(80)
                                        buttonDeliveryDate.setDate(datetime.datetime.strptime(whichTable.cellWidget(item, rowItem).text(), '%d.%m.%Y')) 
                                        if silo == 0:       
                                            buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(4)) 
                                        else: 
                                            buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(3))   
                                        buttonDeliveryDate.setProperty('row', rowPosition)     
                                        buttonDeliveryDate.setDisabled(True)       
                                        otherTable.setCellWidget(rowPosition, rowItem, buttonDeliveryDate)

                                    case 11:
                                        whichBatchSize = QLabel()
                                        whichBatchSize.setText(whichTable.cellWidget(item, rowItem).text())
                                        whichBatchSize.setEnabled(True)
                                        whichBatchSize.setFixedWidth(38)
                                        whichBatchSize.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                        otherTable.setCellWidget(rowPosition, rowItem, whichBatchSize)

                                    case 12:
                                        whichDeliveryDate = QLabel()
                                        whichDeliveryDate.setText(whichTable.cellWidget(item, rowItem).text())                                        
                                        whichDeliveryDate.setFixedWidth(38)
                                        whichDeliveryDate.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                        otherTable.setCellWidget(rowPosition, rowItem, whichDeliveryDate) 

                                        newTimeToDelivery = int(otherTable.cellWidget(rowPosition, rowItem).text() ) 

                                        whichActualLab = 0
                                        match otherTable.cellWidget(rowPosition, 9).text():
                                            case 'Dichte':
                                                whichActualLab = 1 
                                            case 'Mechanik':
                                                whichActualLab = 2 
                                            case 'REACh':
                                                whichActualLab = 3 

                                        if whichActualLab == 1 and newTimeToDelivery < self.timeDensity:            
                                            otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')
                                        elif whichActualLab == 2 and newTimeToDelivery < self.timeMechanics:            
                                            otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                                        elif whichActualLab == 3 and newTimeToDelivery < self.timeReach:            
                                            otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                                        elif newTimeToDelivery < self.timeNormal:            
                                            otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')       
                                        else:
                                            otherTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: white')    

                                        if otherTable.cellWidget(rowPosition, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):             
                                            otherTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: red')
                                        else:
                                            otherTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: white')                    

                                    
                                    case 13:
                                        whichComment = QLineEdit() 
                                    #    whichComment.setText(whichTable.cellWidget(item, rowItem).text())
                                        whichComment.setFixedWidth(350)
                                        otherTable.setCellWidget(rowPosition, rowItem, whichComment)
                                          
    def generateAdditiveUsage(self):
        self.additiveUsageText.clear()
        tableList = [self.tableBatchesExtruder1, self.tableBatchesExtruder2]
       
        dictAdditives = {}
        dictAdditiveUsage = {}


        for key in self.articleList:        

            self.articleListWork = {}                              
            for keys, value in self.articleList[key][3].items():
                 
                additiveConcentration = ast.literal_eval(value)                    
                            
                self.articleListWork[keys] = float(additiveConcentration[0].replace(',','.'))*int(additiveConcentration[1])   

            dictAdditives[self.articleList[key][1]] = self.articleListWork
                  
                
        for table in range(len(tableList)):
            whichTable = tableList[table] 
            dictWeek = {} 
            dictMonth = {}
   
            checkWeek = 0
            checkMonth = 0
            for row in range(whichTable.rowCount()):
                
                currentWeek = datetime.datetime.strptime(whichTable.cellWidget(row, 3).date().toString('dd.MM.yyyy'), '%d.%m.%Y').strftime('%V')
                currentMonth = datetime.datetime.strptime(whichTable.cellWidget(row, 3).date().toString('dd.MM.yyyy'), '%d.%m.%Y').strftime('%m')                    

                if currentWeek != checkWeek:
                    dictAdditiveUsage = {'Granulat': 0}

                if currentMonth != checkMonth:
                    dictAdditiveUsageMonth = {'Granulat': 0}

                try:                        
                    additiveKeyValue = dictAdditives[whichTable.cellWidget(row, 4).currentText()].items()

                except:
                    additiveKeyValue = {}

                else: 
                    if whichTable.cellWidget(row, 11).text() == '':
                        batchMass = 0
                    else:
                        batchMass = int(whichTable.cellWidget(row, 11).text().replace(',','.'))

                    dictAdditiveUsage['Granulat'] = dictAdditiveUsage['Granulat'] + batchMass 
                    dictAdditiveUsageMonth['Granulat'] = dictAdditiveUsageMonth['Granulat'] + batchMass              
                    
                    for key, value in additiveKeyValue:
                        if key in dictAdditiveUsage and value != 0:
                            dictAdditiveUsage[key] = dictAdditiveUsage[key] + value * int(whichTable.cellWidget(row, 11).text().replace(',','.')) / 100
                            
                        elif key not in dictAdditiveUsage and value != 0:
                            dictAdditiveUsage[key] = value * int(whichTable.cellWidget(row, 11).text().replace(',','.')) / 100
                        
                        if key in dictAdditiveUsageMonth and value != 0:
                            dictAdditiveUsageMonth[key] = dictAdditiveUsageMonth[key] + value * int(whichTable.cellWidget(row, 11).text().replace(',','.')) / 100
                            
                        elif key not in dictAdditiveUsageMonth and value != 0:
                            dictAdditiveUsageMonth[key] = value * int(whichTable.cellWidget(row, 11).text().replace(',','.')) / 100
                            
                    dictWeek[currentWeek] = dictAdditiveUsage
                    dictMonth[currentMonth] = dictAdditiveUsageMonth

                    checkWeek = currentWeek
                    checkMonth = currentMonth

            if table == 0:
                dictWeekEx1 = dictWeek
                dictMonthEx1 = dictMonth
            else:
                dictWeekEx2 = dictWeek
                dictMonthEx2 = dictMonth

        dictWeekBoth ={}
        dictMonthBoth ={}

        for week, additives in dictWeekEx1.items():
            if not week in dictWeekEx2:
                dictWeekBoth[week] = additives
            else:
                additiveHelper = {}
                for additive, mass in additives.items():
                    
                    if not additive in dictWeekEx2[week]:                     
                        additiveHelper[additive] = mass                        
                        
                    else:                                              
                        additiveHelper[additive] = mass + dictWeekEx2[week][additive] #Granulat check

                for additive, mass in dictWeekEx2[week].items():
                    
                    if not additive in dictWeekEx1[week]:                     
                        additiveHelper[additive] = mass                        
                        
                    else:                                              
                        additiveHelper[additive] = mass + dictWeekEx1[week][additive]
                        
                dictWeekBoth[week] = additiveHelper
            
        for week, additives in dictWeekEx2.items():
            if not week in dictWeekEx1:
                dictWeekBoth[week] = additives
        
        dictWeekBothSorted = dict(sorted(dictWeekBoth.items(), key=lambda item: item[0]))

        for month, additives in dictMonthEx1.items():
            if not month in dictMonthEx2:
                dictMonthBoth[month] = additives
            else:
                additiveHelper = {}
                for additive, mass in additives.items():
                    
                    if not additive in dictMonthEx2[month]:                     
                        additiveHelper[additive] = mass                        
                        
                    else:                                              
                        additiveHelper[additive] = mass + dictMonthEx2[month][additive] #Granulat check

                for additive, mass in dictMonthEx2[month].items():
                    
                    if not additive in dictMonthEx1[month]:                     
                        additiveHelper[additive] = mass                        
                        
                    else:                                              
                        additiveHelper[additive] = mass + dictMonthEx1[month][additive]
                        
                dictMonthBoth[month] = additiveHelper
            
        for month, additives in dictMonthEx2.items():
            if not month in dictMonthEx1:
                dictMonthBoth[month] = additives
        
        dictMonthBothSorted = dict(sorted(dictMonthBoth.items(), key=lambda item: item[0]))
     
        for week, additives in dictWeekBothSorted.items():

            self.additiveUsageText.append('KW' + week)
            self.additiveUsageText.append('Geplant sind ' + str(additives['Granulat']) + ' t Granulat mit:')
            sumAdditives = 0
            for additive, mass in additives.items():
                if not additive == 'Granulat':                
                    sumAdditives = sumAdditives + round(mass*(1+self.usageFactor/100), 2)
                    self.additiveUsageText.append('   ' + additive + ': ' + str(round(mass*(1+self.usageFactor/100), 2)).replace('.', ',') + ' t')

            massBales = str(round((additives['Granulat'] - sumAdditives) / 0.72)).replace('.', ',')
            self.additiveUsageText.append('   Ballen: ' + massBales + ' t')

            amountPalletes = str(round(additives['Granulat'] / 1.1)).replace('.', ',')
            self.additiveUsageText.append('   Paletten: ' + amountPalletes + ' Stück')
                  
            self.additiveUsageText.append('\n') 

        self.additiveUsageText.append('_____________________________________________________________________\n')

        for month, additives in dictMonthBothSorted.items():
            
            match str(month):
                case '01':
                    monthName = 'Januar'
                case '02':
                    monthName = 'Februar'
                case '03':
                    monthName = 'März'
                case '04':
                    monthName = 'April'
                case '05':
                    monthName = 'Mai'
                case '06':
                    monthName = 'Juni'
                case '07':
                    monthName = 'Juli'
                case '08':
                    monthName = 'August'
                case '09':
                    monthName = 'September'
                case '10':
                    monthName = 'Oktober'
                case '11':
                    monthName = 'November'
                case '12':
                    monthName = 'Dezember'

            self.additiveUsageText.append(monthName)
            self.additiveUsageText.append('Geplant sind ' + str(additives['Granulat']) + ' t Granulat mit:')
            sumAdditives = 0
            for additive, mass in additives.items():
                if not additive == 'Granulat':                
                    sumAdditives = sumAdditives + round(mass*(1+self.usageFactor/100), 2)
                    self.additiveUsageText.append('   ' + additive + ': ' + str(round(mass*(1+self.usageFactor/100), 2)).replace('.', ',') + ' t')

            massBales = str(round((additives['Granulat'] - sumAdditives) / 0.72)).replace('.', ',')
            self.additiveUsageText.append('   Ballen: ' + massBales + ' t')

            amountPalletes = str(round(additives['Granulat'] / 1.1)).replace('.', ',')
            self.additiveUsageText.append('   Paletten: ' + amountPalletes + ' Stück')
                  
            self.additiveUsageText.append('\n')  
            
    def handlePrint(self, table): 
        self.printTable = table

        dialog = QPrintPreviewDialog()
        dialog.paintRequested.connect(self.handlePaintRequest)
        dialog.exec()

    def handlePaintRequest(self, printer): 
        match self.printTable:
            case 1:
                whichTable = self.tableBatchesHomogenisation
                headline = 'Homogenisierungsliste'
            case 2:
                whichTable = self.tableBatchesSilo
                headline = 'Silo-Liste'


        document = QTextDocument()
        cursor = QTextCursor(document)
        printer.setPageOrientation(QPageLayout.Orientation.Landscape) 

        headlineFormat =  QTextCharFormat()
        headlineFormat.setFontWeight(800)
        headlineFormat.setFontUnderline(True)
        headlineFormat.setBaselineOffset(100)
        headlineFormat.setFontFamily('Arial')        
        cursor.insertText(headline, headlineFormat)
        
        tableFormat = QTextTableFormat()        
        tableFormat.setBorder(0.5)
        tableFormat.setCellSpacing(0)
        tableFormat.setTopMargin(0)
        tableFormat.setCellPadding(4)

        tableColorFormat = QTextCharFormat() 

        tableTextFormat = QTextCharFormat()
        tableTextFormat.setFontFamily('Arial')  

        tableColorTextFormat = QTextCharFormat()
        tableColorTextFormat.setFontFamily('Arial') 
        
        table = cursor.insertTable( whichTable.rowCount() + 1, whichTable.columnCount(), tableFormat)
        header = ['KW', 'Schichten', 'Beginn', 'Ende', 'Artikel-Nr.', 'Chargen-Nr.', 'Dispo.-Nr.', 'Kunde', 'Zusatz', 'Labor', 'Abholung', 't', 'Vorlauf', 'Bemerkung']
        for item in header:
            cursor.insertText(item, tableTextFormat)
            cursor.movePosition(QTextCursor.MoveOperation.NextCell) 

        for row in range(whichTable.rowCount()):            
            for col in range(table.columns()):
                match col:
                        case 0: 
                            cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat)
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)                                

                        case 1:  
                            cursor.insertText(whichTable.cellWidget(row, col).currentText(), tableTextFormat)   
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)                                       
                                                   
                        case 2:
                            cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat) 
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)                                 
                        
                        case 3: 
                            cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat)
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)                                
                                                       
                        case 4: 
                            if whichTable.cellWidget(row, col).currentText() in self.articleNoList:                                    
                                            
                                whichColor = self.colorList[self.dictArticleColors[whichTable.cellWidget(row, col).currentText()]][0]
                                match whichColor:
                                    case 'blau':
                                        tableColorTextFormat.setForeground(QColor('#000000'))                                            
                                        tableColorFormat.setBackground(QColor('#096bff'))
                                    case 'gelb':
                                        tableColorTextFormat.setForeground(QColor('#000000'))                                            
                                        tableColorFormat.setBackground(QColor('#fcff09'))
                                    case 'terrakotta':
                                        tableColorTextFormat.setForeground(QColor('#000000'))                                            
                                        tableColorFormat.setBackground(QColor('#db5514'))
                                    case 'taupe':
                                        tableColorTextFormat.setForeground(QColor('#000000'))                                            
                                        tableColorFormat.setBackground(QColor('#bb9624'))
                                    case 'schwarz':
                                        tableColorTextFormat.setForeground(QColor('#FFFFFF'))                                            
                                        tableColorFormat.setBackground(QColor('#000000'))
                                    case 'rot':
                                        tableColorTextFormat.setForeground(QColor('#000000'))                                            
                                        tableColorFormat.setBackground(QColor('#f238de'))
                                    case 'gruen':
                                        tableColorTextFormat.setForeground(QColor('#000000'))                                            
                                        tableColorFormat.setBackground(QColor('#1bb726'))
                                    case 'grau':
                                        tableColorTextFormat.setForeground(QColor('#000000'))                                            
                                        tableColorFormat.setBackground(QColor('#818181'))
                                        

                            cursor.insertText(whichTable.cellWidget(row, col).currentText(), tableColorTextFormat) 
                            table.cellAt(row+1, col).setFormat(tableColorFormat)
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)                              
                                                      
                        case 5:  
                            cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat)
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)                                       

                        case 6: 
                            cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat)  
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)    

                        case 7:                             
                            cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat)                                
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)

                        case 8:  
                            match self.printTable:
                                case 1:
                                    cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat)
                                case 2:
                                    cursor.insertText(whichTable.cellWidget(row, col).currentText(), tableTextFormat)
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)       

                        case 9:                            
                            cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat)
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)                            
                                
                        case 10: 
                            cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat)
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)

                        case 11:
                            cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat)
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)

                        case 12:
                            cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat)
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)   
                        case 13:
                            cursor.insertText(whichTable.cellWidget(row, col).text(), tableTextFormat)
                            cursor.movePosition(QTextCursor.MoveOperation.NextCell)                    
                
            
        document.print(printer)

    def importExcelFile(self):
        fileName = QFileDialog.getOpenFileName(self, 'Importieren...', self.importFilePath, 'Excel-Dateien (*.xlsx)' )   
                
        if fileName != ([], '') and fileName != ('', ''):           
            
            self.importFilePath = os.path.join(os.path.dirname(__file__), (fileName[0]))            

            self.config['PATH']['lastimport'] = self.importFilePath 
            with open(os.path.join(self.imagePath, 'settings.ini'), 'w') as configfile:
                self.config.write(configfile) 


            wb = load_workbook(filename=self.importFilePath)
            sheets = wb.sheetnames

            sheetsNo = 1 ### modify for homogenisation and silo

            for sheet in range(sheetsNo):                

                ws = wb[sheets[sheet]]             

                rx = QRegularExpression("SP\\d{1,7}[a-zA-Z]")
                rx2 = QRegularExpression("3\\d{1}\\.\\d{1,4}\\-\\d{1}")
                rx3 = QRegularExpression("1-\\d{1,2}-\\d{1,3}")
                rx4 = QRegularExpression("\\d{1,2}")
                

                for row in ws.iter_rows(values_only=True):                    
                    if row[3] in self.articleNoList and row[2] not in self.dispoNoList.keys():

                        if self.articleExtruderList[row[3]] == '1':                            
                            whichTable = self.tableBatchesExtruder2
                            extruderNo = '2'
                            table = 2
                        else:                            
                            whichTable = self.tableBatchesExtruder1
                            extruderNo = '1'
                            table = 1

                        rowPosition = whichTable.rowCount()
                        whichTable.insertRow(rowPosition)

                        self.saveFile.setEnabled(True)
                        
                        if row[11] == None:
                            deliveryDate = datetime.datetime.strptime(row[10], '%d.%m.%Y')   
                        else:
                            deliveryDate = datetime.datetime.strptime(row[11], '%d.%m.%Y')   
                                   
                        
                        productionDate = deliveryDate - datetime.timedelta(days=self.timeNormal)  
                            

                        timeToDelivery = (deliveryDate - productionDate).days
                    
                        year =  row[10][-2:]              
                        
                        calendarWeek = productionDate.strftime('%V')

                        for item in range(14):
                            match item:
                                case 0:                    
                                    whichCalendarWeek = QLabel()
                                    whichCalendarWeek.setText(calendarWeek)
                                    whichCalendarWeek.setFixedWidth(40)
                                    whichCalendarWeek.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                    whichTable.setCellWidget(rowPosition, item, whichCalendarWeek)

                                case 1:                                    
                                    whichShift = QComboBox()
                                    whichShift.addItems(self.attrShift)
                                    whichShift.setProperty('row', rowPosition)
                                    if table == 1:                                        
                                        whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(1))
                                    else:
                                        whichShift.currentIndexChanged.connect(lambda: self.shiftChanged(2))  
                                    whichTable.setCellWidget(rowPosition, item, whichShift)

                                case 2:
                                    buttonProductionDate = QDateEdit()
                                    buttonProductionDate.setFixedWidth(80)
                                    buttonProductionDate.setDate(productionDate) 
                                    buttonProductionDate.setProperty('row', rowPosition)
                                    if table == 1:       
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(1))
                                    else:
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionStartDateChangedInTable(2))  
                                    whichTable.setCellWidget(rowPosition, item, buttonProductionDate) 

                                case 3:
                                    buttonProductionDate = QDateEdit()
                                    buttonProductionDate.setFixedWidth(80)
                                    buttonProductionDate.setDate(productionDate)
                                    buttonProductionDate.setProperty('row', rowPosition) 
                                    if table == 1:      
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(1))
                                    else:
                                        buttonProductionDate.dateChanged.connect(lambda: self.productionEndDateChangedInTable(2))
                                    whichTable.setCellWidget(rowPosition, item, buttonProductionDate)                    

                                case 4:                                                    
                                    whichArticle = QComboBox()                                    
                                    whichArticle.addItems(self.articleNoList)                    
                                    whichArticle.setCurrentText(row[3])
                                    whichArticle.setEditable(True)
                                    whichArticle.setValidator(QRegularExpressionValidator(rx2, self))                                    
                                    whichArticle.setProperty('row', rowPosition) 
                                    if table == 1:      
                                        whichArticle.currentIndexChanged.connect(lambda: self.articleChanged(1))
                                    else:
                                        whichArticle.currentIndexChanged.connect(lambda: self.articleChanged(2))
                                    for article in range(len(self.articleNoList)):                       
                                        whichIcon = self.colorList[self.dictArticleColors[self.articleNoList[article]]][0]+'.png'               
                                        whichArticle.setItemIcon(article, QIcon(QIcon(os.path.join(self.imagePath, 'assets', whichIcon))))                                   
                                    whichTable.setCellWidget(rowPosition, item, whichArticle)  

                                case 5:                                         
                                    newBatchNo = QPushButton()
                                    newBatchNo.setText(extruderNo + '-'+year+'-')                                    
                                    newBatchNo.setFixedWidth(100)                                                          
                                    if table == 1: 
                                        newBatchNo.clicked.connect(lambda: self.changeBatchNo(1))
                                    else:
                                        newBatchNo.clicked.connect(lambda: self.changeBatchNo(2))
                                    newBatchNo.setProperty('row', rowPosition)
                                    whichTable.setCellWidget(rowPosition, item, newBatchNo) 

                                case 6:
                                    newDispo = QPushButton()
                                    newDispo.setText(row[2])
                                    newDispo.setFixedWidth(100)
                                    if table == 1:
                                        newDispo.clicked.connect(lambda: self.changeDispoNo(1))
                                    else:
                                        newDispo.clicked.connect(lambda: self.changeDispoNo(2))
                                    newDispo.setProperty('row', rowPosition) 
                                    whichTable.setCellWidget(rowPosition, item, newDispo)
                                    self.dispoNoList[row[2]] = rowPosition  

                                case 7:                                
                                    whichCustomer = QComboBox()
                                    whichCustomer.addItem(' ')
                                    whichCustomer.addItems(self.customerList)
                                    whichCustomer.setCurrentText(row[14][:-3])
                                    whichCustomer.setEditable(True)
                                    whichCustomer.setProperty('row', rowPosition)
                                    if table == 1: 
                                        whichCustomer.currentIndexChanged.connect(lambda: self.customerChanged(1))
                                    else:
                                        whichCustomer.currentIndexChanged.connect(lambda: self.customerChanged(2))  
                                    whichTable.setCellWidget(rowPosition, item, whichCustomer)   

                                case 8:                                                                 
                                    whichPackaging = QComboBox()
                                    whichPackaging.addItems(self.attrPack)
                                    whichPackaging.setProperty('row', rowPosition)
                                    if row[24] == 'Silo':
                                        whichPackaging.setCurrentIndex(2)
                                    elif row[13] == 'HOMO':                                        
                                        whichPackaging.setCurrentIndex(3)  
                                    else:
                                        whichPackaging.setCurrentIndex(0)                                     
                                    if table == 1:
                                        whichPackaging.currentIndexChanged.connect(lambda: self.packagingChanged(1))
                                    else:
                                        whichPackaging.currentIndexChanged.connect(lambda: self.packagingChanged(2))
                                    whichTable.setCellWidget(rowPosition, item, whichPackaging)      

                                case 9:                                    
                                    whichLab = QComboBox()
                                    whichLab.addItems(self.attrLab)
                                    whichLab.setCurrentIndex(0) 
                                    whichLab.setProperty('row', rowPosition)
                                    whichLab.setFixedWidth(80)
                                    if table == 1:
                                        whichLab.currentIndexChanged.connect(lambda: self.labChanged(1))
                                    else:
                                        whichLab.currentIndexChanged.connect(lambda: self.labChanged(2))
                                    whichTable.setCellWidget(rowPosition, item, whichLab)             

                                case 10:
                                    buttonDeliveryDate = QDateEdit()
                                    buttonDeliveryDate.setFixedWidth(80)
                                    buttonDeliveryDate.setDate(deliveryDate)     
                                    buttonDeliveryDate.setProperty('row', rowPosition)  
                                    if table == 1:            
                                        buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(1))  
                                    else:  
                                        buttonDeliveryDate.dateChanged.connect(lambda: self.deliveryDateChangedInTable(2))           
                                    whichTable.setCellWidget(rowPosition, item, buttonDeliveryDate)  
                            
                        
                                case 11:
                                    whichBatchSize = QLineEdit()                                    
                                    batchSize = str(round(row[5]))
                                    whichBatchSize.setText(batchSize)
                                    whichBatchSize.setEnabled(True)
                                    whichBatchSize.setFixedWidth(38)
                                    whichBatchSize.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                    whichBatchSize.setValidator(QRegularExpressionValidator(rx4, self)) 
                                    whichBatchSize.setProperty('row', rowPosition)
                                    if table == 1:            
                                        whichBatchSize.textChanged.connect(lambda: self.batchSizeChanged(1)) 
                                    else:
                                        whichBatchSize.textChanged.connect(lambda: self.batchSizeChanged(2))
                                    whichTable.setCellWidget(rowPosition, item, whichBatchSize)

                                case 12:
                                    whichDeliveryDate = QLabel()
                                    whichDeliveryDate.setText(str(timeToDelivery))                    
                                    whichDeliveryDate.setFixedWidth(38)
                                    whichDeliveryDate.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                    whichTable.setCellWidget(rowPosition, item, whichDeliveryDate)

                                    newTimeToDelivery = int(whichTable.cellWidget(rowPosition, 12).text())

                                    if whichTable.cellWidget(rowPosition, 9).currentIndex() == 1 and newTimeToDelivery < self.timeDensity:            
                                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')
                                    elif whichTable.cellWidget(rowPosition, 9).currentIndex() == 2 and newTimeToDelivery < self.timeMechanics:            
                                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                                    elif whichTable.cellWidget(rowPosition, 9).currentIndex() == 3 and newTimeToDelivery < self.timeReach:            
                                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red') 
                                    elif newTimeToDelivery < self.timeNormal:            
                                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: red')       
                                    else:
                                        whichTable.cellWidget(rowPosition, 12).setStyleSheet('background-color: white')

                                    if whichTable.cellWidget(rowPosition, 3).date().toString('yyyy.MM.dd') <= datetime.datetime.now().strftime('%Y.%m.%d'):             
                                        whichTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: red')
                                    else:
                                        whichTable.cellWidget(rowPosition, 3).setStyleSheet('background-color: white')            
                                case 13:
                                    whichComment = QLineEdit()
                                    whichComment.setText(row[13])
                                    whichComment.setFixedWidth(350)
                                    whichComment.textChanged.connect(self.enableSaving)
                                    whichTable.setCellWidget(rowPosition, item, whichComment)  
    
                        rowPosition = rowPosition + 1 

            self.setLoadedFile = True
        
        self.generateSiloLists()
        self.generateAdditiveUsage()        

def main():               
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(os.path.join(os.path.dirname(__file__), 'assets', 'book-solid.svg')))
    window = MainWindow()
    window.showMaximized()
    app.processEvents()
    app.exec()

if __name__ == "__main__":
    main()